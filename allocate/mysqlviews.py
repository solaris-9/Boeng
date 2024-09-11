from django.shortcuts import render,redirect

# Create your views here.

from datetime import datetime, timedelta
import simplejson
import pymysql
from decimal import Decimal

# import json
# import query_string
import os
import xlrd
# import difflib
import shutil
from openpyxl import Workbook
from openpyxl.styles import Alignment
import smtplib 
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.http import HttpResponse
from pymysql.converters import escape_string
from django.conf import settings
from utils import analyzer_db

File_address = r'C:/ResLog/printlog8.txt'
fa = open(File_address,'a')
Delete_log = r'C:/ResLog/delete.txt'
fb = open(Delete_log,'a')
Modify_log = r'C:/ResLog/modify.txt'
fc = open(Modify_log,'a')
Test_log = r'C:/ResLog/test.txt'
fd = open(Test_log,'a')

def GetJSONList(lTitle, lIssue):
    lResult = []
    for row in lIssue:
        dIssue = {}
    for t, c in zip(lTitle, row):
        dIssue[t] = c
        lResult.append(dIssue)
    return lResult


def List2String(lList):
    sResult = "("
    for i in lList:
        m = escape_string(i)
        m = "'" + m + "'"
        sResult = sResult + m + "," 
    sResult = sResult[:-1] + ")"
    return sResult

def repspecial(string):
    sString = string.strip().replace('\u200b','')
    return sString

def strnum(strN):
    
    strN += 1
    if strN < 10: 
        tr="000000" + str(strN)
    elif strN >9 and strN < 100 :
        tr="00000" + str(strN)
    elif strN > 99 and strN < 1000 :
        tr="0000" + str(strN)
    elif strN > 999 and strN < 10000 :
        tr="000" + str(strN)
    elif strN > 9999 and strN <100000 :
        tr="00" + str(strN)
    elif strN > 99999 and strN <1000000 :
        tr="0" + str(strN)
    else :
        tr= str(strN)                    
    sNCID = "T" + tr     
    return sNCID

def strNum(strN, prefix):
    
    strN += 1
    if strN < 10: 
        tr="000000" + str(strN)
    elif strN >9 and strN < 100 :
        tr="00000" + str(strN)
    elif strN > 99 and strN < 1000 :
        tr="0000" + str(strN)
    elif strN > 999 and strN < 10000 :
        tr="000" + str(strN)
    elif strN > 9999 and strN <100000 :
        tr="00" + str(strN)
    elif strN > 99999 and strN <1000000 :
        tr="0" + str(strN)
    else :
        tr= str(strN)                    
    sID = prefix + tr     
    return sID


def tbl_index(tblname,SQLConn):              
    
    sql = "select count(ID) as num from %s " % tblname
    SQLConn.cur.execute(sql)      
    SQLResult = SQLConn.cur.fetchall()    
    count = SQLResult[0][0]
    
    if count > 0 :    
        sql="SELECT ID FROM %s ORDER BY ID"  % tblname
        SQLConn.cur.execute(sql)
        SQLConn.conn.commit() 
        last_result = [x[0] for x in SQLConn.cur.fetchall()][-1]            
        ST = last_result[-7:]
        strN = int(ST)
    else:
        strN = 0
    return strN


def tbl_filtered_index(tblname, tblType, SQLConn):              
    
    SQLCur2 = SQLConn.cursor()
    sql = "SELECT count(ID) as num from %s WHERE Type = '%s' " % (tblname, tblType)
    SQLCur2.execute(sql)      
    SQLResult = SQLCur2.fetchall()    
    count = SQLResult[0][0]
    
    if count > 0 :    
        sql="SELECT ID from %s WHERE Type = '%s' ORDER BY ID " % (tblname, tblType)
        SQLCur2.execute(sql)
        last_result = [x[0] for x in SQLCur2.fetchall()][-1]            
        ST = last_result[-7:]
        strN = int(ST)
    else:
        strN = 0
    return strN

def prjnum(string):
    prjnumlist = string.split(',')
    prjno =  string.strip()[:4]           
    for r in prjnumlist:           
        if 'BBDPIPL' in r:
            prjno = r.strip()
            break
        elif len(r.strip()) == 4 or (len(r.strip()) == 5 and r[:1]=='W'):
            prjno = r.strip()    
    return prjno

def prjdes(prjno):    
    prjds = ''
    if prjno == 'G001':
       prjds = 'Customer  maintenance feature'           
    elif prjno == 'G002':
       prjds = 'IOP'
    elif prjno == 'G003':
       prjds = 'Preconfig'   
    elif prjno == 'G005':
       prjds = 'R&D internal improvement'
    elif prjno == 'G006':
       prjds = 'China Central Bid legacy proj'
    elif prjno == 'G007':
       prjds = 'FWA legacy projects'  
    elif prjno == 'G008':
       prjds = 'TMO project' 
    elif prjno == 'G009':
       prjds = 'HW maintenance'
    
    elif prjno == 'P011':
       prjds = 'Telecom Argentina G-240W-J'
    elif prjno == 'P012':
       prjds = 'KDDI 212MHz (CDE)'   
    elif prjno == 'P048':
       prjds = 'StarHub 6 Beacon and cloud integration'
    elif prjno == 'P060':
       prjds = 'HyperOptic HA-140W-B'
    elif prjno == 'P112':
       prjds = 'Service cBBD-ONTroller for Hyperoptic 7750 and SoftGRE'   
    elif prjno == 'P133':
       prjds = 'VZ NGPON2 POC'    
    elif prjno == 'P138':
       prjds = 'Hotwire XS-221X-A (CDE)'
    elif prjno == 'P151':
       prjds = 'Telenor Sweden - FON Telenor contract transfer to Nokia'
    elif prjno == 'P166':
       prjds = 'G-2425G-A_India_Variant_CI'
    elif prjno == 'P182':
       prjds = 'G2426G-B XS2426G-B Project File'   
    elif prjno == 'P189':
       prjds = 'Project File KDDI 10G EPON ONU'
    elif prjno == 'P190':
       prjds = 'Oi RFQ'
    elif prjno == 'P198':
       prjds = 'Project File PT Telkom Premium ONT G-2425G-B v1.1'
    elif prjno == 'P204':
       prjds = '5G Receiver 5G14-B window mount'
    elif prjno == 'P217':
       prjds = 'DELTA Netherlands RFQ 14092020'
    elif prjno == 'P220':
       prjds = 'Google Custom WiFi Router RFP'
    elif prjno == 'P231':
       prjds = 'Project File Optus 5G GW3'
    elif prjno == 'P238':
       prjds = 'Globe G-2425G-A'
    elif prjno == 'P244':
       prjds = 'Turkcell-Features'
    elif prjno == 'P246':
       prjds = 'Project File TPG 5G project' 
    elif prjno == 'P247':
        prjds = 'Project File TMX G-1425G-A'
    elif prjno == 'P254':
       prjds = 'VF-opcos-security-requirements-G-2425G-A'
    elif prjno == 'P258':
       prjds = 'FastMile 5G FWA Gateways (Mediatek T750)'
    elif prjno == 'P260':
       prjds = 'Bharti_Third_Party_App_Integration_G-2425G-A'
    elif prjno == 'P265':
       prjds = 'Project File TLAR G-242xG-B'
    elif prjno == 'P274':
       prjds = 'TR-369 porting to ONTs'
    elif prjno == 'P266':
       prjds = 'WorldLink B1.1'
    elif prjno == 'P284':
       prjds = 'Converge B1.1'
    elif prjno == 'P293':
       prjds = 'Project File Japan 5G GW3.2 SKU3ST20210505'
    elif prjno == 'P288':
       prjds = 'XS-2426X-A NAR'
    elif prjno == 'P297':
       prjds = 'Telecom Egypt - G-1425G-A G-2425G-A'       
    elif prjno == 'P303':
       prjds = 'Megacable G-1425G-A (512MB) and G-2425G-B customer specific'
    elif prjno == 'P305':
       prjds = 'Viettel 2021RFP G-2425G-A G-140W-G'
    elif prjno == 'P309':
       prjds = 'XGS SFU Cortina JDM XS-010X-C'
    elif prjno == 'P315':
       prjds = 'STC-2021RFQ-WiFi5-Simple-n-WiFi6-ONTs'
    elif prjno == 'P317':
       prjds = 'BBF247 certification for XS-2426G-A B'
    elif prjno == 'P320':
       prjds = 'BGW320 cost and supply assurance'
    elif prjno == 'P321':
       prjds = 'BGW320 changes required by AT&T'
    elif prjno == 'P324':
       prjds = 'ONT-embedded SD-WAN'
    elif prjno == 'P327':
       prjds = 'NWCC_DeviceConfigManagement' 
    elif prjno == 'P328':
       prjds = 'TTT Ukraine Beacon 1.1 Beacon 2'
    elif prjno == 'P326':
       prjds = 'SingTel G-240G-E CI3'
    elif prjno == 'P330':
       prjds = 'Chorus_Operational_Issues'
    elif prjno == 'P333':
       prjds = 'Sunrise Switzerland 5G GW 3.2'
    elif prjno == 'P335':
       prjds = 'Windstream Beacon 6'
    elif prjno == 'P344':
       prjds = 'OLT-G ME'
    elif prjno == 'P346':
       prjds = 'TMX G-1425G-A 512MB variant'    
    elif prjno == 'P350':
       prjds = 'G-1425G-A_Chorus Variant'
    elif prjno == 'P352':
       prjds = 'XS-230X-A'
    elif prjno == 'P355':
       prjds = 'Project File NBN WNTDv4'
    elif prjno == 'P359':
       prjds = 'Project File Oi Brazil G-x425G-A requirements'
    elif prjno == 'P370':
       prjds = 'SLT-G-1425G-AandG-2425G-A'
    elif prjno == 'P373':
       prjds = 'Go Malta G-120G-E'
    elif prjno == 'P376':
       prjds = 'BGW320 BCM technology change and Wi-Fi 6E'
    elif prjno == 'P378':
       prjds = 'Project File G-1425G Transtel Colombia'
    elif prjno == 'P384':
       prjds = 'FastMile 5G FWA Gateway 2 (Mediatek T750 based Value Tier)'
    elif prjno == 'P390':
       prjds = 'Verizon G-211G-A Project File'    
    elif prjno == 'W001':
       prjds = 'WPA3'
    elif prjno == 'W004':
       prjds = 'VPN Private Browsing'
    elif prjno == 'W0046':
       prjds = 'NWCC Bare Metal mode'   
    elif prjno == 'W005':
       prjds = 'MAPP Pureview'
    elif prjno == 'W007':
       prjds = 'Prioritization'  
    elif prjno == 'W0071':
       prjds = 'Recommendation Engine'      
    
    return prjds

def pool(tech,bl):
    pool =''
    Businessline= bl
    techlist1 =['NWCC', 'Nokia Cloud Services', 'Nokia WiFi Cloud']
    techlist2 = ['WebGUI', 'RGW', 'Device Security','Wifi']
    techlist3 = ['FWA_PS','FWA4G','RGW(FWA)']
    techlist4 = ['WebGUI', 'RGW', 'Device Security']
    if tech in techlist1:
        pool = 'Cloud'
    elif tech =='Mobile App':
        pool = 'Mobile App' 
    elif tech =='Mesh Middleware':
        pool = 'Mesh' 
    elif tech =='Beacon':
        pool = 'Beacon' 
    elif tech in techlist3:
        pool = 'FWA_PS' 
    elif tech =='ONT_PS':
        pool = 'ONT_PS' 
    elif tech =='China Product':
        pool = 'China Product' 
    elif tech =='Voice':
        pool = 'Voice' 
    elif tech =='Container Framework':
        pool = 'HomeOS'     
    elif tech =='Container Application':
        pool = 'Container App' 
    elif bl == 'BBD-NWF' and tech in techlist2:
        pool = 'Beacon'
    elif bl == 'BBD-ONT' and tech in techlist2:
        pool = 'ONT_PS'
    elif bl == 'BBD-FWA' and tech in techlist2:
        pool = 'FWA_PS' 
    elif bl == 'BBD-CSW' and tech in techlist4:
        pool = 'HomeOS'
    if pool == 'ONT_PS' or pool == 'China Product' or pool == 'Voice' :
        Businessline = 'BBD-ONT'
    elif pool == 'Cloud' or pool == 'Mobile App' or pool == 'Mesh' or pool == 'Container App' or pool == 'HomeOS':
        Businessline = 'BBD-CSW'
    elif pool == 'Beacon':
        Businessline = 'BBD-NWF'
    elif pool == 'FWA_PS':
        Businessline = 'BBD-FWA'         
    return pool,Businessline

def fd_type(activities):
    ftype =''
    atlist1 =['PIPL (undefined RCR)', 'FD w/o RCR w/o PIPL', 'HW Common']
    atlist2 = ['Architecture&Design', 'HW Compliance', 'CICD FDT', 'SW Build Manager','Security/FOSS Support','FDT Proj. Management','PT&CFT Test','BBDR Proj. Management','NPI/L3 CustomerSupport','Maintenance','Process Improvement','Competence Management','R&D Operation','Subco Removal','Training/LongTermLeave','Component SW support to FT']
    
    if activities in atlist1:
        ftype = 'FD'
    elif activities in atlist2:
        ftype = 'Fixed' 
    elif activities[:7] == 'BBDPROD': 
        ftype = 'FD'
    return ftype

def phase_type(activities):
    phtype =''    
    atlist1 = ['Architecture&Design']
    atlist2 = ['HW Compliance','PT&CFT Test']
    atlist3 = ['SW Build Manager','Security/FOSS Support','FDT Proj. Management','BBDR Proj. Management','Process Improvement','Competence Management','R&D Operation']
    atlist4 = ['NPI/L3 CustomerSupport','Maintenance']
    atlist5 = ['Training/LongTermLeave']
    atlist6 = ['PIPL (undefined RCR)', 'FD w/o RCR w/o PIPL', 'HW Common', 'CICD FDT','Subco Removal']
    if activities in atlist1:
        phtype = 'Systems'
    elif activities in atlist2:
        phtype = 'Test' 
    elif activities in atlist3:
        phtype = 'Support' 
    elif activities in atlist4:
        phtype = 'Maintenance' 
    elif activities in atlist5:
        phtype = 'Training/Leave' 
    elif activities[:7] == 'BBDPROD' or activities in atlist6:
        phtype = 'Development' 
    return phtype

def history(sID,sYID,sMail):
    pass


def release_list(request):
    try:
        sType = request.GET['type']
        DueDate = datetime.today() + timedelta(weeks=8)
        sYear = str(DueDate)[2:4]
    except:
        return HttpResponse('Invalid Parameters', content_type='application/json')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
    cmd = """
          SELECT 
            FixVersions 
          FROM 
            jira_issues_rcr 
          WHERE 
               Left(FixVersions,5) ='BBDR2' 
              OR Left(FixVersions,5) ='BBDR3' 
             
         ORDER BY FixVersions DESC
         """
    SQLBConn = pymysql.connect(host  = settings.BBD_DB['host'],
                            port     = settings.BBD_DB['port'],
                            user     = settings.BBD_DB['username'],
                            password = settings.BBD_DB['password'],
                            database = settings.BBD_DB['name'],
                            charset  = settings.BBD_DB['charset']
                        )
    SQLBCur = SQLBConn.cursor()
    SQLBCur.execute(cmd)                
    SQLBResult = SQLBCur.fetchall()
    SQLBConn.close()
    rellist =[]
    for row in SQLBResult:        
        relv = row[0].split(',')        
        for r in relv:           
            if int(r[4:6]) >= int(sYear):
                rellist.append(r)   
    relset = set(rellist)
    releaselist = sorted(list(relset))
    
    dItem = {}
    if sType == '0':
        dItem = {}
        dItem['Release'] = 'Common'
        # dResult['data']['items'].append(dItem)
        # dItem = {}
        # dItem['Release'] = 'Support'
    else:
        dItem = {}
        dItem['Release'] = 'All'
        dResult['data']['items'].append(dItem)
        dItem = {}
        dItem['Release'] = 'Category'
        
    dResult['data']['items'].append(dItem)
    for rel in releaselist:        
        dItem = {}
        dItem['Release'] = rel
        dResult['data']['items'].append(dItem)
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')


def rd_resource(request):
    try:
        sUsername = request.GET['username']
        sGrade = request.GET['grade']
        sLevel = request.GET['level']
        sDomain = request.GET['pdomain']
        sYear = request.GET['year']
        sCompetence = request.GET['competence']
        sSite = request.GET['site']
        sBL = request.GET['businessline']        
        
        if sYear == '2023' and sBL == 'BBD-NWF' and (sDomain == 'Cloud' or sDomain == 'Mesh' or sDomain == 'Mobile App' or sDomain == 'Container App'):
            sBL = 'BBD-CSW'
        elif sYear == '2023' and sBL == 'BBD-CSW' and (sDomain == 'Cloud' or sDomain == 'Mesh' or sDomain == 'Mobile App' or sDomain == 'Container App'):
            sBL = 'BBD-NWF' 
        elif sYear == '2023' and sBL == 'BBD-CSW' and sDomain == 'all':
            sBL = 'No'
        sBL2 =''
        if sYear == '2023' and sBL == 'BBD-NWF' and sDomain == 'all':
            sBL2 = 'BBD-CSW'
          
    except:
        return HttpResponse('Invalid Parameters', content_type='application/json')
    
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['column'] = []
    dResult['data']['items'] = []   
    
    cmd = """
            SELECT
            Releases, RCR, Description, State, TechnicalAreas, ProjectNumber, ProjectDescription, ProjectState,
            Businessline,ProductDomain,Type,Site,Activity,Competence,RCRCategories,Siteallocation,Effortjira,
            SumAllocation,Phase,
            Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,
            Year1,Year2,Year3,a.ID,b.ID,SubTasks,b.Modifier,b.RecordTime, ProjectNumber2,FeatureCategory,BusinessPriority
            FROM
            cdb_rd_resource a
            JOIN
              cdb_rd_effort b
            WHERE  
              a.ID = Right(b.ID,8) 
              %s
            """
    if sDomain =='all' and sYear =='all' and sSite =='all' and sCompetence=='all' and sBL=='all' :  
        sRule = 'ORDER BY Releases, RCR, Year'
    elif sDomain !='all' and sYear =='all' and sSite =='all' and sCompetence=='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' ORDER BY Releases, RCR, Year" % (sDomain)
    elif sDomain =='all' and sYear !='all' and sSite =='all' and sCompetence=='all' and sBL=='all' :  
        sRule = "AND Year = '%s' ORDER BY Releases, RCR, Year" % (sYear)
    elif sDomain =='all' and sYear =='all' and sSite !='all' and sCompetence=='all' and sBL=='all' :  
        sRule = "AND Site = '%s' ORDER BY Releases, RCR, Year" % (sSite)
    elif sDomain =='all' and sYear =='all' and sSite =='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sCompetence)
    elif sDomain !='all' and sYear !='all' and sSite =='all' and sCompetence=='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear)
    elif sDomain !='all' and sYear =='all' and sSite !='all' and sCompetence=='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Site = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sSite)
    elif sDomain !='all' and sYear =='all' and sSite =='all' and sCompetence!= 'all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sCompetence)
    elif sDomain !='all' and sYear !='all' and sSite !='all' and sCompetence=='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Site = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sSite)
    elif sDomain !='all' and sYear !='all' and sSite =='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sCompetence)
    elif sDomain !='all' and sYear =='all' and sSite !='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Site = '%s'AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sSite,sCompetence)    
    elif sDomain !='all' and sYear !='all' and sSite !='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Site = '%s' AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sSite,sCompetence)
    elif sDomain =='all' and sYear !='all' and sSite !='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND Year = '%s' AND Site = '%s' AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sYear,sSite,sCompetence)
    elif sDomain =='all' and sYear =='all' and sSite !='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND Site = '%s' AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sSite,sCompetence)
    elif sDomain =='all' and sYear !='all' and sSite !='all' and sCompetence =='all' and sBL=='all' :  
        sRule = "AND Year = '%s' AND Site = '%s' ORDER BY Releases, RCR, Year" % (sYear,sSite)
    elif sDomain =='all' and sYear !='all' and sSite =='all' and sCompetence !='all' and sBL=='all' :  
        sRule = "AND Year = '%s' AND Competence = '%s' ORDER BY Releases, RCR, Year" % (sYear,sCompetence) 
        
    elif sDomain =='all' and sYear =='all' and sSite =='all' and sCompetence=='all' and sBL!='all' :  
        sRule = "AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sBL)
    elif sDomain !='all' and sYear =='all' and sSite =='all' and sCompetence=='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sBL)
    elif sDomain =='all' and sYear !='all' and sSite =='all' and sCompetence=='all' and sBL!='all' : 
        if sYear != '2023':
            sRule = "AND Year = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sYear,sBL)
        else:
            sRule = "AND Year = '%s' AND (Businessline = '%s' OR Businessline = '%s') ORDER BY Releases, RCR, Year" % (sYear,sBL,sBL2)
    elif sDomain =='all' and sYear =='all' and sSite !='all' and sCompetence=='all' and sBL!='all' :  
        sRule = "AND Site = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sSite,sBL)
    elif sDomain =='all' and sYear =='all' and sSite =='all' and sCompetence !='all' and sBL!='all' :  
        sRule = "AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sCompetence,sBL)
    elif sDomain !='all' and sYear !='all' and sSite =='all' and sCompetence=='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sBL)
    elif sDomain !='all' and sYear =='all' and sSite !='all' and sCompetence=='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Site = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sSite,sBL)
    elif sDomain !='all' and sYear =='all' and sSite =='all' and sCompetence!= 'all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sCompetence,sBL)
    elif sDomain !='all' and sYear !='all' and sSite !='all' and sCompetence=='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Site = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sSite,sBL)
    elif sDomain !='all' and sYear !='all' and sSite =='all' and sCompetence !='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sCompetence,sBL)
    elif sDomain !='all' and sYear =='all' and sSite !='all' and sCompetence !='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Site = '%s'AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sSite,sCompetence,sBL)    
    elif sDomain !='all' and sYear !='all' and sSite !='all' and sCompetence !='all' and sBL!='all' :  
        sRule = "AND ProductDomain = '%s' AND Year = '%s' AND Site = '%s' AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sDomain,sYear,sSite,sCompetence,sBL)
    elif sDomain =='all' and sYear !='all' and sSite !='all' and sCompetence !='all' and sBL!='all' : 
        if sYear != '2023':
            sRule = "AND Year = '%s' AND Site = '%s' AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sYear,sSite,sCompetence,sBL)
        else:
            sRule = "AND Year = '%s' AND Site = '%s' AND Competence = '%s' AND (Businessline = '%s' OR Businessline = '%s') ORDER BY Releases, RCR, Year" % (sYear,sSite,sCompetence,sBL,sBL2)
    elif sDomain =='all' and sYear =='all' and sSite !='all' and sCompetence !='all' and sBL!='all' :  
        sRule = "AND Site = '%s' AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sSite,sCompetence,sBL)
    elif sDomain =='all' and sYear !='all' and sSite !='all' and sCompetence =='all' and sBL!='all' : 
        if sYear != '2023':
            sRule = "AND Year = '%s' AND Site = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sYear,sSite,sBL)
        else:
            sRule = "AND Year = '%s' AND Site = '%s' AND (Businessline = '%s' OR Businessline = '%s') ORDER BY Releases, RCR, Year" % (sYear,sSite,sBL,sBL2)    
    elif sDomain =='all' and sYear !='all' and sSite =='all' and sCompetence !='all' and sBL!='all' : 
        if sYear != '2023':
            sRule = "AND Year = '%s' AND Competence = '%s' AND Businessline = '%s' ORDER BY Releases, RCR, Year" % (sYear,sCompetence,sBL) 
        else:
            sRule = "AND Year = '%s' AND Competence = '%s' AND (Businessline = '%s' OR Businessline = '%s') ORDER BY Releases, RCR, Year" % (sYear,sCompetence,sBL,sBL2)
    else:
        sRule =''
    SQLConn = analyzer_db()
    # print('SQL=', sRule,sYear,sCompetence,sBL,file=fa,flush=True )
    
    SQLConn.cur.execute(cmd % sRule)
    SQLResult = SQLConn.cur.fetchall()
    SQLConn.close()
    dAve = {}    
    dAve['items'] = [] 
    for row in SQLResult:       
        Yave = 0        
        if row[20] !='':
           ave2 = float(row[20])
           Yave = Yave + ave2
        if row[21] !='':            
           ave3 = float(row[21])           
           Yave = Yave + ave3
        if row[22] !='':
           ave4 = float(row[22])
           Yave = Yave + ave4
        if row[23] !='':
           ave5 = float(row[23])
           Yave = Yave + ave5
        if row[24] !='':
           ave6 = float(row[24]) 
           Yave = Yave + ave6
        if row[25] !='':
           ave7 = float(row[25])
           Yave = Yave + ave7
        if row[26] !='':
           ave8 = float(row[26])
           Yave = Yave + ave8
        if row[27] !='':
           ave9 = float(row[27])
           Yave = Yave + ave9
        if row[28] !='':
           ave10 = float(row[28])
           Yave = Yave + ave10
        if row[29] !='':
           ave11 = float(row[29])
           Yave = Yave + ave11
        if row[30] !='':
           ave12 = float(row[30])
           Yave = Yave + ave12
        if row[31] !='':
           ave1 = float(row[31])
           Yave = Yave + ave1
        Yave = Yave/12
        RCR = row[1]
        Competence = row[13]
        TA = row[9]
        ID  = row[35]
        YID = row[36]
        Year = row[19]
        Year1 = row[32]
        Year2 = row[33]
        Year3 = row[34]
        flag = 0        
        for item in dAve['items']:
            if RCR == item['RCR'] and TA == item['TA'] and Competence == item['Competence'] and (Year == item['Year1'] or Year == item['Year2'] or Year == item['Year3']):
                flag = 1
                if YID not in item['YID']:                   
                   ave = item['Yave']
                   item['Yave'] = ave + Yave
                   item['YID'].append(YID)
                   
        if flag == 0:            
            dR = {}
            dR['YID'] = []           
            dR['RCR'] = RCR
            dR['TA'] = TA
            dR['Competence'] = Competence
            dR['Year1'] = Year1
            dR['Year2'] = Year2
            dR['Year3'] = Year3
            dR['Yave'] = Yave            
            dR['YID'].append(YID)
            dAve['items'].append(dR)
            
    for row in SQLResult:
        Yave = 0        
        if row[20] !='':
           ave2 = float(row[20])
           Yave = Yave + ave2
        if row[21] !='':
           ave3 = float(row[21])
           Yave = Yave + ave3
        if row[22] !='':
           ave4 = float(row[22])
           Yave = Yave + ave4
        if row[23] !='':
           ave5 = float(row[23])
           Yave = Yave + ave5
        if row[24] !='':
           ave6 = float(row[24]) 
           Yave = Yave + ave6
        if row[25] !='':
           ave7 = float(row[25])
           Yave = Yave + ave7
        if row[26] !='':
           ave8 = float(row[26])
           Yave = Yave + ave8
        if row[27] !='':
           ave9 = float(row[27])
           Yave = Yave + ave9
        if row[28] !='':
           ave10 = float(row[28])
           Yave = Yave + ave10
        if row[29] !='':
           ave11 = float(row[29])
           Yave = Yave + ave11
        if row[30] !='':
           ave12 = float(row[30])
           Yave = Yave + ave12
        if row[31] !='':
           ave1 = float(row[31])
           Yave = Yave + ave1
        Yave = Yave/12
        RCR = row[1]
        TA = row[9]
        Competence =row[13]
        Year = row[19]
        Year1 = row[32]
        Year2 = row[33]
        Year3 = row[34]
        Sumave = 0
        for item in dAve['items']:
            if RCR == item['RCR'] and TA == item['TA'] and Competence == item['Competence'] and (Year == item['Year1'] or Year == item['Year2'] or Year == item['Year3']):
               Sumave = item['Yave']
        Available = 0
        if row[15] != '' and row[15] !=',' :
            Available = round(float(row[15])-Sumave,2)
        elif row[16] !='' and row[16] !=',' :
            try:
                Available = round(float(row[16])-Sumave,2)
            except:
                # print('Yave err =',RCR, row[16], file=fa, flush=True )
                pass
        elif Yave >0 :
            Available = 0 - Sumave
        dItem = {}
        dItem['Releases'] = row[0]
        dItem['RCR'] = row[1]
        dItem['Description'] = row[2]
        dItem['State'] = row[3]
        
        dItem['ProjectNumber'] = row[5]
        dItem['ProjectNumber2'] = row[40]
        dItem['ProjectDescription'] = row[6]
        # dItem['ProjectState'] = row[7]
        dItem['Businessline'] = row[8]
        dItem['ProductDomain'] = row[9]
        dItem['TechnicalAreas'] = row[4]
        dItem['Type'] = row[10]
        dItem['Site'] = row[11]
        # dItem['Activity'] = row[12]
        dItem['Competence'] = row[13]
        dItem['RCRCategories'] = row[14]
        dItem['Phase'] = row[18]
        dItem['FeatureCategory'] = row[41]
        dItem['BusinessPriority'] = row[42]
        dItem['Siteallocation'] = row[15][:5]
        dItem['Effortjira'] = row[16][:5]
        dItem['SumAllocation'] = str("%.2f" % (Available))        
        dItem['SubTasks'] = row[37]                
        dItem['Year'] = row[19] 
        dItem['Sumave'] = str("%.2f" % (Sumave))
        dItem['Yaverage'] = str("%.2f" % (Yave))
        dItem['Jans'] = row[20]
        dItem['Febs'] = row[21]
        dItem['Mars'] = row[22]
        dItem['Aprs'] = row[23]
        dItem['Mays'] = row[24]
        dItem['Juns'] = row[25]
        dItem['Juls'] = row[26]
        dItem['Augs'] = row[27]
        dItem['Seps'] = row[28]
        dItem['Octs'] = row[29]
        dItem['Novs'] = row[30]
        dItem['Decs'] = row[31]
        dItem['Year1'] = row[32]
        dItem['Year2'] = row[33]
        dItem['Year3'] = row[34]
        dItem['Modifier'] = row[38]
        dItem['RecordTime'] = str(row[39])
        dItem['ID'] = row[35]
        dItem['YID'] = row[36]
        
        
        dResult['data']['items'].append(dItem)
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')

def resource_sum(request):
    # ud = request.get_full_path()
    # print(ud, file=fa,flush=True ) 
    try:
        sUsername = request.GET['username']
        sGrade = request.GET['grade']
        sLevel = request.GET['level']
        sDomain = request.GET['pdomain']
        sYear = request.GET['year']
        sCompetence = request.GET['competence']
        sSite = request.GET['site']
        sBL = request.GET['businessline']        
        
        if sBL == 'BBD-ONT':
            sHC = 'ONT'
        elif sBL == 'BBD-NWF':
            sHC = 'Beacon'
        elif sBL == 'BBD-FWA':
            sHC = 'FWA'
        elif sBL == 'BBD-CSW':
            sHC = 'CSW'   
        if sYear == '2023' and sBL == 'BBD-NWF' and (sDomain == 'Cloud' or sDomain == 'Mesh' or sDomain == 'Mobile App' or sDomain == 'Container App'):
            sBL = 'BBD-CSW'
        elif sYear == '2023' and sBL == 'BBD-CSW' and (sDomain == 'Cloud' or sDomain == 'Mesh' or sDomain == 'Mobile App' or sDomain == 'Container App'):
            sBL = 'BBD-NWF'    
        elif sYear == '2023' and sBL == 'BBD-CSW' and sDomain == 'all':
            sBL = 'No'
        sBL2 =''
        if sYear == '2023' and sBL == 'BBD-NWF' and sDomain == 'all':
            sBL2 = 'BBD-CSW'
          
        # sYear = datetime.today().strftime('%Y')
    except:
        return HttpResponse('Invalid Parameters', content_type='application/json')
    
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['column'] = []
    dResult['data']['items'] = []   
    
    cmd = """
            SELECT            
            Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,
            Type,Releases,ProductDomain,
            a.ID,b.ID,Businessline
            FROM
            cdb_rd_resource a
            JOIN
              cdb_rd_effort b
            ON 
              a.ID = Right(b.ID,8) 
              %s
            """
    cmd_hc = """
            SELECT               
               Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,ID,
               DomainPL,HCType1,HCType2,BudgetRollup
            FROM
              cdb_hc_budget
               %s         
            """
    if sDomain =='all':  
        
        if sCompetence == 'all' and sSite =='all' and sBL == 'all':
            sRule = "WHERE Year = '%s' " % sYear
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s'  AND HCType1 = 'R&D'  " % sYear
        elif sCompetence == 'all' and sSite !='all' and sBL == 'all':
            sRule = "WHERE Year = '%s' AND Site = '%s'" % (sYear,sSite)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Country = '%s' AND HCType1 = 'R&D'  " % (sYear,sSite)
        elif sCompetence != 'all' and sSite == 'all' and sBL == 'all':
            sRule = "WHERE Year = '%s' AND Competence = '%s'" % (sYear,sCompetence)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Team = '%s' AND HCType1 = 'R&D' " % (sYear,sCompetence)
        elif sCompetence != 'all' and sSite !='all' and sBL == 'all':
            sRule = "WHERE Year = '%s' AND Competence = '%s' AND Site = '%s'" % (sYear,sCompetence,sSite)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Team = '%s' AND Country = '%s' AND HCType1 = 'R&D'  " % (sYear,sCompetence,sSite)    
    
        elif sCompetence == 'all' and sSite =='all' and sBL != 'all':
            if sYear != '2023':
                sRule = "WHERE Year = '%s' AND Businessline = '%s' " % (sYear,sBL)
            else:
                sRule = "WHERE Year = '%s' AND (Businessline = '%s' or Businessline = '%s') " % (sYear,sBL,sBL2)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND DU = '%s' AND HCType1 = 'R&D'  " % (sYear,sHC)
        elif sCompetence == 'all' and sSite !='all' and sBL != 'all':
            if sYear != '2023': 
                sRule = "WHERE Year = '%s' AND Site = '%s' AND Businessline = '%s' " % (sYear,sSite,sBL)
            else:
                sRule = "WHERE Year = '%s' AND Site = '%s' AND (Businessline = '%s' or Businessline = '%s') " % (sYear,sSite,sBL,sBL2)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Country = '%s' AND DU = '%s'  AND HCType1 = 'R&D'  " % (sYear,sSite,sHC)
        elif sCompetence != 'all' and sSite == 'all' and sBL != 'all':
            if sYear != '2023':
                sRule = "WHERE Year = '%s' AND Competence = '%s' AND Businessline = '%s'" % (sYear,sCompetence,sBL)
            else:
                sRule = "WHERE Year = '%s' AND Competence = '%s' AND (Businessline = '%s' or Businessline = '%s') " % (sYear,sCompetence,sBL,sBL2)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Team = '%s' AND DU = '%s' AND HCType1 = 'R&D' " % (sYear,sCompetence,sHC)
        elif sCompetence != 'all' and sSite !='all' and sBL != 'all':
            if sYear != '2023':
                sRule = "WHERE Year = '%s' AND Competence = '%s' AND Site = '%s' AND Businessline = '%s'" % (sYear,sCompetence,sSite,sBL)
            else:
                sRule = "WHERE Year = '%s' AND Competence = '%s' AND Site = '%s' AND (Businessline = '%s' or Businessline = '%s') " % (sYear,sCompetence,sSite,sBL,sBL2)    
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Team = '%s' AND Country = '%s' AND DU = '%s' AND HCType1 = 'R&D'  " % (sYear,sCompetence,sSite,sHC)         
       
    else:        
        
        if sCompetence == 'all' and sSite =='all' and sBL == 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND ProductDomain = '%s' " % (sYear,sDomain)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND DomainPL = '%s' AND HCType1 = 'R&D'  " % (sYear,sDomain)
        elif sCompetence == 'all' and sSite !='all' and sBL == 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND ProductDomain = '%s' AND Site = '%s' " % (sYear,sDomain,sSite)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND DomainPL = '%s' AND Country = '%s' AND HCType1 = 'R&D'  " % (sYear,sDomain,sSite)
        elif sCompetence != 'all' and sSite == 'all' and sBL == 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND ProductDomain = '%s' AND Competence = '%s' " % (sYear,sDomain,sCompetence)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND DomainPL = '%s' AND Team = '%s' AND HCType1 = 'R&D' " % (sYear,sDomain,sCompetence)
        elif sCompetence != 'all' and sSite !='all' and sBL == 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND ProductDomain = '%s' AND Competence = '%s' AND Site = '%s' " % (sYear,sDomain,sCompetence,sSite)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND DomainPL = '%s' AND Team = '%s' AND Country = '%s' AND HCType1 = 'R&D'  " % (sYear,sDomain,sCompetence,sSite) 
        
        elif sCompetence == 'all' and sSite =='all' and sBL != 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND ProductDomain = '%s' AND Businessline = '%s' " % (sYear,sDomain,sBL)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND DomainPL = '%s' AND DU = '%s' AND HCType1 = 'R&D'  " % (sYear,sDomain,sHC)
        elif sCompetence == 'all' and sSite !='all' and sBL != 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND Site = '%s' AND ProductDomain = '%s' AND Businessline = '%s' " % (sYear,sSite,sDomain,sBL)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Country = '%s' AND DomainPL = '%s' AND DU = '%s' AND HCType1 = 'R&D'  " % (sYear,sSite,sDomain,sHC)        
        elif sCompetence != 'all' and sSite == 'all' and sBL != 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND Competence = '%s' AND ProductDomain = '%s' AND Businessline = '%s'" % (sYear,sCompetence,sDomain,sBL)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Team = '%s' AND DomainPL = '%s' AND DU = '%s' AND HCType1 = 'R&D' " % (sYear,sCompetence,sDomain,sHC)
        elif sCompetence != 'all' and sSite !='all' and sBL != 'all' and sDomain != 'all':
            sRule = "WHERE Year = '%s' AND Competence = '%s' AND Site = '%s' AND ProductDomain = '%s' AND Businessline = '%s'" % (sYear,sCompetence,sSite,sDomain,sBL)
            sRule1 = "WHERE BudgetRollup = 'Rollup' AND Year = '%s' AND Team = '%s' AND Country = '%s' AND DomainPL = '%s' AND DU = '%s' AND HCType1 = 'R&D'  " % (sYear,sCompetence,sSite,sDomain,sHC) 
     
    SQLConn = analyzer_db() 
    SQLConn.cur.execute(cmd % sRule)
    SQLResult = SQLConn.cur.fetchall()
    SQLConn.cur.execute(cmd_hc % sRule1)
    SQLResult_hc = SQLConn.cur.fetchall()
    SQLConn.close()
    Yave1 = 0
    Yave2 = 0
    Yave3 = 0
    Yave4 = 0
    Yave5 = 0
    Yave6 = 0
    Yave7 = 0
    Yave8 = 0
    Yave9 = 0
    Yave10 = 0
    Yave11 = 0
    Yave12 = 0
    k = 0
    for row_hc in SQLResult_hc:
        k += 1                
        if row_hc[1] !='':
           Yave1 += float(row_hc[1])
        if row_hc[2] !='':
           Yave2 += float(row_hc[2])
        if row_hc[3] !='':
           Yave3 += float(row_hc[3])
        if row_hc[4] !='':
           Yave4 += float(row_hc[4])
        if row_hc[5] !='':
           Yave5 += float(row_hc[5])
        if row_hc[6] !='':
           Yave6 += float(row_hc[6])
        if row_hc[7] !='':
           Yave7 += float(row_hc[7])
        if row_hc[8] !='':
           Yave8 += float(row_hc[8])
        if row_hc[9] !='':
           Yave9 += float(row_hc[9])
        if row_hc[10] !='':
           Yave10 += float(row_hc[10])
        if row_hc[11] !='':
           Yave11 += float(row_hc[11])
        if row_hc[12] !='':
           Yave12 += float(row_hc[12])   
    sum_hc = (Yave1 + Yave2 + Yave3 + Yave4  + Yave5 + Yave6 + Yave7 + Yave8  +Yave9 + Yave10 + Yave11 + Yave12)/12
    # print('hc num =',str(k), sDomain,file=fa,flush=True )
    dItem = {}
    dItem['Summary'] = str("%.1f" % (sum_hc))    
    dItem['Jans'] = str("%.1f" % (Yave1))
    dItem['Febs'] = str("%.1f" % (Yave2))
    dItem['Mars'] = str("%.1f" % (Yave3))
    dItem['Aprs'] = str("%.1f" % (Yave4))
    dItem['Mays'] = str("%.1f" % (Yave5))
    dItem['Juns'] = str("%.1f" % (Yave6))
    dItem['Juls'] = str("%.1f" % (Yave7))
    dItem['Augs'] = str("%.1f" % (Yave8))
    dItem['Seps'] = str("%.1f" % (Yave9))
    dItem['Octs'] = str("%.1f" % (Yave10))
    dItem['Novs'] = str("%.1f" % (Yave11))
    dItem['Decs'] = str("%.1f" % (Yave12))    
        
    dResult['data']['items'].append(dItem)
    
    fd_Yave1 = 0
    fd_Yave2 = 0
    fd_Yave3 = 0
    fd_Yave4 = 0
    fd_Yave5 = 0
    fd_Yave6 = 0
    fd_Yave7 = 0
    fd_Yave8 = 0
    fd_Yave9 = 0
    fd_Yave10 = 0
    fd_Yave11 = 0
    fd_Yave12 = 0
    
    fa_Yave1 = 0
    fa_Yave2 = 0
    fa_Yave3 = 0
    fa_Yave4 = 0
    fa_Yave5 = 0
    fa_Yave6 = 0
    fa_Yave7 = 0
    fa_Yave8 = 0
    fa_Yave9 = 0
    fa_Yave10 = 0
    fa_Yave11 = 0
    fa_Yave12 = 0
    
    fx_Yave1 = 0
    fx_Yave2 = 0
    fx_Yave3 = 0
    fx_Yave4 = 0
    fx_Yave5 = 0
    fx_Yave6 = 0
    fx_Yave7 = 0
    fx_Yave8 = 0
    fx_Yave9 = 0
    fx_Yave10 = 0
    fx_Yave11 = 0
    fx_Yave12 = 0
    g = 0
    h = 0
    a = 0
    for row in SQLResult: 
        if row[13] =='FD':
            g += 1 
        elif row[13] =='Fixed':
            h += 1  
        # elif row[13] =='FD Available':
        #     a += 1             
        if row[1] !='':
           if row[13] =='FD':               
               fd_Yave1 += float(row[1])
           elif row[13] =='Fixed':              
               fx_Yave1 += float(row[1]) 
           # elif row[13] =='FD Available':              
           #     fa_Yave1 += float(row[1])    
        if row[2] !='':
           if row[13] =='FD':
               fd_Yave2 += float(row[2])
           elif row[13] =='Fixed':
               fx_Yave2 += float(row[2])
           # elif row[13] =='FD Available':              
           #     fa_Yave2 += float(row[2])
        if row[3] !='':
           if row[13] =='FD':
               fd_Yave3 += float(row[3])
           elif row[13] =='Fixed':
               fx_Yave3 += float(row[3])
           # elif row[13] =='FD Available':              
           #     fa_Yave3 += float(row[3])
        if row[4] !='':
           if row[13] =='FD':
               fd_Yave4 += float(row[4])
           elif row[13] =='Fixed':
               fx_Yave4 += float(row[4]) 
           # elif row[13] =='FD Available':              
           #     fa_Yave4 += float(row[4])
        if row[5] !='':
           if row[13] =='FD':
               fd_Yave5 += float(row[5])
           elif row[13] =='Fixed':
               fx_Yave5 += float(row[5]) 
           # elif row[13] =='FD Available':              
           #     fa_Yave5 += float(row[5])
        if row[6] !='':
           if row[13] =='FD':
               fd_Yave6 += float(row[6])
           elif row[13] =='Fixed':
               fx_Yave6 += float(row[6])
           # elif row[13] =='FD Available':              
           #     fa_Yave6 += float(row[6])
        if row[7] !='':
           if row[13] =='FD':
               fd_Yave7 += float(row[7])
           elif row[13] =='Fixed':
               fx_Yave7 += float(row[7])
           # elif row[13] =='FD Available':              
           #     fa_Yave7 += float(row[7])
        if row[8] !='':
           if row[13] =='FD':
               fd_Yave8 += float(row[8])
           elif row[13] =='Fixed':
               fx_Yave8 += float(row[8])
           # elif row[13] =='FD Available':              
           #     fa_Yave8 += float(row[8])
        if row[9] !='':
           if row[13] =='FD':
               fd_Yave9 += float(row[9])
           elif row[13] =='Fixed':
               fx_Yave9 += float(row[9]) 
           # elif row[13] =='FD Available':              
           #     fa_Yave9 += float(row[9])
        if row[10] !='':
           if row[13] =='FD':
               fd_Yave10 += float(row[10])
           elif row[13] =='Fixed':
               fx_Yave10 += float(row[10])
           # elif row[13] =='FD Available':              
           #     fa_Yave10 += float(row[10])    
        if row[11] !='':
           if row[13] =='FD':
               fd_Yave11 += float(row[11])
           elif row[13] =='Fixed':
               fx_Yave11 += float(row[11])
           # elif row[13] =='FD Available':              
           #     fa_Yave11 += float(row[11])
        if row[12] !='':
           if row[13] =='FD':
               fd_Yave12 += float(row[12])
           elif row[13] =='Fixed':
               fx_Yave12 += float(row[12]) 
           # elif row[13] =='FD Available':              
           #     fa_Yave12 += float(row[12])     
    
    fe_Yave1 = Yave1 - fx_Yave1  - fd_Yave1
    fe_Yave2 = Yave2 - fx_Yave2  - fd_Yave2
    fe_Yave3 = Yave3 - fx_Yave3  - fd_Yave3
    fe_Yave4 = Yave4 - fx_Yave4  - fd_Yave4
    fe_Yave5 = Yave5 - fx_Yave5  - fd_Yave5
    fe_Yave6 = Yave6 - fx_Yave6  - fd_Yave6
    fe_Yave7 = Yave7 - fx_Yave7  - fd_Yave7
    fe_Yave8 = Yave8 - fx_Yave8  - fd_Yave8
    fe_Yave9 = Yave9 - fx_Yave9  - fd_Yave9
    fe_Yave10 = Yave10 - fx_Yave10 - fd_Yave10
    fe_Yave11 = Yave11 - fx_Yave11 - fd_Yave11
    fe_Yave12 = Yave12 - fx_Yave12 - fd_Yave12
          
    sum_fd = (fd_Yave1 + fd_Yave2 + fd_Yave3 + fd_Yave4  + fd_Yave5 + fd_Yave6 + fd_Yave7 + fd_Yave8  +fd_Yave9 + fd_Yave10 + fd_Yave11 + fd_Yave12)/12  
    sum_fx = (fx_Yave1 + fx_Yave2 + fx_Yave3 + fx_Yave4  + fx_Yave5 + fx_Yave6 + fx_Yave7 + fx_Yave8  +fx_Yave9 + fx_Yave10 + fx_Yave11 + fx_Yave12)/12  
    sum_fa = (fa_Yave1 + fa_Yave2 + fa_Yave3 + fa_Yave4  + fa_Yave5 + fa_Yave6 + fa_Yave7 + fa_Yave8  +fa_Yave9 + fa_Yave10 + fa_Yave11 + fa_Yave12)/12
    sum_fe = sum_hc - sum_fx - sum_fd
    # print('fd num =',str(g), 'fx num =',str(h),sDomain,file=fa,flush=True )
    dItem = {}
    dItem['Summary'] = str("%.1f" % (sum_fx))    
    dItem['Jans'] = str("%.1f" % (fx_Yave1))
    dItem['Febs'] = str("%.1f" % (fx_Yave2))
    dItem['Mars'] = str("%.1f" % (fx_Yave3))
    dItem['Aprs'] = str("%.1f" % (fx_Yave4))
    dItem['Mays'] = str("%.1f" % (fx_Yave5))
    dItem['Juns'] = str("%.1f" % (fx_Yave6))
    dItem['Juls'] = str("%.1f" % (fx_Yave7))
    dItem['Augs'] = str("%.1f" % (fx_Yave8))
    dItem['Seps'] = str("%.1f" % (fx_Yave9))
    dItem['Octs'] = str("%.1f" % (fx_Yave10))
    dItem['Novs'] = str("%.1f" % (fx_Yave11))
    dItem['Decs'] = str("%.1f" % (fx_Yave12))
    dResult['data']['items'].append(dItem)
    
    # dItem = {}
    # dItem['Summary'] = str("%.1f" % (sum_fa))    
    # dItem['Jans'] = str("%.1f" % (fa_Yave1))
    # dItem['Febs'] = str("%.1f" % (fa_Yave2))
    # dItem['Mars'] = str("%.1f" % (fa_Yave3))
    # dItem['Aprs'] = str("%.1f" % (fa_Yave4))
    # dItem['Mays'] = str("%.1f" % (fa_Yave5))
    # dItem['Juns'] = str("%.1f" % (fa_Yave6))
    # dItem['Juls'] = str("%.1f" % (fa_Yave7))
    # dItem['Augs'] = str("%.1f" % (fa_Yave8))
    # dItem['Seps'] = str("%.1f" % (fa_Yave9))
    # dItem['Octs'] = str("%.1f" % (fa_Yave10))
    # dItem['Novs'] = str("%.1f" % (fa_Yave11))
    # dItem['Decs'] = str("%.1f" % (fa_Yave12))
    # dResult['data']['items'].append(dItem)
    
    dItem = {}
    dItem['Summary'] = str("%.1f" % (sum_fd))    
    dItem['Jans'] = str("%.1f" % (fd_Yave1))
    dItem['Febs'] = str("%.1f" % (fd_Yave2))
    dItem['Mars'] = str("%.1f" % (fd_Yave3))
    dItem['Aprs'] = str("%.1f" % (fd_Yave4))
    dItem['Mays'] = str("%.1f" % (fd_Yave5))
    dItem['Juns'] = str("%.1f" % (fd_Yave6))
    dItem['Juls'] = str("%.1f" % (fd_Yave7))
    dItem['Augs'] = str("%.1f" % (fd_Yave8))
    dItem['Seps'] = str("%.1f" % (fd_Yave9))
    dItem['Octs'] = str("%.1f" % (fd_Yave10))
    dItem['Novs'] = str("%.1f" % (fd_Yave11))
    dItem['Decs'] = str("%.1f" % (fd_Yave12))
    dResult['data']['items'].append(dItem)
    
    dItem = {}
    dItem['Summary'] = str("%.1f" % (sum_fe))    
    dItem['Jans'] = str("%.1f" % (fe_Yave1))
    dItem['Febs'] = str("%.1f" % (fe_Yave2))
    dItem['Mars'] = str("%.1f" % (fe_Yave3))
    dItem['Aprs'] = str("%.1f" % (fe_Yave4))
    dItem['Mays'] = str("%.1f" % (fe_Yave5))
    dItem['Juns'] = str("%.1f" % (fe_Yave6))
    dItem['Juls'] = str("%.1f" % (fe_Yave7))
    dItem['Augs'] = str("%.1f" % (fe_Yave8))
    dItem['Seps'] = str("%.1f" % (fe_Yave9))
    dItem['Octs'] = str("%.1f" % (fe_Yave10))
    dItem['Novs'] = str("%.1f" % (fe_Yave11))
    dItem['Decs'] = str("%.1f" % (fe_Yave12))
    dResult['data']['items'].append(dItem)
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')

def rd_project(request):
    try:
        sType = request.GET['type']
        sProjectNumber = request.GET['ProjectNumber']
    except:
        print('PJ 6 =','error', file=fa,flush=True ) 
        return HttpResponse('Invalid Parameters', content_type='application/json')
     
    dResult = {}        
    dResult['data'] = {}
    dResult['data']['status'] = []
    SQLBConn = pymysql.connect(host  = settings.BBD_DB['host'],
                            port     = settings.BBD_DB['port'],
                            user     = settings.BBD_DB['username'],
                            password = settings.BBD_DB['password'],
                            database = settings.BBD_DB['name'],
                            charset  = settings.BBD_DB['charset']
                        )
    SQLBCur = SQLBConn.cursor()
    SQLB_pipl = """
    SELECT
    `Key`,Summary      
    FROM
    jira_issues_bbdpipl           
    """        
    SQLBCur.execute(SQLB_pipl)             
    SQLBResult_pipl = SQLBCur.fetchall()      
    Project =''
    if  'BBDPIPL' in sProjectNumber:
        for row in SQLBResult_pipl:
            if row[0] == sProjectNumber:
                print('PJ 1 =',sProjectNumber,row[0], file=fa,flush=True ) 
                Project = row[1].replace("'","\\'")                                         
                dResult['data']['status']= Project
                break
           
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json') 
   
def rd_resource_edit(request):
    ud = request.get_full_path()
    print(ud, file=fa,flush=True ) 
    try:
        sType = request.GET['stype'] 
        sMail = request.GET['mail']
        sGrade = request.GET['roles']
        sLevel = request.GET['level']
        sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
        sUpdate = datetime.today().strftime('%Y-%m-%d')
        Modifier = sMail + '[' + sUpdate + ']' 
        if sType == '0':
           sID = request.GET['ID']
           sYID = request.GET['YID']
           sYear = request.GET['Year']
           sYear1 = request.GET['Y1']
           sYear2 = request.GET['Y2']
           sYear3 = request.GET['Y3']
           sSite = request.GET['Site']
           sRCR  = request.GET['RCR']
           sRelease = request.GET['Release']
           sProductDomain = request.GET['ProductDomain']
           sCompetence = request.GET['Competence']
           pType = fd_type(sRCR)
        elif sType == '8':           
           sYear = request.GET['Year']
           sYear1 = request.GET['Y1']
           sYear2 = request.GET['Y2']
           sYear3 = request.GET['Y3']
           sSite = request.GET['Site']
           sRCR  = request.GET['RCR']
           sRelease = request.GET['Release']
           sProductDomain = request.GET['ProductDomain']
           sCompetence = request.GET['Competence']
           pType = fd_type(sRCR)
           # print(sID,sYear, file=fa,flush=True )
        elif sType == '6':           
           sYear = request.GET['Year']
           sProductDomain = request.GET['ProductDomain']
           sCompetence = request.GET['Competence']
           sRCR  = request.GET['RCR']
           sSite = request.GET['Site']   
        elif sType == '2' or sType == '5':
            sID = request.GET['ID']
            sYID = request.GET['YID']
        elif sType == '22':            
            sYID = request.GET['YID']            
            sVal = request.GET['val']
            sMonth = request.GET['month']
                
        elif sType == '7':
            sID = request.GET['ID']
            sYID = request.GET['YID']
            sJans = request.GET['Jans']
            sFebs = request.GET['Febs']
            sMars = request.GET['Mars']
            sAprs = request.GET['Aprs']
            sMays = request.GET['Mays']
            sJuns = request.GET['Juns']
            sJuls = request.GET['Juls']
            sAugs = request.GET['Augs']
            sSeps = request.GET['Seps']
            sOcts = request.GET['Octs']
            sNovs = request.GET['Novs']
            sDecs = request.GET['Decs']
        elif sType == '3':            
            sYID = request.GET['YID']
            # print(sYID, file=fa,flush=True )
            
        if sType == '1' or sType == '2' or sType == '4' or sType == '5':
            sProductDomain = request.GET['ProductDomain']
            sRelease = request.GET['Release']
            sRCR  = request.GET['RCR']          
            sDescription = request.GET['Description']
            sDescription = repspecial(sDescription)
            # sTechnicalAreas = request.GET['TechnicalAreas']
            sBusinessLine = ''
            sProjectNumber = request.GET['ProjectNumber']
            sProjectDescription = request.GET['ProjectDescription']
            sProjectDescription = repspecial(sProjectDescription)
            phase = phase_type(sRCR)   
            if sRelease == 'Common':
                pType = fd_type(sRCR)
            else:
                pType ='FD'
            sSite = request.GET['Site']
            # sActivity = request.GET['Activity']
            sCompetence = request.GET['Competence']
            sRCRCategories = request.GET['RCRCategories']           
            # sPhase = request.GET['Phase']
            sYear = request.GET['Year']
            sJans = request.GET['Jans']
            sFebs = request.GET['Febs']
            sMars = request.GET['Mars']
            sAprs = request.GET['Aprs']
            sMays = request.GET['Mays']
            sJuns = request.GET['Juns']
            sJuls = request.GET['Juls']
            sAugs = request.GET['Augs']
            sSeps = request.GET['Seps']
            sOcts = request.GET['Octs']
            sNovs = request.GET['Novs']
            sDecs = request.GET['Decs']
            sYear2 =''
            sYear3 =''
                    
        if sType == '4' or sType == '5' :             
            
            sEffort = request.GET['Effort']               
            sSite2 = request.GET['Site2']            
            sYear2 = request.GET['Year2']
            sJans2 = request.GET['Jans2']
            sFebs2 = request.GET['Febs2']
            sMars2 = request.GET['Mars2']
            sAprs2 = request.GET['Aprs2']
            sMays2 = request.GET['Mays2']
            sJuns2 = request.GET['Juns2']
            sJuls2 = request.GET['Juls2']
            sAugs2 = request.GET['Augs2']
            sSeps2 = request.GET['Seps2']
            sOcts2 = request.GET['Octs2']
            sNovs2 = request.GET['Novs2']
            sDecs2 = request.GET['Decs2'] 
            sSite3 = request.GET['Site3']            
            sYear3 = request.GET['Year3']
            sJans3 = request.GET['Jans3']
            sFebs3 = request.GET['Febs3']
            sMars3 = request.GET['Mars3']
            sAprs3 = request.GET['Aprs3']
            sMays3 = request.GET['Mays3']
            sJuns3 = request.GET['Juns3']
            sJuls3 = request.GET['Juls3']
            sAugs3 = request.GET['Augs3']
            sSeps3 = request.GET['Seps3']
            sOcts3 = request.GET['Octs3']
            sNovs3 = request.GET['Novs3']
            sDecs3 = request.GET['Decs3']
            sSite4 = request.GET['Site4']            
            sYear4 = request.GET['Year4']
            sJans4 = request.GET['Jans4']
            sFebs4 = request.GET['Febs4']
            sMars4 = request.GET['Mars4']
            sAprs4 = request.GET['Aprs4']
            sMays4 = request.GET['Mays4']
            sJuns4 = request.GET['Juns4']
            sJuls4 = request.GET['Juls4']
            sAugs4 = request.GET['Augs4']
            sSeps4 = request.GET['Seps4']
            sOcts4 = request.GET['Octs4']
            sNovs4 = request.GET['Novs4']
            sDecs4 = request.GET['Decs4']
            
            sSite5 = request.GET['Site5']
            sYear5 = request.GET['Year5']
            sJans5 = request.GET['Jans5']
            sFebs5 = request.GET['Febs5']
            sMars5 = request.GET['Mars5']
            sAprs5 = request.GET['Aprs5']
            sMays5 = request.GET['Mays5']
            sJuns5 = request.GET['Juns5']
            sJuls5 = request.GET['Juls5']
            sAugs5 = request.GET['Augs5']
            sSeps5 = request.GET['Seps5']
            sOcts5 = request.GET['Octs5']
            sNovs5 = request.GET['Novs5']
            sDecs5 = request.GET['Decs5']
            
            sSite6 = request.GET['Site6']
            sYear6 = request.GET['Year6']
            sJans6 = request.GET['Jans6']
            sFebs6 = request.GET['Febs6']
            sMars6 = request.GET['Mars6']
            sAprs6 = request.GET['Aprs6']
            sMays6 = request.GET['Mays6']
            sJuns6 = request.GET['Juns6']
            sJuls6 = request.GET['Juls6']
            sAugs6 = request.GET['Augs6']
            sSeps6 = request.GET['Seps6']
            sOcts6 = request.GET['Octs6']
            sNovs6 = request.GET['Novs6']
            sDecs6 = request.GET['Decs6']
            
            sSite7 = request.GET['Site7']
            sYear7 = request.GET['Year7']
            sJans7 = request.GET['Jans7']
            sFebs7 = request.GET['Febs7']
            sMars7 = request.GET['Mars7']
            sAprs7 = request.GET['Aprs7']
            sMays7 = request.GET['Mays7']
            sJuns7 = request.GET['Juns7']
            sJuls7 = request.GET['Juls7']
            sAugs7 = request.GET['Augs7']
            sSeps7 = request.GET['Seps7']
            sOcts7 = request.GET['Octs7']
            sNovs7 = request.GET['Novs7']
            sDecs7 = request.GET['Decs7']
            
            sSite8 = request.GET['Site8']
            sYear8 = request.GET['Year8']
            sJans8 = request.GET['Jans8']
            sFebs8 = request.GET['Febs8']
            sMars8 = request.GET['Mars8']
            sAprs8 = request.GET['Aprs8']
            sMays8 = request.GET['Mays8']
            sJuns8 = request.GET['Juns8']
            sJuls8 = request.GET['Juls8']
            sAugs8 = request.GET['Augs8']
            sSeps8 = request.GET['Seps8']
            sOcts8 = request.GET['Octs8']
            sNovs8 = request.GET['Novs8']
            sDecs8 = request.GET['Decs8']
            
            # Latestupdate = datetime.today().strftime("%Y-%m-%d")
            # print(sProductDomain,sRCR,sYear2, sJans2,sSite2,sYear3,sSite3, file=fa,flush=True )
    except:
        return HttpResponse('Invalid Parameters', content_type='application/json')    
    
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
    
    if sType == '0':         
        if sYear == sYear1:
            year1 = sYear2
            year2 = sYear3
        elif sYear == sYear2:
            year1 = sYear1
            year2 = sYear3    
        elif sYear == sYear3:
            year1 = sYear1
            year2 = sYear2 
        print('sType 0 =',sRelease,sRCR,sProductDomain,pType,sCompetence,sID,sYear,year1,year2,file=fc, flush=True )    
        cmd_s = """
                SELECT                
                Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,ID                
                FROM                
                  cdb_rd_effort
                WHERE                     
                  '%s' = Right(ID,8) 
                  AND
                  (Year = '%s' OR Year = '%s')
                 
                 ORDER BY Year
                """ % (sID,year1,year2)
        cmd = """
                SELECT
                
                Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,b.ID,Site
                FROM
                cdb_rd_resource a
                JOIN
                  cdb_rd_effort b
                WHERE  
                  a.ID = Right(b.ID,8) 
                  %s
                """  
        sRule = """AND Releases = '%s' AND RCR = '%s' AND ProductDomain = '%s'
                AND Type = '%s' AND Competence = '%s' AND a.ID <> '%s' 
                ORDER BY Site,Year""" % (sRelease,sRCR,sProductDomain,pType,sCompetence,sID)       
        
           
        SQLConn = analyzer_db()
        SQLConn.cur.execute(cmd_s)
        SQLResult_s = SQLConn.cur.fetchall()
        SQLConn.cur.execute(cmd % sRule)
        SQLResult = SQLConn.cur.fetchall()
        SQLConn.close()
        n = 0
        for row in SQLResult_s:           
            # print('Year=',row[0],row[13], file=fa,flush=True )
            n += 1
            dItem = {}
            
            dItem['Year'] = row[0]            
            dItem['Jans'] = row[1]
            dItem['Febs'] = row[2]
            dItem['Mars'] = row[3]
            dItem['Aprs'] = row[4]
            dItem['Mays'] = row[5]
            dItem['Juns'] = row[6]
            dItem['Juls'] = row[7]
            dItem['Augs'] = row[8]
            dItem['Seps'] = row[9]
            dItem['Octs'] = row[10]
            dItem['Novs'] = row[11]
            dItem['Decs'] = row[12]            
            dItem['ID'] = row[13] 
            dItem['Site'] = sSite
            
            dResult['data']['items'].append(dItem)
        for row in SQLResult:           
            # print('Year=',row[0],row[13], file=fa,flush=True )
            n += 1
            dItem = {}
            
            dItem['Year'] = row[0]            
            dItem['Jans'] = row[1]
            dItem['Febs'] = row[2]
            dItem['Mars'] = row[3]
            dItem['Aprs'] = row[4]
            dItem['Mays'] = row[5]
            dItem['Juns'] = row[6]
            dItem['Juls'] = row[7]
            dItem['Augs'] = row[8]
            dItem['Seps'] = row[9]
            dItem['Octs'] = row[10]
            dItem['Novs'] = row[11]
            dItem['Decs'] = row[12]            
            dItem['ID'] = row[13]     
            dItem['Site'] = row[14]
            
            dResult['data']['items'].append(dItem)
        for k in range(7-n):
            dItem = {}
            dItem['Site'] = ''
            dItem['Year'] = ''            
            dItem['Jans'] = ''
            dItem['Febs'] = ''
            dItem['Mars'] = ''
            dItem['Aprs'] = ''
            dItem['Mays'] = ''
            dItem['Juns'] = ''
            dItem['Juls'] = ''
            dItem['Augs'] = ''
            dItem['Seps'] = ''
            dItem['Octs'] = ''
            dItem['Novs'] = ''
            dItem['Decs'] = ''            
            dItem['ID'] = ''     
            
            dResult['data']['items'].append(dItem)
        return HttpResponse(simplejson.dumps(dResult), content_type='application/json')
    
    elif sType == '8':         
        if sYear == sYear1:
            year1 = sYear2
            year2 = sYear3
        elif sYear == sYear2:
            year1 = sYear1
            year2 = sYear3    
        elif sYear == sYear3:
            year1 = sYear1
            year2 = sYear2
        
        cmd = """
                SELECT
                
                Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,b.ID,Site
                FROM
                cdb_rd_resource a
                JOIN
                  cdb_rd_effort b
                WHERE  
                  a.ID = Right(b.ID,8) 
                  %s
                """  
        sRule = """AND Releases = '%s' AND RCR = '%s' AND ProductDomain = '%s'
                AND Type = '%s' AND Competence = '%s' AND (Site <> '%s' or Year <>'%s')
                ORDER BY Site,Year""" % (sRelease,sRCR,sProductDomain,pType,sCompetence,sSite,sYear)       
        
           
        SQLConn = analyzer_db()
        
        SQLConn.cur.execute(cmd % sRule)
        SQLResult = SQLConn.cur.fetchall()
        SQLConn.close()
        n = 0
        
        for row in SQLResult:           
            # print('Year=',row[0],row[13], file=fa,flush=True )
            n += 1
            dItem = {}
            
            dItem['Year'] = row[0]            
            dItem['Jans'] = row[1]
            dItem['Febs'] = row[2]
            dItem['Mars'] = row[3]
            dItem['Aprs'] = row[4]
            dItem['Mays'] = row[5]
            dItem['Juns'] = row[6]
            dItem['Juls'] = row[7]
            dItem['Augs'] = row[8]
            dItem['Seps'] = row[9]
            dItem['Octs'] = row[10]
            dItem['Novs'] = row[11]
            dItem['Decs'] = row[12]            
            dItem['ID'] = row[13]     
            dItem['Site'] = row[14]
            
            dResult['data']['items'].append(dItem)
        for k in range(7-n):
            dItem = {}
            dItem['Site'] = ''
            dItem['Year'] = ''            
            dItem['Jans'] = ''
            dItem['Febs'] = ''
            dItem['Mars'] = ''
            dItem['Aprs'] = ''
            dItem['Mays'] = ''
            dItem['Juns'] = ''
            dItem['Juls'] = ''
            dItem['Augs'] = ''
            dItem['Seps'] = ''
            dItem['Octs'] = ''
            dItem['Novs'] = ''
            dItem['Decs'] = ''            
            dItem['ID'] = ''     
            
            dResult['data']['items'].append(dItem)
        return HttpResponse(simplejson.dumps(dResult), content_type='application/json')
    
    elif sType == '6':         
        
        cmd = """
                SELECT                
                Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,b.ID,Year1,Year2,Year3                
                FROM
                cdb_rd_resource a
                LEFT JOIN
                  cdb_rd_effort b
                ON  
                  a.ID = Right(b.ID,8) 
                  %s
                """
        
        sRule = """ WHERE Site = '%s' AND ProductDomain = '%s' 
               AND RCR = '%s' AND Competence = '%s' AND Year = '%s'
               """ % (sSite,sProductDomain,sRCR,sCompetence,sYear)
        # print(sSite,sProductDomain,sRCR,sCompetence,sYear, file=fa, flush=True )
        SQLConn = analyzer_db() 
        # print(cmd % sRule, file=fa, flush=True )
        SQLConn.cur.execute(cmd % sRule)
        SQLResult = SQLConn.cur.fetchall()
        SQLConn.close()
        n = 0
        for row in SQLResult:
            n += 1
            dItem = {}
            dItem['Site'] = sSite
            dItem['Year'] = sYear            
            dItem['Jans'] = row[1]
            dItem['Febs'] = row[2]
            dItem['Mars'] = row[3]
            dItem['Aprs'] = row[4]
            dItem['Mays'] = row[5]
            dItem['Juns'] = row[6]
            dItem['Juls'] = row[7]
            dItem['Augs'] = row[8]
            dItem['Seps'] = row[9]
            dItem['Octs'] = row[10]
            dItem['Novs'] = row[11]
            dItem['Decs'] = row[12]            
            dItem['ID'] = row[13]
            dItem['Y1'] = row[14]
            dItem['Y2'] = row[15]
            dItem['Y3'] = row[16]
            
            dResult['data']['items'].append(dItem)
        if n == 0:
            dItem = {}
            dItem['Site'] = sSite
            dItem['Year'] = sYear           
            dItem['Jans'] = ''
            dItem['Febs'] = ''
            dItem['Mars'] = ''
            dItem['Aprs'] = ''
            dItem['Mays'] = ''
            dItem['Juns'] = ''
            dItem['Juls'] = ''
            dItem['Augs'] = ''
            dItem['Seps'] = ''
            dItem['Octs'] = ''
            dItem['Novs'] = ''
            dItem['Decs'] = ''            
            dItem['ID'] = ''     
            dItem['Y1'] = ''
            dItem['Y2'] = ''
            dItem['Y3'] = ''
            dResult['data']['items'].append(dItem)
        return HttpResponse(simplejson.dumps(dResult), content_type='application/json') 
    
    # 1 add
    elif sType == '1':
        cmd = """
                SELECT
                   count(RCR) as num                 
                FROM
                cdb_rd_resource a
                LEFT JOIN
                  cdb_rd_effort b
                ON  
                  a.ID = Right(b.ID,8) 
                  %s
                """
        if sYear == '':
            sYear = datetime.today().strftime('%Y')
        sRule = " WHERE Site = '%s' AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' AND Year = '%s' " % (sSite,sProductDomain,sRCR,sCompetence,sYear)
        # print(sSite,sProductDomain,sRCR,sCompetence,sYear, file=fa, flush=True )
        SQLConn = analyzer_db() 
        # print(cmd % sRule, file=fa, flush=True )
        SQLConn.cur.execute(cmd % sRule)
        SQLResult = SQLConn.cur.fetchall()
        count = SQLResult[0][0]    
        if count == 0 or sRelease == 'Common' or sRCR == 'PIPL (undefined RCR)' or sRCR == 'FD w/o RCR w/o PIPL': 
            if sProductDomain == 'ONT_PS' or sProductDomain == 'China Product' or sProductDomain == 'Voice' :
                sBusinessLine = 'BBD-ONT'
            elif sProductDomain == 'Cloud' or sProductDomain == 'Mobile App' or sProductDomain == 'Mesh' or sProductDomain == 'Container App' or sProductDomain == 'HomeOS':
                sBusinessLine = 'BBD-CSW'
            elif sProductDomain == 'Beacon':
                sBusinessLine = 'BBD-NWF'
            elif sProductDomain == 'FWA_PS':
                sBusinessLine = 'BBD-FWA'
            tblname = 'cdb_rd_resource'    
            stN = tbl_index(tblname,SQLConn)
            sTD = strnum(stN) 
            InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Phase,Year1,Modifier) VALUES 
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
            values = (sTD, sRelease, sRCR, sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,phase,sYear,sMail)
            # print('STD=',sTD,InUpSql, file=fa, flush=True )
            SQLConn.cur.execute(InUpSql, values)        
            InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """         
            YID = sYear + sTD
            values2 = (YID, sYear, sJans, sFebs,sMars, sAprs, sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs,Modifier)                           
            SQLConn.cur.execute(InUpSql2, values2)
            SQLConn.conn.commit()
            dResult['data']['status']="Add successful"  
        else:
            dResult['data']['status']="The data is exist, do not add again!"
            
    # 4 add more
    elif sType == '4':        
        cmd = """
                SELECT
                   count(RCR) as num                 
                FROM
                cdb_rd_resource a
                LEFT JOIN
                  cdb_rd_effort b
                ON  
                  a.ID = Right(b.ID,8) 
                  %s
                """
        if sYear == '':
            sYear = datetime.today().strftime('%Y')
        sRule = " WHERE Site = '%s' AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' AND Year = '%s' " % (sSite,sProductDomain,sRCR,sCompetence,sYear)
        # print(sSite,sProductDomain,sRCR,sCompetence,sYear, file=fa, flush=True )
        SQLConn = analyzer_db() 
        # print(cmd % sRule, file=fa, flush=True )
        SQLConn.cur.execute(cmd % sRule)
        SQLResult = SQLConn.cur.fetchall()
        count = SQLResult[0][0]    
        if count > 0 and sRelease != 'Common' and sRCR != 'PIPL (undefined RCR)' and sRCR != 'FD w/o RCR w/o PIPL':
            dResult['data']['status']="The first data is exist, do not add again!"
        else:
            if sProductDomain == 'ONT_PS' or sProductDomain == 'China Product' or sProductDomain == 'Voice' :
                sBusinessLine = 'BBD-ONT'
            elif sProductDomain == 'Cloud' or sProductDomain == 'Mobile App' or sProductDomain == 'Mesh' or sProductDomain == 'Container App' or sProductDomain == 'HomeOS':
                sBusinessLine = 'BBD-CSW'
            elif sProductDomain == 'Beacon':
                sBusinessLine = 'BBD-NWF'
            elif sProductDomain == 'FWA_PS':
                sBusinessLine = 'BBD-FWA'
            tblname = 'cdb_rd_resource'    
            stN = tbl_index(tblname,SQLConn)
            sID = strnum(stN) 
            InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Phase,Year1,Modifier) VALUES 
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
            values = (sID, sRelease, sRCR, sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,phase,sYear,sMail)
            # print('STD=',sTD,InUpSql, file=fa, flush=True )
            SQLConn.cur.execute(InUpSql, values)        
            InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """         
            YID = sYear + sID
            values2 = (YID, sYear, sJans, sFebs,sMars, sAprs, sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs,Modifier)                           
            SQLConn.cur.execute(InUpSql2, values2)
            SQLConn.conn.commit()
            if sRelease == 'Common' or sRCR == 'PIPL (undefined RCR)' or sRCR == 'FD w/o RCR w/o PIPL' :
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite8,sYear, sJans, sFebs,sMars,sAprs,sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs,sMail,sID,SQLConn)
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite2,sYear2, sJans2, sFebs2,sMars2,sAprs2,sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,sMail,sID,SQLConn)                    
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite3,sYear3, sJans3, sFebs3,sMars3,sAprs3,sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3,sMail,sID,SQLConn)
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite4,sYear4, sJans4, sFebs4,sMars4,sAprs4,sMays4,sJuns4,sJuls4,sAugs4,sSeps4,sOcts4,sNovs4,sDecs4,sMail,sID,SQLConn)
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite5,sYear5, sJans5, sFebs5,sMars5,sAprs5,sMays5,sJuns5,sJuls5,sAugs5,sSeps5,sOcts5,sNovs5,sDecs5,sMail,sID,SQLConn)
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite6,sYear6, sJans6, sFebs6,sMars6,sAprs6,sMays6,sJuns6,sJuls6,sAugs6,sSeps6,sOcts6,sNovs6,sDecs6,sMail,sID,SQLConn)
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite7,sYear7, sJans7, sFebs7,sMars7,sAprs7,sMays7,sJuns7,sJuls7,sAugs7,sSeps7,sOcts7,sNovs7,sDecs7,sMail,sID,SQLConn)
                create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite8,sYear8, sJans8, sFebs8,sMars8,sAprs8,sMays8,sJuns8,sJuls8,sAugs8,sSeps8,sOcts8,sNovs8,sDecs8,sMail,sID,SQLConn) 
                SQLConn.conn.close()
            else :
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite8,sYear, sJans, sFebs,sMars,sAprs,sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs,sMail,sID,SQLConn)
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite2,sYear2, sJans2, sFebs2,sMars2,sAprs2,sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,sMail,sID,SQLConn)                    
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite3,sYear3, sJans3, sFebs3,sMars3,sAprs3,sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3,sMail,sID,SQLConn)
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite4,sYear4, sJans4, sFebs4,sMars4,sAprs4,sMays4,sJuns4,sJuls4,sAugs4,sSeps4,sOcts4,sNovs4,sDecs4,sMail,sID,SQLConn)
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite5,sYear5, sJans5, sFebs5,sMars5,sAprs5,sMays5,sJuns5,sJuls5,sAugs5,sSeps5,sOcts5,sNovs5,sDecs5,sMail,sID,SQLConn)
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite6,sYear6, sJans6, sFebs6,sMars6,sAprs6,sMays6,sJuns6,sJuls6,sAugs6,sSeps6,sOcts6,sNovs6,sDecs6,sMail,sID,SQLConn)
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite7,sYear7, sJans7, sFebs7,sMars7,sAprs7,sMays7,sJuns7,sJuls7,sAugs7,sSeps7,sOcts7,sNovs7,sDecs7,sMail,sID,SQLConn)
                create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite8,sYear8, sJans8, sFebs8,sMars8,sAprs8,sMays8,sJuns8,sJuls8,sAugs8,sSeps8,sOcts8,sNovs8,sDecs8,sMail,sID,SQLConn) 
                SQLConn.conn.close()
            dResult['data']['status']="Add successful"     
    # 2 edit
    elif sType == '2':
        # print(sType, file=fa, flush=True )
        SQLConn = analyzer_db()
        if sProductDomain == 'ONT_PS' or sProductDomain == 'China Product' or sProductDomain == 'Voice' :
            sBusinessLine = 'BBD-ONT'
        elif sProductDomain == 'Cloud' or sProductDomain == 'Mobile App' or sProductDomain == 'Mesh' or sProductDomain == 'Container App' or sProductDomain == 'HomeOS':
            sBusinessLine = 'BBD-CSW'
        elif sProductDomain == 'Beacon':
            sBusinessLine = 'BBD-NWF'
        elif sProductDomain == 'FWA_PS':
            sBusinessLine = 'BBD-FWA'
        YID = sYear + sID    
        sql="""
            UPDATE cdb_rd_resource set Releases = '%s', BusinessLine= '%s', RCR = '%s',Description = '%s',
            ProjectNumber = '%s',ProjectDescription = '%s',Type = '%s',Site = '%s',
            Competence= '%s',RCRCategories = '%s',Phase  = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (sRelease, sBusinessLine, sRCR,sDescription, sProjectNumber, sProjectDescription,pType,sSite,sCompetence,sRCRCategories,phase,sMail, sID)
        sql1="""
            UPDATE cdb_rd_effort set Year = '%s', Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
            Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
            Novs= '%s',Decs = '%s',Modifier= '%s',ID = '%s'
            WHERE ID = '%s'           
            """ % (sYear, sJans, sFebs,sMars, sAprs, sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs, sMail,YID,sYID)
        
        print('update 2 rd=',sql,sMail,sLastupdate,file=fc, flush=True )
        print('update 2 ef=',sql1,sMail,sLastupdate,file=fc, flush=True )
        SQLConn.cur.execute(sql)        
        SQLConn.cur.execute(sql1)
        SQLConn.conn.commit()
        sql = """
           SELECT
              Year 
           FROM 
             cdb_rd_resource a
           JOIN
             cdb_rd_effort b
           ON
             a.ID = Right(b.ID,8)
           WHERE  
             a.ID = '%s'
           ORDER BY Year  
        """ % sID
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
        Yearlist =[]
        for row in SQLResult:
            Yearlist.append(row[0]) 
        print('Yearlist =',Yearlist,sLastupdate,file=fc, flush=True )
        if len(Yearlist) == 1:
            Year1 = Yearlist[0]
            Year2 = ''
            Year3 = ''
        elif len(Yearlist) == 2:
            Year1 = Yearlist[0]
            Year2 = Yearlist[1]
            Year3 = ''
        elif len(Yearlist) == 3:
            Year1 = Yearlist[0]
            Year2 = Yearlist[1]
            Year3 = Yearlist[2]
        sql="""
            UPDATE cdb_rd_resource set Year1 = '%s',Year2 = '%s',Year3 = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (Year1,Year2,Year3,sMail,sID)                        
        SQLConn.cur.execute(sql)       
        SQLConn.conn.commit()
        SQLConn.conn.close()  
        dResult['data']['status']="Edit successful" 
        
    # 22 edit data
    elif sType == '22':        
        SQLConn = analyzer_db()
        if sMonth == 'Jans':
            sql="""
                UPDATE cdb_rd_effort set Jans = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Febs':       
            sql="""
                UPDATE cdb_rd_effort set Febs = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Mars':       
            sql="""
                UPDATE cdb_rd_effort set Mars = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Aprs':       
            sql="""
                UPDATE cdb_rd_effort set Aprs = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Mays':       
            sql="""
                UPDATE cdb_rd_effort set Mays = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Juns':       
            sql="""
                UPDATE cdb_rd_effort set Juns = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Juls':       
            sql="""
                UPDATE cdb_rd_effort set Juls = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Augs':       
            sql="""
                UPDATE cdb_rd_effort set Augs = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Seps':       
            sql="""
                UPDATE cdb_rd_effort set Seps = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Octs':       
            sql="""
                UPDATE cdb_rd_effort set Octs = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID) 
        elif sMonth == 'Novs':       
            sql="""
                UPDATE cdb_rd_effort set Novs = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
        elif sMonth == 'Decs':       
            sql="""
                UPDATE cdb_rd_effort set Decs = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (sVal,sMail,sYID)
                
        print('update =',sql,sMail,sLastupdate,file=fc, flush=True )
        SQLConn.cur.execute(sql) 
        SQLConn.conn.commit() 
        SQLConn.conn.close()  
        dResult['data']['status']="data update successful" 
        
    # 7 edit month data
    elif sType == '7':
        
        SQLConn = analyzer_db()
        sql="""
            UPDATE cdb_rd_effort set Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
            Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
            Novs= '%s',Decs = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (sJans, sFebs,sMars, sAprs, sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs, sMail, sYID)
        print('update =',sql,sMail,sLastupdate,file=fc, flush=True )
        SQLConn.cur.execute(sql) 
        SQLConn.conn.commit() 
        SQLConn.conn.close()  
        dResult['data']['status']="data update successful" 
        
    # 5 edit more
    elif sType == '5':
        print('para 5 =',sRelease, sBusinessLine, sRCR,sDescription, sProjectNumber, sProjectDescription,pType,sSite,sCompetence,sRCRCategories,phase,sID,file=fc, flush=True )
        SQLConn = analyzer_db()
        if sProductDomain == 'ONT_PS' or sProductDomain == 'China Product' or sProductDomain == 'Voice' :
                sBusinessLine = 'BBD-ONT'
        elif sProductDomain == 'Cloud' or sProductDomain == 'Mobile App' or sProductDomain == 'Mesh' or sProductDomain == 'Container App' or sProductDomain == 'HomeOS':
            sBusinessLine = 'BBD-CSW'
        elif sProductDomain == 'Beacon':
            sBusinessLine = 'BBD-NWF'
        elif sProductDomain == 'FWA_PS':
            sBusinessLine = 'BBD-FWA'   
        YID = sYear + sID
        sql="""
            UPDATE cdb_rd_resource set Releases = '%s', BusinessLine= '%s', RCR = '%s',Description = '%s',
            ProjectNumber = '%s',ProjectDescription = '%s',Type = '%s',Site = '%s',
            Competence= '%s',RCRCategories = '%s',Phase = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (sRelease, sBusinessLine, sRCR,sDescription, sProjectNumber, str(sProjectDescription),pType,sSite,sCompetence,sRCRCategories,phase,sMail, sID)
        sql1="""
            UPDATE cdb_rd_effort set Year = '%s', Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
            Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
            Novs= '%s',Decs = '%s',Modifier= '%s',ID = '%s'
            WHERE ID = '%s'           
            """ % (sYear, sJans, sFebs,sMars, sAprs, sMays,sJuns,sJuls,sAugs,sSeps,sOcts,sNovs,sDecs, sMail,YID, sYID)
        print('update 5-0 rd=',sql, sID,file=fc, flush=True )
        try:
            SQLConn.cur.execute(sql)
        except:
            print('error 1=',sLastupdate,file=fc, flush=True )
        try:
            SQLConn.cur.execute(sql1)
        except:
            print('error 2=',sLastupdate,file=fc, flush=True ) 
        # SQLConn.cur.execute(sql)    
        # SQLConn.cur.execute(sql1)
        SQLConn.conn.commit()
        print('update 5 rd=',sql,sMail,sLastupdate,file=fc, flush=True )
        print('update 5 ef=',sql1,sMail,sLastupdate,file=fc, flush=True )
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite2,sYear2, sJans2, sFebs2,sMars2,sAprs2,sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,sMail,sID,SQLConn)                    
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite3,sYear3, sJans3, sFebs3,sMars3,sAprs3,sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3,sMail,sID,SQLConn)
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite4,sYear4, sJans4, sFebs4,sMars4,sAprs4,sMays4,sJuns4,sJuls4,sAugs4,sSeps4,sOcts4,sNovs4,sDecs4,sMail,sID,SQLConn)
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite5,sYear5, sJans5, sFebs5,sMars5,sAprs5,sMays5,sJuns5,sJuls5,sAugs5,sSeps5,sOcts5,sNovs5,sDecs5,sMail,sID,SQLConn)
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite6,sYear6, sJans6, sFebs6,sMars6,sAprs6,sMays6,sJuns6,sJuls6,sAugs6,sSeps6,sOcts6,sNovs6,sDecs6,sMail,sID,SQLConn)
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite7,sYear7, sJans7, sFebs7,sMars7,sAprs7,sMays7,sJuns7,sJuls7,sAugs7,sSeps7,sOcts7,sNovs7,sDecs7,sMail,sID,SQLConn)
        create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite8,sYear8, sJans8, sFebs8,sMars8,sAprs8,sMays8,sJuns8,sJuls8,sAugs8,sSeps8,sOcts8,sNovs8,sDecs8,sMail,sID,SQLConn)
        sql = """
           SELECT
              Year 
           FROM 
             cdb_rd_resource a
           JOIN
             cdb_rd_effort b
           ON
             a.ID = Right(b.ID,8)
           WHERE  
             a.ID = '%s'
           ORDER BY Year  
        """ % sID
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
        Yearlist =[]
        for row in SQLResult:
            Yearlist.append(row[0]) 
        
        if len(Yearlist) == 1:
            Year1 = Yearlist[0]
            Year2 = ''
            Year3 = ''
        elif len(Yearlist) == 2:
            Year1 = Yearlist[0]
            Year2 = Yearlist[1]
            Year3 = ''
        elif len(Yearlist) == 2:
            Year1 = Yearlist[0]
            Year2 = Yearlist[1]
            Year3 = Yearlist[2]
        sql="""
            UPDATE cdb_rd_resource set Year1 = '%s',Year2 = '%s',Year3 = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (Year1,Year2,Year3,sMail,sID)                        
        SQLConn.cur.execute(sql)       
        SQLConn.conn.commit()            
        SQLConn.conn.close()  
        dResult['data']['status']="Edit successful" 
    
    # 3 delete
    elif sType == '3':
        lYID = sYID.split(',')
        SQLConn = analyzer_db() 
        sql="DELETE FROM cdb_rd_effort WHERE ID IN %s" % List2String(lYID) 
        print('delete ef =',sql,sMail,sLastupdate,file=fb, flush=True )
        SQLConn.cur.execute(sql)
        SQLConn.conn.commit()
        for td in lYID:
            sql = """
               SELECT
                  count(RCR) as num 
               FROM 
                 cdb_rd_resource a
               JOIN
                 cdb_rd_effort b
               ON
                 a.ID = Right(b.ID,8)
               WHERE  
                 a.ID = '%s'
            """ % td[-8:]
            SQLConn.cur.execute(sql)      
            SQLResult = SQLConn.cur.fetchall()
            count = SQLResult[0][0]
            # print('count=',str(count),sql,file=fa, flush=True )
            if count == 0 : 
                sqlt="DELETE FROM cdb_rd_resource WHERE ID = '%s'" % td[-8:]
                print('delete rd=',sqlt,sMail,sLastupdate,file=fb, flush=True )
                SQLConn.cur.execute(sqlt)
            else:
                sql = """
                   SELECT
                      ID,Year1,Year2,Year3 
                   FROM 
                     cdb_rd_resource                    
                   WHERE  
                     ID = '%s'
                """ % td[-8:]
                SQLConn.cur.execute(sql)      
                SQLResult = SQLConn.cur.fetchall()
                for row in SQLResult:
                    Year1 = row[1]
                    Year2 = row[2]
                    Year3 = row[3]
                    if Year1 == td[:4]:
                        Year1 =''
                    elif Year2 == td[:4]:
                        Year2 =''
                    elif Year3 == td[:4]:
                        Year3 =''
                sql="""
                    UPDATE cdb_rd_resource set Year1 = '%s',Year2 = '%s',Year3 = '%s',Modifier= '%s'
                    WHERE ID = '%s'           
                    """ % (Year1,Year2,Year3,sMail,td[-8:])                        
                SQLConn.cur.execute(sql)       
            SQLConn.conn.commit()                
                
         
        SQLConn.conn.close()  
        dResult['data']['status']="Delete successful"
        
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')

def create_new(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite2,sYear2, sJans2, sFebs2,sMars2,sAprs2,sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,sMail,sID,SQLConn):
    
    sql = """
            SELECT                    
             ID,Year1,Year2,Year3
            FROM
            cdb_rd_resource            
            WHERE  
             ID = '%s' 
            """  % sID
    SQLConn.cur.execute(sql)      
    SQLResult = SQLConn.cur.fetchall()
    for row in SQLResult:
        sY1 = row[1]
        sY2 = row[2]
        sY3 = row[3]
    print('site=',sID,sBusinessLine,sProductDomain,pType,sCompetence,sSite2,sYear,sYear2, sY1,sY2,sY3,file=fa, flush=True )    
    if sSite2 == sSite and sYear2 !='' and sYear2 != sYear  :
        sql = """
                SELECT                    
                  count(RCR) as num
                FROM
                cdb_rd_resource a
                JOIN
                  cdb_rd_effort b
                WHERE  
                  a.ID = Right(b.ID,8) 
                  AND Year = '%s' AND a.ID = '%s'
                """  % (sYear2,sID)
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
        count = SQLResult[0][0]            
        if count > 0 :
            YID = sYear2 + sID
            sql="""
            UPDATE cdb_rd_effort set Year = '%s', Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
            Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
            Novs= '%s',Decs = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail,YID) 
            print('N1=',sql,sYear,sYear2, file=fa, flush=True )
            SQLConn.cur.execute(sql) 
            SQLConn.conn.commit()
        else:
            year_list =[]
            if sY1 !='':
                year_list.append(sY1)
            if sY2 !='':
                year_list.append(sY2)
            if sY3 !='':
                year_list.append(sY3)  
            year_list.append(sYear2)
            year_list.sort(key=lambda x: x)
            if len(year_list) > 3:
                year_list.remove(year_list[0])
            Y1 = year_list[0]
            Y2 = year_list[1]
            Y3 = ''
            if len(year_list) > 2 :
                Y3=year_list[2]                       
            InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
            YID2 = sYear2 + sID
            values = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail)
            print('Nytd2=',YID2, file=fa, flush=True )
            SQLConn.cur.execute(InUpSql2, values)
            
            sql="""
                UPDATE cdb_rd_resource set Year1 = '%s', Year2 = '%s',Year3 = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (Y1,Y2,Y3,sMail,sID)
            SQLConn.cur.execute(sql)
            SQLConn.conn.commit()
              
    elif sSite2 != sSite and sSite2 != '' and sYear2 !='':       
        tblname = 'cdb_rd_resource'    
        stN = tbl_index(tblname,SQLConn)
        sTD = strnum(stN)
        InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Phase,Effortjira,Year1,Modifier) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        values = (sTD, sRelease, sRCR, sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite2,sCompetence,sRCRCategories,phase,sEffort,sYear2, sMail)
        # print('STD=',sTD,InUpSql, file=fa, flush=True )
        SQLConn.cur.execute(InUpSql, values)
        
        InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
        YID2 = sYear2 + sTD
        values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail)
        print('N3=',YID2, file=fa, flush=True )                    
        SQLConn.cur.execute(InUpSql2, values2)
        SQLConn.conn.commit()

def create_item(sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sCompetence,sRCRCategories,phase,sEffort,sSite,sYear,sSite2,sYear2, sJans2, sFebs2,sMars2,sAprs2,sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,sMail,sID,SQLConn):
    # print('site=',sSite2,sYear,sYear2, file=fa, flush=True )
    sql = """
            SELECT                    
              b.ID,a.ID,Year1,Year2,Year3,Year
            FROM
            cdb_rd_resource a
            JOIN
              cdb_rd_effort b
            WHERE  
              a.ID = Right(b.ID,8) 
              AND Site = '%s' AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' 
              AND Type = '%s' 
            """  % (sSite,sProductDomain,sRCR,sCompetence,pType)
    SQLConn.cur.execute(sql)      
    SQLResult = SQLConn.cur.fetchall()
    for row in SQLResult:
        sY1 = row[2]
        sY2 = row[3]
        sY3 = row[4]
    if sSite2 == sSite and sYear2 !='' and sYear2 != sYear  :
        sql = """
                SELECT                    
                  count(RCR) as num
                FROM
                cdb_rd_resource a
                JOIN
                  cdb_rd_effort b
                WHERE  
                  a.ID = Right(b.ID,8) 
                  AND Year = '%s' AND a.ID = '%s'
                """  % (sYear2,sID)
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
        count = SQLResult[0][0]            
        if count > 0 :
            YID = sYear2 + sID
            sql="""
            UPDATE cdb_rd_effort set Year = '%s', Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
            Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
            Novs= '%s',Decs = '%s',Modifier= '%s'
            WHERE ID = '%s'           
            """ % (sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail,YID) 
            print('A1=',sql,sYear,sYear2, file=fc, flush=True )
            SQLConn.cur.execute(sql) 
            SQLConn.conn.commit()
        else:
            year_list =[]
            if sY1 !='':
                year_list.append(sY1)
            if sY2 !='':
                year_list.append(sY2)
            if sY3 !='':
                year_list.append(sY3)  
            year_list.append(sYear2)
            year_list.sort(key=lambda x: x)
            if len(year_list) > 3:
                year_list.remove(year_list[0])
            Y1 = year_list[0]
            Y2 = year_list[1]
            Y3 = ''
            if len(year_list) > 2 :
                Y3=year_list[2]                       
            InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
            YID2 = sYear2 + sID
            values = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail)
            print('ytd2=',YID2, file=fa, flush=True )
            SQLConn.cur.execute(InUpSql2, values)
            
            sql="""
                UPDATE cdb_rd_resource set Year1 = '%s', Year2 = '%s',Year3 = '%s',Modifier= '%s'
                WHERE ID = '%s'           
                """ % (Y1,Y2,Y3,sMail,sID)
            SQLConn.cur.execute(sql)
            SQLConn.conn.commit()
              
    elif sSite2 != sSite and sSite2 != '' and sYear2 !='':
            
            sql = """
                    SELECT                    
                      count(RCR) as num
                    FROM
                    cdb_rd_resource a
                    JOIN
                      cdb_rd_effort b
                    WHERE  
                      a.ID = Right(b.ID,8) 
                      AND Site = '%s' AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' 
                      AND Type = '%s' 
                    """  % (sSite2,sProductDomain,sRCR,sCompetence,pType)
            SQLConn.cur.execute(sql)      
            SQLResult = SQLConn.cur.fetchall()
            count = SQLResult[0][0]            
            # print('count=',str(count),sSite,sSite2,sYear2,file=fa, flush=True )
            if count == 0 :
                tblname = 'cdb_rd_resource'    
                stN = tbl_index(tblname,SQLConn)
                sTD = strnum(stN)
                InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Phase,Effortjira,Year1,Modifier) VALUES 
                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                values = (sTD, sRelease, sRCR, sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite2,sCompetence,sRCRCategories,phase,sEffort,sYear2, sMail)
                # print('STD=',sTD,InUpSql, file=fa, flush=True )
                SQLConn.cur.execute(InUpSql, values)
                
                InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                YID2 = sYear2 + sTD
                values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail)
                # print('A3=',YID2, file=fa, flush=True )                    
                SQLConn.cur.execute(InUpSql2, values2)
                SQLConn.conn.commit()
            
            else:
                sql = """
                        SELECT                    
                          b.ID,a.ID,Year1,Year2,Year3,Year
                        FROM
                        cdb_rd_resource a
                        JOIN
                          cdb_rd_effort b
                        WHERE  
                          a.ID = Right(b.ID,8) 
                          AND Site = '%s' AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' 
                          AND Type = '%s' 
                        """  % (sSite2,sProductDomain,sRCR,sCompetence,pType)
                SQLConn.cur.execute(sql)      
                SQLResult = SQLConn.cur.fetchall()
                flag = 0                    
                for row in SQLResult:
                    # print('A5=',row[0],row[1], file=fa, flush=True ) 
                    sID2 = row[1] 
                    year_list=[]
                    year_list.append(row[2])
                    if row[3] !='':
                        year_list.append(row[3])
                    if row[4] !='':    
                        year_list.append(row[4])
                    if sYear2 in year_list and sYear2 == row[5]:                           
                        flag = 1
                        YID2 = sYear2 + sID2                       
                        sql2="""
                               UPDATE cdb_rd_effort set Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
                                Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
                                Novs= '%s',Decs = '%s',Modifier= '%s'
                                WHERE ID = '%s'           
                            """ % (sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail,YID2)                                   
                        # print('A4=',YID2, file=fa, flush=True ) 
                        SQLConn.cur.execute(sql2)
                        SQLConn.conn.commit()
                        break
                    
                if flag == 0:                        
                    year_list.append(sYear2)    
                    year_list.sort(key=lambda x: x)                        
                    if len(year_list) > 3:
                        year_list.remove(year_list[0])
                    Y1 = year_list[0]
                    Y2 = year_list[1]
                    if len(year_list) == 2:                        
                        Y3 = ''
                    else:
                        Y3 = year_list[2]
                    InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                    YID2 = sYear2 + sID2
                    print('A6=',YID2,year_list, file=fa, flush=True ) 
                    
                    values = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2, sMail)
                    SQLConn.cur.execute(InUpSql2, values)
                    
                    sql="""
                        UPDATE cdb_rd_resource set Year1 = '%s', Year2 = '%s',Year3 = '%s',Modifier= '%s'
                        WHERE ID = '%s'           
                        """ % (Y1,Y2,Y3,sMail,sID2)
                    SQLConn.cur.execute(sql)
                    SQLConn.conn.commit()
                     

def rd_resource_update(request):
    ud = request.get_full_path()    
    try:
        sRelease = request.GET['release']
        DueDate = datetime.today() + timedelta(weeks=8)
        sYear = str(DueDate)[:4]        
        # sYear = datetime.today().strftime('%Y')
        # sYear ='2024'
        sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
        print('Sync =', ud,sLastupdate,file=fa,flush=True )         
    except:
        return HttpResponse('Invalid username or password', content_type='application/json')
    cmd = """
          SELECT 
            FixVersions 
          FROM 
            jira_issues_rcr 
          WHERE 
               Left(FixVersions,5) ='BBDR2' 
              OR Left(FixVersions,5) ='BBDR3' 
             
         ORDER BY FixVersions DESC
         """
    SQLBConn = pymysql.connect(host  = settings.BBD_DB['host'],
                            port     = settings.BBD_DB['port'],
                            user     = settings.BBD_DB['username'],
                            password = settings.BBD_DB['password'],
                            database = settings.BBD_DB['name'],
                            charset  = settings.BBD_DB['charset']
                        )
    SQLBCur = SQLBConn.cursor()
    SQLConn = analyzer_db()
    if sRelease == 'All' :      
        SQLBCur.execute(cmd)                
        SQLBResult = SQLBCur.fetchall()
        rellist =[]
        for row in SQLBResult:        
            relv = row[0].split(',')        
            for r in relv:
                if int(r[4:6]) >= int(sYear[2:4]):
                    rellist.append(r)   
        relset = set(rellist)
        releaselist = sorted(list(relset))
        num1 = 0
        num2 = 0
        num3 = 0
        num4 = 0
        rel_list =[]
         
        for rel in releaselist:
            result = rd_update(rel,sYear,SQLBConn,SQLConn)
            num1 += int(result[0])
            num2 += int(result[1])
            num3 += int(result[2])
            num4 += int(result[3])
            if num1 + num2 + num3 + num4 >0:
                rel_list.append(rel)
        new_mum = str(num1)
        up_num = str(num2)
        del_num = str(num3)
        ins_num = str(num4)
    elif sRelease == 'Category':
        result = rd_update_cat(sRelease,SQLBConn,SQLConn)
        new_mum = 0
        up_num = result[0]
        del_num = 0
        ins_num = 0
        rel_list = sRelease
    else:
        result = rd_update(sRelease,sYear,SQLBConn,SQLConn)        
        new_mum = result[0]
        up_num = result[1]
        del_num = result[2]
        ins_num = result[3]
        # result = rd_update_cat2(sRelease,SQLConn)
        # new_mum = '0'
        # up_num = '0'
        # del_num = result[0]
        rel_list = sRelease
    
    SQLBConn.close()
    SQLConn.close()    
    dResult = {}        
    dResult['data'] = {}
    dResult['data']['items'] = []    
    dItem = {}
    dItem['Result'] = 'Update successful'    
    dItem['New RCR'] = new_mum
    dItem['Update RCR'] = up_num
    dItem['Insert RCR'] = ins_num
    dItem['Delete RCR'] = del_num
    dItem['Insert RCR'] = ins_num
    if len(rel_list) > 0:
        dItem['Release'] = rel_list
    else:
        dItem['Release'] = 'All'
    dResult['data']['items'].append(dItem)
    
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')


def rd_update_cat2(sRelease,SQLConn):
    cmd = """
        SELECT
        ID,Effortjira,SubTasks,Siteallocation,RCR,ProductDomain,
        Competence,Type,Site,Year1,Year2,Year3,Releases,RCRCategories                
        FROM
        cdb_rd_resource
        WHERE 
        
        Releases ='%s' AND Left(RCR,7)='BBDPROD' 

        """  % sRelease   
        # RCR LIKE CONCAT('%','BBDPROD','%')
      
    SQLConn.cur.execute(cmd)             
    SQLResult = SQLConn.cur.fetchall()
    num_del = 0 
    rcr_list=[]
    lYID =[]
    for row in SQLResult:
       cst =  rd_zero(row[0],SQLConn)
       if cst[0]:
          rcr_list.append(cst[2])
          lYID.append(cst[3])
          # print('RD zero =',cst[1],cst[2],cst[3],row[4],row[0],row[8],row[1],row[2],row[3],row[9],row[10],row[11],file=fa, flush=True ) 
          num_del += 1 
    print('RD zero =',rcr_list,lYID,file=fd, flush=True )
    # delete
    
    # lYID = sYID.split(',')
    SQLConn = analyzer_db() 
    sql="DELETE FROM cdb_rd_effort WHERE ID IN %s" % List2String(lYID) 
    print('delete LIST =',sql,file=fd, flush=True )
    SQLConn.cur.execute(sql)
    SQLConn.conn.commit()
    for td in lYID:
        sql = """
           SELECT
              count(RCR) as num 
           FROM 
             cdb_rd_resource a
           JOIN
             cdb_rd_effort b
           ON
             a.ID = Right(b.ID,8)
           WHERE  
             a.ID = '%s'
        """ % td[-8:]
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
        count = SQLResult[0][0]
        # print('count=',str(count),sql,file=fd, flush=True )
        if count == 0 : 
            sqlt="DELETE FROM cdb_rd_resource WHERE ID = '%s'" % td[-8:]
            # print('delete rd=',sqlt,file=fd, flush=True )
            SQLConn.cur.execute(sqlt)
        else:
            sql = """
               SELECT
                  ID,Year1,Year2,Year3 
               FROM 
                 cdb_rd_resource                    
               WHERE  
                 ID = '%s'
            """ % td[-8:]
            SQLConn.cur.execute(sql)      
            SQLResult = SQLConn.cur.fetchall()
            for row in SQLResult:
                Year1 = row[1]
                Year2 = row[2]
                Year3 = row[3]
                if Year1 == td[:4]:
                    Year1 =''
                elif Year2 == td[:4]:
                    Year2 =''
                elif Year3 == td[:4]:
                    Year3 =''
            sql="""
                UPDATE cdb_rd_resource set Year1 = '%s',Year2 = '%s',Year3 = '%s'
                WHERE ID = '%s'           
                """ % (Year1,Year2,Year3,td[-8:])                        
            SQLConn.cur.execute(sql)       
        SQLConn.conn.commit() 
              
    return str(num_del) 

     
def rd_update_cat(sRelease,SQLBConn,SQLConn):
    SQLB_num = """
        SELECT
        count(Status) as num        
        FROM
        jira_issues_rcr        
        """    
        
    SQLB = """
        SELECT
        `Key`,PortfolioTypes,BusinessRationale,FeatureCategory,BusinessPriority1       
        FROM
        jira_issues_rcr
        
       """
    SQLB_pipl = """
    SELECT
    `Key`,Summary      
    FROM
    jira_issues_bbdpipl           
    """
    SQLBCur = SQLBConn.cursor()
    SQLBCur.execute(SQLB_num)          
    SQLResult = SQLBCur.fetchall()            
    count = SQLResult[0][0]
    
    SQLBCur.execute(SQLB)             
    SQLBResult = SQLBCur.fetchall()
    
    SQLBCur.execute(SQLB_pipl)             
    SQLBResult_pipl = SQLBCur.fetchall()
    rowvalues =[]                  
    
    RCR = [row[0] for row in SQLBResult]    
    RCRCategories = [row[1] for row in SQLBResult]    
    ProjectNumber = [prjnum(row[2]) for row in SQLBResult]   
    ProjectNumber2 = [row[2] for row in SQLBResult]
    FeatureCategory = [row[3] for row in SQLBResult]
    BusinessPriority = [row[4].replace('.0','') for row in SQLBResult]
    cmd = """
        SELECT
        ID,Effortjira,SubTasks,Siteallocation,RCR,ProductDomain,
        Competence,Type,Site,Year1,Year2,Year3,Releases,RCRCategories,ProjectNumber                
        FROM
        cdb_rd_resource
        WHERE 
        RCR LIKE CONCAT('%','BBDPROD','%') 
        OR RCR LIKE CONCAT('%','PIPL','%') 
        
        
        """     
    SQLConn.cur.execute(cmd)             
    SQLResult = SQLConn.cur.fetchall()
    num_up = 0    
    for row in SQLResult:
        for i in range(count):
           if RCR[i] == row[4] :
               num_up += 1                
               Project = ''                      
               if  'BBDPIPL' in ProjectNumber[i]:
                   for row_p in SQLBResult_pipl:
                       if row_p[0] == ProjectNumber[i]:
                           Project = repspecial(row_p[1][:1024].replace("'","\'"))                           
                           break
               if Project == '':
                   Project = prjdes(ProjectNumber[i])         
               try:
                  Sql = """UPDATE cdb_rd_resource SET ProjectNumber= '%s',ProjectNumber2= '%s', ProjectDescription= '%s',
                   ,FeatureCategory= %s,BusinessPriority = %s
                   WHERE RCR = '%s' """ % (ProjectNumber[i],ProjectNumber2[i][:100],Project,FeatureCategory[i].strip(),BusinessPriority[i].strip(),RCR[i])
                  SQLConn.cur.execute(Sql)
                  SQLConn.commit()                
                 
               except:
                    print('pipl error =',row[4],ProjectNumber[i],ProjectNumber2[i],Project[:30],file=fa, flush=True )
                   
               rowvalues.append((RCRCategories[i].strip(),FeatureCategory[i].strip(),BusinessPriority[i].strip(),RCR[i])) 
                   
        if 'PIPL' in row[4]:
            Project = ''                      
            if  'BBDPIPL' in row[14]:
                for row_p in SQLBResult_pipl:
                    if row_p[0] == row[14]:
                        Project = repspecial(row_p[1][:1024].replace("'","\'"))                           
                        try:
                          Sql = """UPDATE cdb_rd_resource SET ProjectDescription= '%s'
                           WHERE ID = '%s' """ % (Project,row[0])
                          SQLConn.cur.execute(Sql)
                          SQLConn.commit()                
                          # print('pipl prj=',row[0],row[4],row[14],row_p[1][:30],file=fa, flush=True )
                        except:
                            print('pipl error =',row[4],Project[:30],file=fa, flush=True )
                        break
                        
    InUpSql ="""UPDATE cdb_rd_resource SET RCRCategories = %s,FeatureCategory= %s,BusinessPriority = %s
      WHERE RCR=%s"""    
    SQLConn.cur.executemany(InUpSql,rowvalues)              
    SQLConn.commit()
    
    PDomain1 = 'ONT_PS'
    PDomain2 = 'China Product'
    PDomain3 = 'Voice'            
    PDomain4 = 'Cloud'
    PDomain5 = 'Mobile App'
    PDomain6 = 'Mesh'
    PDomain7 = 'Container App'
    PDomain8 = 'HomeOS'
    PDomain9 = 'Beacon'
    PDomain10 = 'FWA_PS'
    
    BL1 = 'BBD-ONT'
    BL2 = 'BBD-NWF'
    BL3 = 'BBD-FWA'
    BL4 = 'BBD-CSW'
    Sql = """UPDATE cdb_rd_resource SET Businessline = '%s'
     WHERE  ProductDomain = '%s' OR ProductDomain = '%s' OR ProductDomain = '%s'""" % (BL1,PDomain1,PDomain2,PDomain3)
    SQLConn.cur.execute(Sql)
    SQLConn.commit() 
    Sql = """UPDATE cdb_rd_resource SET Businessline = '%s'
     WHERE  ProductDomain = '%s' """ % (BL2,PDomain9)
    SQLConn.cur.execute(Sql)
    SQLConn.commit() 
    Sql = """UPDATE cdb_rd_resource SET Businessline = '%s'
     WHERE  ProductDomain = '%s' """ % (BL3,PDomain10)
    SQLConn.cur.execute(Sql)
    SQLConn.commit() 
    Sql = """UPDATE cdb_rd_resource SET Businessline = '%s'
     WHERE  ProductDomain = '%s' OR ProductDomain = '%s' OR ProductDomain = '%s' OR ProductDomain = '%s' OR ProductDomain = '%s' """ % (BL4,PDomain4,PDomain5,PDomain6,PDomain7,PDomain8)
    SQLConn.cur.execute(Sql)
    SQLConn.commit()
        
    return str(num_up)

def rd_zero(sID,SQLConn): 
   
    sql = """
            SELECT                    
              Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Year1,Year2,Year3,RCR,b.ID
            FROM
              cdb_rd_resource a
            JOIN
              cdb_rd_effort b
            WHERE  
              a.ID = Right(b.ID,8) 
              AND a.ID = '%s' 
                """  % (sID)
    SQLConn.cur.execute(sql)      
    SQLResult = SQLConn.cur.fetchall()
    flag = 0
    for row in SQLResult:
        sY1 = row[13]
        sY2 = row[14]
        sY3 = row[15]
        Yave = 0 
        RCR =row[16]
        YID = row[17]
        if row[0] == sY1 or row[0] == sY2 or row[0] == sY3:            
            if row[1] !='':
               ave1 = float(row[1])
               Yave = Yave + ave1
            if row[2] !='':
               ave2 = float(row[2])
               Yave = Yave + ave2
            if row[3] !='':
               ave3 = float(row[3])
               Yave = Yave + ave3
            if row[4] !='':
               ave4 = float(row[4])
               Yave = Yave + ave4
            if row[5] !='':
               ave5 = float(row[5])
               Yave = Yave + ave5
            if row[6] !='':
               ave6 = float(row[6]) 
               Yave = Yave + ave6
            if row[7] !='':
               ave7 = float(row[7])
               Yave = Yave + ave7
            if row[8] !='':
               ave8 = float(row[8])
               Yave = Yave + ave8
            if row[9] !='':
               ave9 = float(row[9])
               Yave = Yave + ave9
            if row[10] !='':
               ave10 = float(row[10])
               Yave = Yave + ave10
            if row[11] !='':
               ave11 = float(row[11])
               Yave = Yave + ave11
            if row[12] !='':
               ave12 = float(row[12])
               Yave = Yave + ave12            
            # print('ave =',str("%.2f" % (Yave)),row[0],sY1,sY2,sY1,row[1],row[2], file=fa, flush=True )
            if Yave > 0:
                flag = 1
                break
        else:
           print('year not match =',sID,row[16],row[0],sY1,sY2,sY3,row[1],row[2], file=fa, flush=True )
    if flag == 0:
        return True,str(Yave),RCR,YID
    else:
        return False,str(Yave),RCR,YID

    
def rd_update(sRelease,sYear,SQLBConn,SQLConn):
    cmd = """
        SELECT
        ID,Effortjira,SubTasks,Siteallocation,RCR,ProductDomain,
        Competence,Type,Site,Year1,Year2,Year3,Releases,RCRCategories,TechnicalAreas,Phase                
        FROM
        cdb_rd_resource
        WHERE 
           Releases = "%s" OR Left(Releases,4)='BBDP' 
                            
             """ % sRelease
         
    SQLConn.cur.execute(cmd)             
    SQLResult = SQLConn.cur.fetchall()    
    rcr_rd=[]
    num_rcr =0
    for row in SQLResult:
        rcr_rd.append(row[4])
        num_rcr +=1
    rcrset = set(rcr_rd)
    rcr_all = sorted(list(rcrset)) 
    print('S1 =',sRelease,num_rcr,file=fd, flush=True )    
    
    if num_rcr > 0 :
        SQLB_num = """
            SELECT
            count(Status) as num        
            FROM
            jira_issues_rcr
            WHERE 
            `Key` in %s 
            """ % List2String(rcr_all)    
            
        SQLB = """
            SELECT
            `Key`,Summary,Status, BusinessLine, FixVersions,SubTasks,
            TotalFeatureTeamEffort,HWEffort,Site,BusinessRationale,ProductLine,PortfolioTypes,
            FeatureCategory,BusinessPriority1            
            FROM
            jira_issues_rcr
            WHERE
              `Key` in %s 
            """ % List2String(rcr_all)
        SQLB_task = """
            SELECT
            `Key`,TechnicalArea,TimeOriginalEstimate,BusinessLine      
            FROM
            jira_issues
            WHERE `Type` = 'Sub-task' AND Status <> 'Rejected'   
            """
        SQLB_pipl = """
            SELECT
            `Key`,Summary      
            FROM
            jira_issues_bbdpipl           
            """
        SQLBCur = SQLBConn.cursor()
        SQLBCur.execute(SQLB_num)          
        SQLBResult = SQLBCur.fetchall()            
        count = SQLBResult[0][0]
        
        SQLBCur.execute(SQLB)             
        SQLBResult = SQLBCur.fetchall()
        
        SQLBCur.execute(SQLB_task)             
        SQLBResult_task = SQLBCur.fetchall()
        
        SQLBCur.execute(SQLB_pipl)             
        SQLBResult_pipl = SQLBCur.fetchall()
                    
        rowvalues =[]
        rowvalues1 =[]                   
        Releases = [row[4] for row in SQLBResult]
        RCR = [row[0] for row in SQLBResult]
        Description  = [row[1].replace("'","\\'") for row in SQLBResult]
        State = [row[2] for row in SQLBResult] 
        Site_rcr = [row[8] for row in SQLBResult]
        ProjectNumber = [prjnum(row[9]) for row in SQLBResult]  
        ProjectNumber2 = [row[9] for row in SQLBResult]
        ProjectDescription = ['' for row in SQLBResult]
        ProductDomain = [row[10] for row in SQLBResult]
        RCRCategories = [row[11] for row in SQLBResult]
        FeatureCategory = [row[12] for row in SQLBResult]
        BusinessPriority = [row[13].replace('.0','') for row in SQLBResult]
        SubTasks = [row[5][:1024] for row in SQLBResult]
        Businessline = [row[3] for row in SQLBResult]
        Effortjira = [row[6] for row in SQLBResult]
        SEffort = [row[6] for row in SQLBResult]
        HEffort = [row[7] for row in SQLBResult]
        num_new = 0
        num_up = 0
        num_del = 0  
                     
        for i in range(count):
            
            Phase = phase_type(RCR[i])
            Project = ProjectDescription[i]        
            if  'BBDPIPL' in ProjectNumber[i]:
                for row in SQLBResult_pipl:
                    if row[0] == ProjectNumber[i]:
                        Project = row[1].replace("'","\\'")
                        break
            if Project == '':
                Project = prjdes(ProjectNumber[i])                   
            rowvalues.append((Releases[i],Description[i].strip(),State[i].strip(),ProjectNumber[i],ProjectNumber2[i][:100],Project,RCRCategories[i].strip(),Phase,FeatureCategory[i].strip(),BusinessPriority[i].strip(),sRelease,RCR[i]))
            
        InUpSql ="""UPDATE cdb_rd_resource SET Releases=%s,Description=%s,State=%s,ProjectNumber=%s,ProjectNumber2=%s,
             ProjectDescription=%s,RCRCategories = %s,Phase = %s,FeatureCategory= %s,BusinessPriority= %s
         WHERE Releases=%s AND RCR=%s"""
       
        SQLConn.cur.executemany(InUpSql,rowvalues) 
        SQLConn.commit() 
        cmd = """
            SELECT
            ID,RCR,ProductDomain,Releases,Businessline,TechnicalAreas,SubTasks,Competence                
            FROM
            cdb_rd_resource
            WHERE 
               Releases = "%s" 
                                
                 """ % sRelease 
        SQLConn.cur.execute(cmd)             
        SQLResult = SQLConn.cur.fetchall() 
        for row in  SQLResult:
            ID = row[0]
            ProductDomain = row[2]
            Businessline = row[4]
            CP = row[7]
            RCR = row[1]
            ts = row[5].split(',')
            st = row[6].split(',')
            # if RCR == 'BBDPROD-63732':
            #     print('RCR item =',row[0],RCR,ts,st,file=fa, flush=True )
            
            task =''
            if ts[0] !='':
                BL=''
                for row_tsk in SQLBResult_task:
                    if row_tsk[0] == st[0] and st[0] !='':
                       BL =row_tsk[3]
                       task = row_tsk[0]
                Apool = pool(ts[0],BL)            
                 
                # print('TA BL NOW =',row[0],row[1],row[5],row[2],row[4],'--',Apool[0],Apool[1], file=fa, flush=True )     
                if (Apool[0] != row[2] or Apool[1] != row[4]) and task !='' and task == st[0] :                   
                   Sql= """UPDATE cdb_rd_resource SET ProductDomain = '%s',Businessline = '%s'
                    WHERE  ID = '%s' """ % (Apool[0],Apool[1],row[0])
                   num_up += 1
                   SQLConn.cur.execute(Sql) 
                   SQLConn.commit() 
                   print('TA BL change =',row[0],row[1],row[5],row[2],row[4],'--',Apool[0],Apool[1], file=fa, flush=True )
                elif st[0] =='' : 
                   for row_b in SQLBResult:
                       if RCR == row_b[0]:
                           task = row_b[5]
                           break
                   Sql= """UPDATE cdb_rd_resource SET ProductDomain = '%s',Businessline = '%s', SubTasks = '%s'
                    WHERE  ID = '%s' """ % (Apool[0],Apool[1],task,row[0])
                   num_up += 1
                   SQLConn.cur.execute(Sql) 
                   SQLConn.commit() 
                   print('TA BL TK change =',row[0],row[1],row[5],row[2],row[4],'--',Apool[0],Apool[1],task, file=fa, flush=True )   
            elif Businessline == '':
                for row_rcr in SQLBResult:
                    if RCR == row_rcr[0]:
                        BL = row_rcr[3]
                        Sql= """UPDATE cdb_rd_resource SET Businessline = '%s'
                         WHERE  ID = '%s' """ % (BL,ID)
                        num_up += 1
                        SQLConn.cur.execute(Sql) 
                        SQLConn.commit()
                        break
            elif ProductDomain == '':                
                for row_rcr in SQLBResult:
                    if RCR == row_rcr[0]:
                        PL = row_rcr[10]
                        if PL == 'PON ONTs':
                            ProductDomain = 'ONT_PS' 
                        elif PL == 'Beacon':
                            ProductDomain = 'Beacon'
                        elif PL == 'Mesh':
                            ProductDomain = 'Mesh'
                        elif PL == 'Cloud':
                            ProductDomain = 'Cloud' 
                        elif PL == 'Container App':
                            ProductDomain = 'Container App'
                        elif PL == 'mobile App' or PL == 'Mobile App':
                            ProductDomain = 'Mobile App'
                        elif PL == 'FWA-4G'or PL == 'FWA-FastMile-4G':
                            ProductDomain = 'FWA_PS'
                        elif PL == 'FWA-5G' or PL == 'FWA-FastMile-5G':
                            ProductDomain = 'FWA_PS'
                
                        Sql= """UPDATE cdb_rd_resource SET ProductDomain = '%s'
                         WHERE  ID = '%s' """ % (ProductDomain,ID)
                        num_up += 1
                        SQLConn.cur.execute(Sql) 
                        SQLConn.commit()
                        break
                
    # Create new RCR=========================================================================
    cmd = """
            SELECT
            RCR, Releases           
            FROM
            cdb_rd_resource            
            WHERE  
              Left(Releases,5) ='BBDR2' 
              OR Left(Releases,5) ='BBDR3' 
            ORDER BY RCR
            """ 
   
    SQLB_rel = """
        SELECT
        `Key`,Summary,Status, BusinessLine, FixVersions,SubTasks,
        TotalFeatureTeamEffort,HWEffort,Site,BusinessRationale,ProductLine,PortfolioTypes,
        FeatureCategory,BusinessPriority1
        FROM
        jira_issues_rcr
        WHERE
          FixVersions = '%s'
        """ % sRelease
    SQLConn.cur.execute(cmd)
    SQLResult_r = SQLConn.cur.fetchall()
    
    SQLBCur.execute(SQLB_rel)             
    SQLBResult = SQLBCur.fetchall()     
    rcr_rd=[]
    for row_r in SQLResult_r:        
        rcr_rd.append(row_r[0])
    rcrset = set(rcr_rd)
    rcr_all = sorted(list(rcrset))    
    rcr_list=[]
    count = 0
    for row in SQLBResult:
        if row[0] not in rcr_all:
            rcr_list.append(row[0])
            count += 1
    print('new RCR count =',str(count),rcr_list,file=fa, flush=True ) 
    num_add = count       
    if count > 0 :        
        SQLB = """
            SELECT
            `Key`,Summary,Status, BusinessLine, FixVersions,SubTasks,
            TotalFeatureTeamEffort,HWEffort,Site,BusinessRationale,ProductLine,PortfolioTypes,
            FeatureCategory,BusinessPriority1
            FROM
            jira_issues_rcr
            WHERE FixVersions = '%s' and `Key` IN %s   
           """ % (sRelease,List2String(rcr_list))
        print('stN =',List2String(rcr_list), file=fa, flush=True ) 
        SQLBCur.execute(SQLB)             
        SQLBResult = SQLBCur.fetchall()
        SQLBConn.close()
        tblname = 'cdb_rd_resource'    
        stN = tbl_index(tblname,SQLConn)
        print('stN 4 =',str(stN),rcr_list, file=fa, flush=True )               
        rowvalues =[]
        rowvalues1 =[]
        
        ID = [strnum(stN+i) for i in range(count)]            
        Releases = [row[4] for row in SQLBResult]
        RCR = [row[0] for row in SQLBResult]
        Description  = [row[1].replace("'","\\'") for row in SQLBResult]
        State = [row[2] for row in SQLBResult]
        ProjectNumber = [prjnum(row[9]) for row in SQLBResult]  
        ProjectNumber2 = [row[9] for row in SQLBResult]
        ProjectDescription = ['' for row in SQLBResult]
        SubTasks = [row[5][:1024] for row in SQLBResult]
        Businessline = [row[3] for row in SQLBResult]
        ProductDomain = [row[10] for row in SQLBResult]
        Type = ['FD' for row in SQLBResult]
        Site_rcr = [row[8] for row in SQLBResult]
        RCRCategories = [row[11] for row in SQLBResult]
        FeatureCategory = [row[12] for row in SQLBResult]
        BusinessPriority = [row[13].replace('.0','') for row in SQLBResult]
        Competence = ['SW' for row in SQLBResult]                
        Effortjira = [row[6] for row in SQLBResult]
        Competence1 = ['HW' for row in SQLBResult]                
        Effortjira1 = [row[7] for row in SQLBResult]
        SEffort = [row[6] for row in SQLBResult]
        HEffort = [row[7] for row in SQLBResult]
        SumAllocation = ['' for row in SQLBResult]
        Phase = ['' for row in SQLBResult]
        Year1 = [sYear for row in SQLBResult]
        
        
        ID1 = [sYear + strnum(stN+i) for i in range(count)]        
        Jan1 = ['' for row in SQLBResult]
        Feb1 = ['' for row in SQLBResult]
        Mar1 = ['' for row in SQLBResult]
        Apr1 = ['' for row in SQLBResult]
        May1 = ['' for row in SQLBResult]
        Jun1 = ['' for row in SQLBResult]
        Jul1 = ['' for row in SQLBResult]
        Aug1 = ['' for row in SQLBResult]
        Sep1 = ['' for row in SQLBResult]
        Oct1 = ['' for row in SQLBResult]
        Nov1 = ['' for row in SQLBResult]
        Dec1 = ['' for row in SQLBResult]
        
        for i in range(count):
            if ProductDomain[i] == 'PON ONTs':
                ProductDomain[i] = 'ONT_PS'           
            elif ProductDomain[i] == 'mobile App':
                ProductDomain[i] = 'Mobile App'
            elif ProductDomain[i] == 'FWA-4G'or ProductDomain[i] == 'FWA-FastMile-4G':
                ProductDomain[i] = 'FWA_PS'
            elif ProductDomain[i] == 'FWA-5G' or ProductDomain[i] == 'FWA-FastMile-5G':
                ProductDomain[i] = 'FWA_PS'            
                
            if ProductDomain[i] == 'ONT_PS' or ProductDomain[i] == 'China Product' or ProductDomain[i] == 'Voice' :
                Businessline[i] = 'BBD-ONT'
            elif ProductDomain[i] == 'Cloud' or ProductDomain[i] == 'Mobile App' or ProductDomain[i] == 'Mesh' or ProductDomain[i] == 'Container App' or ProductDomain[i] == 'HomeOS':
                Businessline[i] = 'BBD-CSW'
            elif ProductDomain[i] == 'Beacon':
                Businessline[i] = 'BBD-NWF'
            elif ProductDomain[i] == 'FWA_PS':
                Businessline[i] = 'BBD-FWA'
              
            Project = ProjectDescription[i]        
            if  'BBDPIPL' in ProjectNumber[i]:
                for row in SQLBResult_pipl:
                    if row[0] == ProjectNumber[i]:
                        Project = row[1].replace("'","\\'")
                        break
            Phase = phase_type(RCR[i])
            Site =''
            Subtk =''
            if Site_rcr[i] == 'ODM':
                Site ='ODM' 
            if  HEffort[i] != '0.0' and HEffort[i] != '' and HEffort[i] != '0' and HEffort[i] != ' ':
                rowvalues.append((ID[i],Releases[i].strip(),RCR[i].strip(),Description[i].strip(),State[i].strip(),ProjectNumber[i],ProjectNumber2[i][:100],Project,FeatureCategory[i].strip(),BusinessPriority[i].strip(),Subtk,Businessline[i].strip(),ProductDomain[i],Type[i].strip(),Site,Competence1[i].strip(),RCRCategories[i].strip(),Phase,Effortjira1[i].strip(),SumAllocation[i].strip(),Year1[i]))
            else:
                rowvalues.append((ID[i],Releases[i].strip(),RCR[i].strip(),Description[i].strip(),State[i].strip(),ProjectNumber[i],ProjectNumber2[i][:100],Project,FeatureCategory[i].strip(),BusinessPriority[i].strip(),Subtk,Businessline[i].strip(),ProductDomain[i],Type[i].strip(),Site,Competence[i].strip(),RCRCategories[i].strip(),Phase,Effortjira[i].strip(),SumAllocation[i].strip(),Year1[i]))   
            print('New RCR =',ID[i],RCR[i], file=fa, flush=True )            
            rowvalues1.append((ID1[i],Year1[i],Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip()))
           
        InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,State,ProjectNumber,ProjectNumber2,ProjectDescription,
        FeatureCategory,BusinessPriority,SubTasks,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Phase,Effortjira,
        SumAllocation,Year1) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        
        InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        print('New 11 =',InUpSql, file=fa, flush=True ) 
        SQLConn.cur.executemany(InUpSql,rowvalues)
        SQLConn.cur.executemany(InUpSql2,rowvalues1)
        SQLConn.commit()
        print('New 12 =',InUpSql, file=fa, flush=True ) 
        num_new += 1
        # result = rd_update2(sRelease,sYear,SQLBConn,SQLConn) 
        
    # update new ====================
    cmd = """
        SELECT
        ID,Effortjira,SubTasks,Siteallocation,RCR,ProductDomain,
        Competence,Type,Site,Year1,Year2,Year3,Releases,RCRCategories,TechnicalAreas,Phase                
        FROM
        cdb_rd_resource
        WHERE 
           Releases = "%s" 
                            
             """ % sRelease
         
    SQLConn.cur.execute(cmd)             
    SQLResult = SQLConn.cur.fetchall()    
    rcr_rd=[]
    for row in SQLResult:
        rcr_rd.append(row[4])
    rcrset = set(rcr_rd)
    rcr_all = sorted(list(rcrset)) 
    
    SQLB_num = """
        SELECT
        count(Status) as num        
        FROM
        jira_issues_rcr
        WHERE 
        `Key` in %s 
        """ % List2String(rcr_all)    
        
    SQLB = """
        SELECT
        `Key`,Summary,Status, BusinessLine, FixVersions,SubTasks,
        TotalFeatureTeamEffort,HWEffort,Site,BusinessRationale,ProductLine,PortfolioTypes,
        FeatureCategory,BusinessPriority1
        FROM
        jira_issues_rcr
        WHERE
          `Key` in %s 
        """ % List2String(rcr_all)
    SQLB_task = """
        SELECT
        `Key`,TechnicalArea,TimeOriginalEstimate,BusinessLine      
        FROM
        jira_issues
        WHERE `Type` = 'Sub-task' AND Status <> 'Rejected'   
        """
    SQLB_pipl = """
        SELECT
        `Key`,Summary      
        FROM
        jira_issues_bbdpipl           
        """
    SQLBCur = SQLBConn.cursor()
    SQLBCur.execute(SQLB_num)          
    SQLBResult = SQLBCur.fetchall()            
    count = SQLBResult[0][0]
    
    SQLBCur.execute(SQLB)             
    SQLBResult = SQLBCur.fetchall()
    
    SQLBCur.execute(SQLB_task)             
    SQLBResult_task = SQLBCur.fetchall()
    
    SQLBCur.execute(SQLB_pipl)             
    SQLBResult_pipl = SQLBCur.fetchall()
                
    rowvalues =[]
    rowvalues1 =[]                   
    Releases = [row[4] for row in SQLBResult]
    RCR = [row[0] for row in SQLBResult]
    Description  = [row[1].replace("'","\\'") for row in SQLBResult]
    State = [row[2] for row in SQLBResult] 
    Site_rcr = [row[8] for row in SQLBResult]
    ProjectNumber = [prjnum(row[9]) for row in SQLBResult]  
    ProjectNumber2 = [row[9] for row in SQLBResult]
    ProjectDescription = ['' for row in SQLBResult]
    ProductDomain = [row[10] for row in SQLBResult]
    RCRCategories = [row[11] for row in SQLBResult]
    FeatureCategory = [row[12] for row in SQLBResult]
    BusinessPriority = [row[13].replace('.0','') for row in SQLBResult]
    SubTasks = [row[5][:1024] for row in SQLBResult]
    Businessline = [row[3] for row in SQLBResult]
    Effortjira = [row[6] for row in SQLBResult]
    Effortjira1 = [row[7] for row in SQLBResult]
    SEffort = [row[6] for row in SQLBResult]
    HEffort = [row[7] for row in SQLBResult]
    
    num_up = 0
    num_del = 0
    for i in range(count):        
        # st = Effortjira[i].find(',')
        SWeft = SEffort[i]
        HWeft = HEffort[i]
        tasks = SubTasks[i]
        Phase = phase_type(RCR[i])       
        
        if ProductDomain[i] == 'PON ONTs':
            ProductDomain[i] = 'ONT_PS'
        elif ProductDomain[i] == 'mobile App':
            ProductDomain[i] = 'Mobile App'
        elif ProductDomain[i] == 'FWA-4G'or ProductDomain[i] == 'FWA-FastMile-4G':
            ProductDomain[i] = 'FWA_PS'
        elif ProductDomain[i] == 'FWA-5G' or ProductDomain[i] == 'FWA-FastMile-5G':
            ProductDomain[i] = 'FWA_PS'        
            
        for row in SQLResult:
            if RCR[i] == row[4] and row[8] =='' and Site_rcr[i]== 'ODM':
                sql="""
                    UPDATE cdb_rd_resource set Site = '%s'
                    WHERE ID = '%s'                
                    """ % (Site_rcr[i],row[0]) 
                print('HW2 =',row[4],row[1],HWeft, file=fa, flush=True )
                num_up += 1
                SQLConn.cur.execute(sql)
                SQLConn.conn.commit()
        if HWeft != '0.0' and HWeft != '' and HWeft != '0':                
            flag = 0
            for row in SQLResult:   
               if RCR[i] == row[4] and row[6] =='HW' and HWeft == row[1] and ProductDomain[i]==row[5]:
                    flag = 1
                    print('HW NO CHANGE =',row[4],HWeft,row[1],row[6], ProductDomain[i],file=fa, flush=True )
                    break
            if flag == 0:
                for row in SQLResult:                
                   if RCR[i] == row[4] and row[6] =='HW' and HWeft != row[1] and ProductDomain[i]==row[5]:                            
                        flag = 1
                        sql="""
                            UPDATE cdb_rd_resource set Effortjira = '%s'
                            WHERE ID = '%s'                
                            """ % (HWeft,row[0]) 
                        print('HW CHANGE =',row[4],row[1],HWeft, file=fa, flush=True )
                        num_up += 1
                        SQLConn.cur.execute(sql)
                        SQLConn.conn.commit()
                        break
            if flag == 0:
                for row in SQLResult:                       
                   if RCR[i] == row[4] and ProductDomain[i]==row[5] and row[6] =='SW' and (SWeft =='' or SWeft =='0.0' or SWeft == '0'):
                        flag = 1
                        hw_site = 'China'
                        if Site_rcr[i] == 'ODM':
                            hw_site = 'ODM'                    
                        Competence ='HW'
                        sql="""
                            UPDATE cdb_rd_resource set Effortjira = '%s',Competence = '%s',Site ='%s'
                            WHERE ID = '%s'                
                            """ % (HWeft,Competence,hw_site,row[0]) 
                        print('SW change to HW=',HWeft,row[4],row[5],file=fa, flush=True ) 
                        num_up += 1
                        SQLConn.cur.execute(sql)
                        SQLConn.conn.commit()
                        break
                if flag == 0:
                    for row in SQLResult:
                       if RCR[i] == row[4] and ProductDomain[i]==row[5] and row[6] !='HW': 
                            flag = 1
                            hw_site = 'China'
                            if Site_rcr[i] == 'ODM':
                                hw_site = 'ODM'                    
                            Competence ='HW'
                            tblname = 'cdb_rd_resource'    
                            stN = tbl_index(tblname,SQLConn)
                            ID = strnum(stN)
                            InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,ProductDomain,Competence,
                                Type,Site,Effortjira,RCRCategories,Year1) VALUES 
                            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                            values = (ID,sRelease,row[4],row[5],Competence,'FD',hw_site,HWeft,row[13],sYear)
                            SQLConn.cur.execute(InUpSql, values)
                            InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year) VALUES 
                            (%s,%s) """
                            TID = sYear + ID
                            print('New HW 1=',TID,row[4],ProductDomain[i],HWeft, file=fa, flush=True ) 
                            values1 = (TID,sYear)
                            num_new += 1                   
                            SQLConn.cur.execute(InUpSql2, values1)
                            SQLConn.commit()
                                
        elif HWeft== '0' or HWeft=='0,0' or HWeft=='' :
            for row in SQLResult:
               if RCR[i] == row[4] and row[6] =='HW' and row[1] != '':
                    ID = row[0]                    
                    sqlt="DELETE FROM cdb_rd_resource WHERE ID = '%s'" % ID        
                    SQLConn.cur.execute(sqlt)
                    sql="DELETE FROM cdb_rd_effort WHERE Right(ID,8) = '%s'" % ID                     
                    SQLConn.cur.execute(sql)
                    SQLConn.conn.commit()
                    num_del += 1
                    print('Dekete HW =',ID,row[4],row[6],Effortjira, file=fa, flush=True )
                            
        if SWeft !='' and SWeft !='0.0' and SWeft != '0':
            flag2 = 0
            for row in SQLResult:
                if RCR[i] == row[4] and row[6] =='SW' and SWeft == row[1] :
                    flag2 = 1
                    break
            if flag2 == 0:
                for row in SQLResult:
                    if RCR[i] == row[4] and row[6] =='SW' and SWeft != row[1]:
                        flag2 = 1
                        sql="""
                            UPDATE cdb_rd_resource set Effortjira = '%s'
                            WHERE RCR = '%s' AND Competence = '%s'                 
                            """ % (SWeft,row[4],row[6]) 
                        print('SW change 1 =',row[4],row[1],SWeft, row[1],file=fa, flush=True )
                        num_up += 1
                        SQLConn.cur.execute(sql)
                        SQLConn.conn.commit()  
               
                if flag2 == 0:                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                              
                    Competence ='SW'
                    tblname = 'cdb_rd_resource'    
                    stN = tbl_index(tblname,SQLConn)
                    ID = strnum(stN)
                    
                    InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,ProductDomain,Competence,
                        Type,Effortjira,RCRCategories,Year1,Businessline,Phase) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                    values = (ID,sRelease,RCR[i],ProductDomain[i],Competence,'FD',SWeft,RCRCategories[i],sYear,Businessline[i],Phase)
                    SQLConn.cur.execute(InUpSql, values)
                    SQLConn.commit()
                    InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year) VALUES 
                    (%s,%s) """
                    TID = sYear + ID
                    print('INSERT SW 1 =',TID,RCR[i],SWeft,ProductDomain[i],sYear,Businessline[i],file=fa, flush=True ) 
                    values1 = (TID,sYear)
                    num_up += 1                    
                    SQLConn.cur.execute(InUpSql2, values1)
                    SQLConn.commit()
                    print('INSERT SW 2 =',TID,sRelease,RCR[i],SWeft, file=fa, flush=True ) 
               
                
        if tasks !="":
            dR = {}        
            dR['data'] = {}
            dR['data']['items'] = []
            k = 0
            tk = tasks.split(',')
            print('subtask =',RCR[i],tasks, file=fa, flush=True )
            for s in tk:
                fls = 0 
                for ts in SQLBResult_task:
                    if s == ts[0]:
                        tarea = ts[1]
                        bl_tsk= ts[3]
                        if bl_tsk =='':
                            bl_tsk = Businessline[i]
                        Apool = pool(tarea,bl_tsk)
                        print('Apool =',RCR[i],s,Apool[0],Apool[1],tarea, file=fa, flush=True )
                        try:
                          tsw = round(Decimal(int(ts[2])/6048000),2)                         
                        except:
                          tsw = 0 
                          print('RCR ERROR1 =',RCR[i],ts[2],tasks, file=fa, flush=True )
                        for item in dR['data']['items']:            
                            if Apool[0] == item['Apool'] :
                              fls = 1                              
                              item['SWeffort'] = round((Decimal(tsw) + Decimal(item['SWeffort'])),2) 
                              # if RCR[i] =='BBDPROD-68865':
                              #     print('subeff =',RCR[i],item['SWeffort'],s,tsw, file=fa, flush=True )
                              techtemp = [item['TechnicalArea'][i] for i in range(len(item['TechnicalArea']))]
                              if tarea not in techtemp:
                                    techtemp.append(tarea)
                                    item['TechnicalArea']=sorted(techtemp)
                              tsktemp = [item['SubTasks'][i] for i in range(len(item['SubTasks']))]
                              if s not in tsktemp:
                                    tsktemp.append(s)
                                    item['SubTasks']=sorted(tsktemp)
                              print('subtask1 =',RCR[i],str(tsktemp), file=fa, flush=True )      
                              
                        if fls == 0:
                            fls = 1
                            k +=1 
                            # if RCR[i] =='BBDPROD-68865':
                            #     print('subeff 1 =',RCR[i],s,tsw, file=fa, flush=True )
                            dItem = {}
                            dItem['SubTasks']=[]
                            dItem['Apool'] = Apool[0]
                            dItem['TechnicalArea'] = []                            
                            dItem['TechnicalArea'].append(tarea)  
                            dItem['Businessline'] = Apool[1]
                            dItem['SWeffort'] = tsw
                            # dItem['BusinessLine'] = bl_tsk
                            dItem['SubTasks'].append(s)
                            print('new subtask =',RCR[i],dItem['Apool'],dItem['TechnicalArea'],dItem['SubTasks'], file=fa, flush=True ) 
                            dR['data']['items'].append(dItem) 
                
                # if fls == 0:
                #     for row in SQLResult:                       
                #         if RCR[i] == row[4] and row[6] =='SW' and row[3]!= '' and s in row[2] :
                #             sql="""
                #                 UPDATE cdb_rd_resource set Siteallocation = '',SubTasks = ''
                #                 WHERE ID = '%s'                
                #                 """ % row[0] 
                #             print('remove task2 =',row[4],row[3],row[2],row[0],row[5],file=fa, flush=True )
                #             num_up += 1
                #             SQLConn.cur.execute(sql)
                #             SQLConn.conn.commit()
            for item in dR['data']['items']:
                flag3 = 0
                for row in SQLResult:
                    # if RCR[i] == 'BBDPROD-54729' and row[4]== 'BBDPROD-54729' and row[5] =='ONT_PS':
                    #     print('RCR item 1 =',RCR[i],row[0],row[8],','.join(item['TechnicalArea']),'row[14]=',row[14],item['SubTasks'],str(item['SWeffort']),row[3],row[2].replace(' ',''),','.join(item['SubTasks']),file=fa, flush=True )
                    if RCR[i] == row[4] and row[6] =='SW' and row[5] == item['Apool'] and row[14]==  ','.join(item['TechnicalArea'])  and row[3] == str(item['SWeffort']) :
                          # if RCR[i] == 'BBDPROD-68865' and row[4]== 'BBDPROD-68865' and row[5] =='Beacon':
                          #     print('RCR item 3 =',RCR[i],row[0],row[8],','.join(item['TechnicalArea']),item['SubTasks'],str(item['SWeffort']),row[3],file=fa, flush=True )
                          flag3 = 1
                    elif RCR[i] == row[4] and row[6] =='SW' and row[5] == item['Apool'] and row[14]==  ','.join(item['TechnicalArea'])  and (row[3] != str(item['SWeffort']) or row[2].replace(' ','') != ','.join(item['SubTasks'])):
                          flag3 = 1
                          sql="""
                              UPDATE cdb_rd_resource set Siteallocation = '%s',SubTasks = '%s'
                              WHERE ID = '%s'                
                              """ % (item['SWeffort'],','.join(item['SubTasks']),row[0]) 
                          print('task change R1 =',row[4],row[3],str(item['SWeffort']),','.join(item['SubTasks']), ','.join(item['TechnicalArea']),row[0],row[5],row[14],row[2],file=fa, flush=True )
                          num_up += 1
                          SQLConn.cur.execute(sql)
                          SQLConn.conn.commit()
                    elif RCR[i] == row[4] and row[6] =='SW' and row[5] == item['Apool'] and (row[14] == '' or row[14] == ','.join(item['TechnicalArea']) or row[14] in ','.join(item['TechnicalArea']))  and (row[3] != str(item['SWeffort']) or row[2].replace(' ','') != ','.join(item['SubTasks'])):
                          flag3 = 1
                          sql="""
                              UPDATE cdb_rd_resource set Siteallocation = '%s',SubTasks = '%s',TechnicalAreas = '%s'
                              WHERE ID = '%s'                
                              """ % (item['SWeffort'],','.join(item['SubTasks']), ','.join(item['TechnicalArea']),row[0]) 
                          print('task change 1 =',row[4],row[3],str(item['SWeffort']),','.join(item['SubTasks']), ','.join(item['TechnicalArea']),row[0],row[5],row[14],file=fa, flush=True )
                          num_up += 1
                          SQLConn.cur.execute(sql)
                          SQLConn.conn.commit() 
                
                if flag3 == 0:
                    for row in SQLResult:    
                         if RCR[i] == row[4] and row[6] =='SW' and row[5] == item['Apool'] and (row[14]==  ','.join(item['TechnicalArea']) or  row[14] in ','.join(item['TechnicalArea'])) and row[3] != str(item['SWeffort']) :
                          flag3 = 1
                          sql="""
                              UPDATE cdb_rd_resource set Siteallocation = '%s',SubTasks = '%s',TechnicalAreas = '%s'
                              WHERE ID = '%s'                
                              """ % (item['SWeffort'],','.join(item['SubTasks']), ','.join(item['TechnicalArea']),row[0]) 
                          print('task change 2 =',row[4],row[3],str(item['SWeffort']),','.join(item['SubTasks']), ','.join(item['TechnicalArea']),row[0],row[5],row[14],file=fa, flush=True )
                          num_up += 1
                          SQLConn.cur.execute(sql)
                          SQLConn.conn.commit()
                          break
                if flag3 == 0:
                    for row in SQLResult:      
                          if RCR[i] == row[4] and row[6] =='SW' and row[5] == item['Apool'] and row[14] == '' and row[3] == str(item['SWeffort']) :
                              flag3 = 1
                              sql="""
                                  UPDATE cdb_rd_resource set TechnicalAreas = '%s'
                                  WHERE ID = '%s'                
                                  """ % ( ','.join(item['TechnicalArea']),row[0]) 
                              print('task change 3  =',row[4],row[3], ','.join(item['TechnicalArea']),row[0],row[5],row[14],file=fa, flush=True )
                              num_up += 1
                              SQLConn.cur.execute(sql)
                              SQLConn.conn.commit()                              
                              break
                if flag3 == 0:
                    for row in SQLResult:          
                          if RCR[i] == row[4] and row[6] =='SW' and row[14] == '' and ','.join(item['TechnicalArea']) !='' :
                              flag3 = 1
                              sql="""
                                  UPDATE cdb_rd_resource set TechnicalAreas = '%s'
                                  WHERE ID = '%s'                
                                  """ % ( ','.join(item['TechnicalArea']),row[0]) 
                              print('task change 4 =',row[4],','.join(item['TechnicalArea']),row[0],row[5],row[14],file=fa, flush=True )
                              num_up += 1
                              SQLConn.cur.execute(sql)
                              SQLConn.conn.commit() 
                              break
                if flag3 == 0:
                    for row in SQLResult:
                        # if RCR[i] == 'BBDPROD-63732':
                        #     print('RCR item 1 =',RCR[i],item['Apool'],','.join(item['TechnicalArea']),item['SubTasks'],str(item['SWeffort']),file=fa, flush=True )
                        if RCR[i] == row[4] and row[6] =='SW' and row[5] == item['Apool'] and row[14]== '' and  ','.join(item['TechnicalArea']) !='':                                    
                              flag3 = 1  
                              sql="""
                                  UPDATE cdb_rd_resource set TechnicalAreas = '%s',Siteallocation = '%s'
                                  WHERE ID = '%s'                
                                  """ % (item['TechnicalArea'],item['SWeffort'],row[0]) 
                              print('Change TA1 =',row[4],str(item['SWeffort']),row[0],row[5], ','.join(item['TechnicalArea']),file=fa, flush=True )
                              num_up += 1
                              SQLConn.cur.execute(sql)
                              SQLConn.conn.commit()
                              break
                
                if flag3 == 0:
                    flag3 = 1                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                              
                    Competence ='SW'
                    tblname = 'cdb_rd_resource'    
                    stN = tbl_index(tblname,SQLConn)
                    ID = strnum(stN)
                    Apool = pool(','.join(item['TechnicalArea']),item['Businessline'])
                    InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,ProductDomain,Competence,
                        Type,Siteallocation,SubTasks,Effortjira,RCRCategories,FeatureCategory,BusinessPriority,Year1,TechnicalAreas,Businessline,Phase) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                    values = (ID,sRelease,RCR[i],Apool[0],Competence,'FD',item['SWeffort'],','.join(item['SubTasks']),SWeft,row[13],FeatureCategory[i].strip(),BusinessPriority[i].strip(),sYear,','.join(item['TechnicalArea']),Apool[1],Phase)
                    SQLConn.cur.execute(InUpSql, values)
                    SQLConn.commit()
                    InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year) VALUES 
                    (%s,%s) """
                    TID = sYear + ID
                    print('INSERT STN 31 =',TID,RCR[i],SWeft,Apool[0],','.join(item['SubTasks']),row[13],sYear,','.join(item['TechnicalArea']),Apool[1],file=fa, flush=True ) 
                    values1 = (TID,sYear)
                    num_up += 1                    
                    SQLConn.cur.execute(InUpSql2, values1)
                    SQLConn.commit()
                    print('INSERT STN 31-2 =',TID,sRelease,RCR[i],SWeft, file=fa, flush=True ) 
                            
        tk = tasks.split(',')                     
        for row in SQLResult: 
              if RCR[i] == row[4] and row[6] =='SW' and row[2] !='' and ',' not in row[2]:
                if row[2] not in tk:
                    ID = row[0]                    
                    sqlt="DELETE FROM cdb_rd_resource WHERE ID = '%s'" % ID        
                    SQLConn.cur.execute(sqlt)
                    sql="DELETE FROM cdb_rd_effort WHERE Right(ID,8) = '%s'" % ID                     
                    SQLConn.cur.execute(sql)
                    SQLConn.conn.commit()
                    num_del += 1
                    print('Delete 2 =',ID,row[4],row[6],Effortjira[i],row[2],tk, file=fa, flush=True )
        
              elif RCR[i] == row[4] and row[6] =='SW' and row[2] !='' and ',' in row[2]:
                ts = row[2].split(',')
                flag_t = 0
                for s in ts:                    
                    if s in tk:
                       flag_t = 1 
                if flag_t == 0:
                    ID = row[0]                    
                    sqlt="DELETE FROM cdb_rd_resource WHERE ID = '%s'" % ID        
                    SQLConn.cur.execute(sqlt)
                    sql="DELETE FROM cdb_rd_effort WHERE Right(ID,8) = '%s'" % ID                     
                    SQLConn.cur.execute(sql)
                    SQLConn.conn.commit()
                    num_del += 1
                    print('Delete 3 =',ID,row[4],row[6],Effortjira[i],row[2],tk, file=fa, flush=True )         
    for i in range(count):
        Phase = phase_type(RCR[i])
        Project = ProjectDescription[i]        
        if  'BBDPIPL' in ProjectNumber[i]:
            for row in SQLBResult_pipl:
                if row[0] == ProjectNumber[i]:
                    Project = row[1].replace("'","\\'")
                    break
        if Project == '':
            Project = prjdes(ProjectNumber[i])                    
        rowvalues.append((Releases[i],Description[i].strip(),State[i].strip(),ProjectNumber[i],ProjectNumber2[i][:100],Project,RCRCategories[i].strip(),Phase,FeatureCategory[i].strip(),BusinessPriority[i].strip(),sRelease,RCR[i]))
        
    InUpSql ="""UPDATE cdb_rd_resource SET Releases=%s,Description=%s,State=%s,ProjectNumber=%s,ProjectNumber2=%s,
         ProjectDescription=%s,RCRCategories = %s,Phase = %s,FeatureCategory= %s,BusinessPriority= %s
     WHERE Releases=%s AND RCR=%s"""
   
    SQLConn.cur.executemany(InUpSql,rowvalues)              
    SQLConn.commit() 
        
    return str(num_add),str(num_up),str(num_del),str(num_new)

def rd_rcr_refresh(request):    
    
    SQLBConn = pymysql.connect(host  = settings.BBD_DB['host'],
                            port     = settings.BBD_DB['port'],
                            user     = settings.BBD_DB['username'],
                            password = settings.BBD_DB['password'],
                            database = settings.BBD_DB['name'],
                            charset  = settings.BBD_DB['charset']
                        )
    SQLBCur = SQLBConn.cursor()   
    SQLB_num = """
    SELECT
       count(Status) as num        
    FROM
       jira_issues_rcr
    WHERE 
       FixVersions NOT LIKE CONCAT('%','BBDR2','%')
       
    """ 
    
    SQLB = """
    SELECT
     `Key`,Summary,Status, BusinessLine, FixVersions,SubTasks,
     TotalFeatureTeamEffort,HWEffort,Site,BusinessRationale,ProductLine,PortfolioTypes        
    FROM
     jira_issues_rcr
    WHERE 
     FixVersions NOT LIKE CONCAT('%','BBDR2','%')     
     
       """ 
       #FixVersions LIKE CONCAT('%','BBD','%')  AND  FixVersions NOT LIKE CONCAT('%','BBDR2','%') 
    SQLB_task = """
    SELECT
    `Key`,TechnicalArea,TimeOriginalEstimate      
    FROM
    jira_issues
    WHERE `Type` = 'Sub-task' AND Status <> 'Rejected'   
    """
    SQLB_pipl = """
    SELECT
    `Key`,Summary      
    FROM
    jira_issues_bbdpipl           
    """
    SQLBCur.execute(SQLB_num)          
    SQLResult = SQLBCur.fetchall()            
    count = SQLResult[0][0]
    
    SQLBCur.execute(SQLB)             
    SQLBResult = SQLBCur.fetchall()
    
    SQLBCur.execute(SQLB_task)             
    SQLBResult_task = SQLBCur.fetchall()
    
    SQLBCur.execute(SQLB_pipl)             
    SQLBResult_pipl = SQLBCur.fetchall()
    
    SQLConn = analyzer_db()    
            
    rowvalues =[]                   
    Releases = [row[4] for row in SQLBResult]
    RCR = [row[0] for row in SQLBResult]
    Description  = [row[1].replace("'","\\'") for row in SQLBResult]
    State = [row[2] for row in SQLBResult]
    BusinessLine = [row[3] for row in SQLBResult]
    ProjectNumber = [row[9] for row in SQLBResult]
    ProjectDescription = ['' for row in SQLBResult]
    SubTasks = [row[5][:100] for row in SQLBResult]
    Businessline = [row[3] for row in SQLBResult]
    Effortjira = [row[6]+','+row[7] for row in SQLBResult]
    SWeft = [row[6] for row in SQLBResult]
    HWeft = [row[7] for row in SQLBResult]
       
    cmd = """
    SELECT
    ID,Effortjira,SubTasks,Siteallocation,RCR,ProductDomain,
    Competence,Type,Site,Year1,Year2,Year3,Releases,RCRCategories                
    FROM
    cdb_rd_resource
    WHERE 
    RCR LIKE CONCAT('%','BBDPROD','%') 
    AND Releases LIKE CONCAT('%','BBD','%') 
    AND Releases NOT LIKE CONCAT('%','BBDR2','%')
    """ 
    SQLConn.cur.execute(cmd)             
    SQLResult = SQLConn.cur.fetchall()
    num_up = 0
    for row in SQLResult:        
        try:
            k = RCR.index(row[4])
        except:
            k = -1
        if k != -1 :
            Effort =''
            subtsk = ''
            subtaskid = ''
            if row[6] == 'SW':
                Effort = SWeft[k]
                tasks = SubTasks[k]
                dR = {}        
                dR['data'] = {}
                dR['data']['items'] = []
                if tasks !="":                    
                    h = 0
                    tk = tasks.split(',')
                    for s in tk:
                        fls = 0 
                        for ts in SQLBResult_task:
                            if s == ts[0]:
                                tarea = ts[1]
                                bl_tsk= ts[3]
                                # if bl_tsk =='':
                                #     bl_tsk = BusinessLinee[i]
                                Apool = pool(tarea,bl_tsk)
                                try:
                                  tsw = round(Decimal(int(ts[2])/6048000),2)                         
                                except:
                                  tsw = 0                               
                                for item in dR['data']['items']:            
                                    if Apool == item['Apool']:
                                      fls = 1                              
                                      item['SWeffort'] = round((Decimal(tsw) + Decimal(item['SWeffort'])),2) 
                                      techtemp = [item['TechnicalArea'][i] for i in range(len(item['TechnicalArea']))]
                                      if tarea not in techtemp:
                                           techtemp.append(tarea)
                                           item['TechnicalArea']=sorted(techtemp)
                                      tsktemp = [item['SubTasks'][i] for i in range(len(item['SubTasks']))]
                                      if s not in tsktemp:
                                          tsktemp.append(s)
                                          item['SubTasks']=sorted(tsktemp)
                                      break
                                if fls == 0:
                                    fls = 1
                                    h +=1                                    
                                    dItem = {}
                                    dItem['Apool'] = Apool
                                    dItem['TechnicalArea'] = []                            
                                    dItem['TechnicalArea'].append(tarea) 
                                    dItem['Businessline'] = bl_tsk
                                    dItem['SWeffort'] = tsw
                                    dItem['SubTasks'].append(s)
                                    dR['data']['items'].append(dItem)
                for item in dR['data']['items']: 
                    if row[5] == ','.join(item['TechnicalArea']):
                        subtsk =  item['SWeffort']
                        subtaskid = ','.join(item['SubTasks'])
            elif row[6] == 'HW':
                Effort = HWeft[k] 
            Project = ''        
            if  'BBDPIPL' in ProjectNumber[k]:
                for row_p in SQLBResult_pipl:
                    if row_p[0] == ProjectNumber[k]:
                        Project = repspecial(row_p[1][:1024].replace("'","\'"))
                        break
                
            sql="""
                UPDATE cdb_rd_resource set Releases = '%s',Description = '%s',State = '%s',ProjectNumber = '%s',
                ProjectDescription = '%s',SubTasks = '%s',Businessline = '%s',Effortjira = '%s',Siteallocation = '%s'
                WHERE ID = '%s'                
                """ % (Releases[k],repspecial(Description[k]),State[k],ProjectNumber[k][:30],Project,subtaskid,Businessline[k],Effort,subtsk,row[0]) 
            # print('update RCR =',row[4],RCR[k],Releases[k],State[k],ProjectNumber[k],Project,subtaskid,Effort,subtsk,file=fa, flush=True )
            print('update RCR =',row[4],sql,file=fa, flush=True )
            num_up += 1
            SQLConn.cur.execute(sql)
            SQLConn.conn.commit()           
            
    SQLConn.close()
    dResult = {}        
    dResult['data'] = {}
    dResult['data']['items'] = []    
    dItem = {}
    dItem['Result'] = 'Refresh data from Jira successful'
    dItem['Update'] = str(num_up)
    dResult['data']['items'].append(dItem)
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json') 

def rd_resource_upload(request):    
    file = request.FILES.get('file')
    # print('uplaod:%s'% file,file=fa,flush=True)
    # 创建upload文件夹
    if not os.path.exists(settings.UPLOAD_ROOT):
        os.makedirs(settings.UPLOAD_ROOT)
    try:
        if file is None:
            return HttpResponse('Please select the file you want to upload')
        # 循环二进制写入
        with open(settings.UPLOAD_ROOT + "/" + file.name, 'wb') as f:
            for i in file.readlines():
                f.write(i)

        # 写入 mysql
        filename = settings.UPLOAD_ROOT + "/" + file.name
        resource_insert_db(filename)
    except Exception as e:
        return HttpResponse(e) 
    return HttpResponse('Update successful')


def resource_insert_db(filename):    
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    # ncols = sheet.ncols 
    # print('nrows new =',nrows,file=fa,flush=True)
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    SQLConn = analyzer_db()    
    SQLConn.cur.execute('truncate table cdb_rd_resource') 
    SQLConn.cur.execute('truncate table cdb_rd_effort')      
        
    limitall = 50000
    k = 1
    tblname = 'cdb_rd_resource'    
    stN = tbl_index(tblname,SQLConn)
    # print('nrows =',nrows,stN,file=fa,flush=True)
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]
        rowvalues1 =[]
        rowvalues2 =[]
        rowvalues3 =[]
        ID = [strnum(stN+i) for i in range(limit)]            
        Releases = [str(sheet.cell_value(i+k, 0)) for i in range(limit)]
        RCR = [str(sheet.cell_value(i+k, 1)) for i in range(limit)]
        Description  = [str(sheet.cell_value(i+k, 2)) for i in range(limit)]
        State = [str(sheet.cell_value(i+k, 3)) for i in range(limit)]
        TechnicalAreas = [str(sheet.cell_value(i+k, 4)) for i in range(limit)]
        ProjectNumber = [str(sheet.cell_value(i+k, 5)) for i in range(limit)]
        ProjectDescription = [str(sheet.cell_value(i+k, 6)) for i in range(limit)]
        ProjectState = [str(sheet.cell_value(i+k, 7)) for i in range(limit)]
        Businessline = [str(sheet.cell_value(i+k, 8)) for i in range(limit)]
        ProductDomain = [str(sheet.cell_value(i+k, 9)) for i in range(limit)]
        Type = [str(sheet.cell_value(i+k, 10)) for i in range(limit)]
        Site = [str(sheet.cell_value(i+k, 11)) for i in range(limit)]
        Activity = [str(sheet.cell_value(i+k, 12)) for i in range(limit)]
        Competence = [str(sheet.cell_value(i+k, 13)) for i in range(limit)]
        RCRCategories = [str(sheet.cell_value(i+k, 14)) for i in range(limit)]        
        Effortjira = [str(sheet.cell_value(i+k, 15)) for i in range(limit)]
        SumAllocation = [str(sheet.cell_value(i+k, 16)) for i in range(limit)]
        Phase = [str(sheet.cell_value(i+k, 17)) for i in range(limit)]
        Year1 = [str(sheet.cell_value(0, 18))[1:5] for i in range(limit)]
        Year2 = [str(sheet.cell_value(0, 19))[1:5] for i in range(limit)]
        Year3 = [str(sheet.cell_value(0, 20))[1:5] for i in range(limit)]
        
        ID1 = [str(sheet.cell_value(0, 18))[1:5] + strnum(stN+i) for i in range(limit)]        
        Jan1 = [str(sheet.cell_value(i+k, 21)) for i in range(limit)]
        Feb1 = [str(sheet.cell_value(i+k, 22)) for i in range(limit)]
        Mar1 = [str(sheet.cell_value(i+k, 23)) for i in range(limit)]
        Apr1 = [str(sheet.cell_value(i+k, 24)) for i in range(limit)]
        May1 = [str(sheet.cell_value(i+k, 25)) for i in range(limit)]
        Jun1 = [str(sheet.cell_value(i+k, 26)) for i in range(limit)]
        Jul1 = [str(sheet.cell_value(i+k, 27)) for i in range(limit)]
        Aug1 = [str(sheet.cell_value(i+k, 28)) for i in range(limit)]
        Sep1 = [str(sheet.cell_value(i+k, 29)) for i in range(limit)]
        Oct1 = [str(sheet.cell_value(i+k, 30)) for i in range(limit)]
        Nov1 = [str(sheet.cell_value(i+k, 31)) for i in range(limit)]
        Dec1 = [str(sheet.cell_value(i+k, 32)) for i in range(limit)]
        ID2 = [str(sheet.cell_value(0, 19))[1:5] + strnum(stN+i) for i in range(limit)] 
        Jan2 = [str(sheet.cell_value(i+k, 33)) for i in range(limit)]
        Feb2 = [str(sheet.cell_value(i+k, 34)) for i in range(limit)]
        Mar2 = [str(sheet.cell_value(i+k, 35)) for i in range(limit)]
        Apr2 = [str(sheet.cell_value(i+k, 36)) for i in range(limit)]
        May2 = [str(sheet.cell_value(i+k, 37)) for i in range(limit)]
        Jun2 = [str(sheet.cell_value(i+k, 38)) for i in range(limit)]
        Jul2 = [str(sheet.cell_value(i+k, 39)) for i in range(limit)]
        Aug2 = [str(sheet.cell_value(i+k, 40)) for i in range(limit)]
        Sep2 = [str(sheet.cell_value(i+k, 41)) for i in range(limit)]
        Oct2 = [str(sheet.cell_value(i+k, 42)) for i in range(limit)]
        Nov2 = [str(sheet.cell_value(i+k, 43)) for i in range(limit)]
        Dec2 = [str(sheet.cell_value(i+k, 44)) for i in range(limit)]
        ID3 = [str(sheet.cell_value(0, 20))[1:5] + strnum(stN+i) for i in range(limit)] 
        Jan3 = [str(sheet.cell_value(i+k, 45)) for i in range(limit)]
        Feb3 = [str(sheet.cell_value(i+k, 46)) for i in range(limit)]
        Mar3 = [str(sheet.cell_value(i+k, 47)) for i in range(limit)]
        Apr3 = [str(sheet.cell_value(i+k, 48)) for i in range(limit)]
        May3 = [str(sheet.cell_value(i+k, 49)) for i in range(limit)]
        Jun3 = [str(sheet.cell_value(i+k, 50)) for i in range(limit)]
        Jul3 = [str(sheet.cell_value(i+k, 51)) for i in range(limit)]
        Aug3 = [str(sheet.cell_value(i+k, 52)) for i in range(limit)]
        Sep3 = [str(sheet.cell_value(i+k, 53)) for i in range(limit)]
        Oct3 = [str(sheet.cell_value(i+k, 54)) for i in range(limit)]
        Nov3 = [str(sheet.cell_value(i+k, 55)) for i in range(limit)]
        Dec3 = [str(sheet.cell_value(i+k, 56)) for i in range(limit)]
        
        k+=1
        # print('nrows =',nrows, 'ftype =', Ftype1, Ftype2, Ftype3, str(k), str(limit),file=fa,flush=True)
        # print(City, file=fa,flush=True) 
        for i in range(limit): 
            if Site[i] == 'SHANGHAI':
                Site[i] = 'Shanghai'
                
            if Releases[i] == 'Maintenance':
                Releases[i] = 'Common'
            elif Releases[i] == 'training':
                Releases[i] = 'Common'
            
            if ProductDomain[i] == 'RGW(ONT)':
                ProductDomain[i] = 'ONT_PS'
            elif ProductDomain[i] == 'Beacon Cloud':
                ProductDomain[i] = 'Cloud'
            elif ProductDomain[i] == 'Beacon MAPP':
                ProductDomain[i] = 'Mobile App'
            elif ProductDomain[i] == 'FWA-4G':
                ProductDomain[i] = 'FWA_PS'
            elif ProductDomain[i] == 'FWA-5G' :
                ProductDomain[i] = 'FWA_PS'
            elif ProductDomain[i] == 'Beacon Mesh':
                ProductDomain[i] = 'Mesh'
            rowvalues.append((ID[i],Releases[i].strip(),RCR[i].strip(),Description[i].strip(),State[i].strip(),TechnicalAreas[i].strip(),ProjectNumber[i].strip(),ProjectDescription[i].strip(),ProjectState[i].strip(),Businessline[i].strip(),ProductDomain[i].strip(),Type[i].strip(),Site[i].strip(),Activity[i].strip(),Competence[i].strip(),RCRCategories[i].strip(),Effortjira[i].strip(),SumAllocation[i].strip(),Phase[i].strip(),Year1[i],Year2[i],Year3[i]))
            rowvalues1.append((ID1[i],Year1[i],Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip()))
            rowvalues2.append((ID2[i],Year2[i],Jan2[i].strip(),Feb2[i].strip(),Mar2[i].strip(),Apr2[i].strip(),May2[i].strip(),Jun2[i].strip(),Jul2[i].strip(),Aug2[i].strip(),Sep2[i].strip(),Oct2[i].strip(),Nov2[i].strip(),Dec2[i].strip()))
            rowvalues3.append((ID3[i],Year3[i],Jan3[i].strip(),Feb3[i].strip(),Mar3[i].strip(),Apr3[i].strip(),May3[i].strip(),Jun3[i].strip(),Jul3[i].strip(),Aug3[i].strip(),Sep3[i].strip(),Oct3[i].strip(),Nov3[i].strip(),Dec3[i].strip()))
            
        InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,State,TechnicalAreas,ProjectNumber,ProjectDescription,ProjectState,Businessline,ProductDomain,Type,Site,Activity,Competence,RCRCategories,Effortjira,SumAllocation,Phase,Year1,Year2,Year3) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        
        InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        
        SQLConn.cur.executemany(InUpSql,rowvalues)
        SQLConn.cur.executemany(InUpSql2,rowvalues1)
        SQLConn.cur.executemany(InUpSql2,rowvalues2)
        SQLConn.cur.executemany(InUpSql2,rowvalues3)             
        SQLConn.commit()
        # print('new nrows =',nrows, str(h), str(t), str(k), str(limit),file=fa,flush=True)
        k += limit
    rowx = nrows-1     
    SQLConn.close()    
    return rowx

def rd_file_upload(request):    
    file = request.FILES.get('file')
    print('uplaod:%s'% file,file=fa,flush=True)
    # 创建upload文件夹
    if not os.path.exists(settings.UPLOAD_ROOT):
        os.makedirs(settings.UPLOAD_ROOT)
    try:
        if file is None:
            return HttpResponse('Please select the file you want to upload')
        # 循环二进制写入
        with open(settings.UPLOAD_ROOT + "/" + file.name, 'wb') as f:
            for i in file.readlines():
                f.write(i)

        # 写入 mysql
        filename = settings.UPLOAD_ROOT + "/" + file.name
        row_num = file_cleaning_db(filename)
        # row_num = file_reload_db(filename)
    except Exception as e:
        return HttpResponse(e) 
    return HttpResponse('Update successful, total num =',str(row_num))

def file_reload_db1(filename):    
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    print('nrows =',nrows,file=fa,flush=True)
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    SQLConn = analyzer_db() 
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)] 
    col_release = Title.index('Release/Categories')
    col_RCR = Title.index('RCR/Activities')
    col_Year = Title.index('Year')
    col_site = Title.index('Site')
    col_des = Title.index('Description')
    col_State = Title.index('State')
    col_Prn = Title.index('Project Number')
    col_Project = Title.index('Project Description')
    col_BL = Title.index('Businessline')
    col_TA = Title.index('Technical Area')
    col_Type = Title.index('Type')
    col_Comp = Title.index('Competence')
    col_Cat = Title.index('Categories')
    col_Subtask = Title.index('Sub Task')
    col_Effort = Title.index('Effort jira')
    col_Year1 = Title.index('Year1')
    col_Year2 = Title.index('Year2')
    col_Year3 = Title.index('Year3')
    
    col_ID = Title.index('ID')
    col_YID = Title.index('YID')
    print('cols =',str(col_ID), str(col_site), str(col_RCR), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]
        
        ID = [str(sheet.cell_value(i+k, col_ID)) for i in range(limit)] 
        Site = [str(sheet.cell_value(i+k, col_site)) for i in range(limit)]              
        Release = [str(sheet.cell_value(i+k, col_release)) for i in range(limit)]        
        RCR = [str(sheet.cell_value(i+k, col_RCR)) for i in range(limit)]        
        Des = [str(sheet.cell_value(i+k, col_des)) for i in range(limit)]
        State = [str(sheet.cell_value(i+k, col_State)) for i in range(limit)]
        Prn = [str(sheet.cell_value(i+k, col_Prn)) for i in range(limit)]
        Project = [str(sheet.cell_value(i+k, col_Project)) for i in range(limit)]
        BL = [str(sheet.cell_value(i+k, col_BL)) for i in range(limit)]
        TA = [str(sheet.cell_value(i+k, col_TA)) for i in range(limit)]
        Type = [str(sheet.cell_value(i+k, col_Type)) for i in range(limit)]
        Comp = [str(sheet.cell_value(i+k, col_Comp)) for i in range(limit)]
        Cat = [str(sheet.cell_value(i+k, col_Cat)) for i in range(limit)]
        Subtask = [str(sheet.cell_value(i+k, col_Subtask)) for i in range(limit)]
        Effort = [str(sheet.cell_value(i+k, col_Effort)) for i in range(limit)]
        Year1 = [str(sheet.cell_value(i+k, col_Year1)) for i in range(limit)]
        Year2 = [str(sheet.cell_value(i+k, col_Year2)) for i in range(limit)]
        Year3 = [str(sheet.cell_value(i+k, col_Year3)) for i in range(limit)]
        
    
        for i in range(limit):
            rowvalues.append((ID[i],Release[i],RCR[i],Site[i],Des[i],State[i],Prn[i],Project[i],BL[i],TA[i],Type[i],Comp[i],Cat[i],Subtask[i],Effort[i],Year1[i],Year2[i],Year3[i]))            
            
        InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases, RCR, Site, Description, State,ProjectNumber, ProjectDescription,
        Businessline,ProductDomain,Type,Competence,RCRCategories,Siteallocation,Effortjira,
        Year1,Year2,Year3) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        print('SQL =',InUpSql,file=fa,flush=True)
        SQLConn.executemany(InUpSql,rowvalues)        
        SQLConn.commit()
        k += limit
    rowx = nrows-1    
    SQLConn.close()
    return rowx


def file_reload_db(filename):
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    print('nrows =',nrows,file=fa,flush=True)
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    SQLConn = analyzer_db() 
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)] 
    
    col_Year = Title.index('Year')
    # print('Title =',Title,file=fa,flush=True)
    col_Jan = Title.index('Jan')
    col_Feb = Title.index('Feb')
    col_Mar = Title.index('Mar')
    col_Apr = Title.index('Apr')
    col_May = Title.index('May')
    col_Jun = Title.index('Jun')
    col_Jul = Title.index('Jul')
    col_Aug = Title.index('Aug')
    col_Sep = Title.index('Sep')
    col_Oct = Title.index('Oct')
    col_Nov = Title.index('Nov')
    col_Dec = Title.index('Dec')
    col_ID = Title.index('ID')
    col_YID = Title.index('YID')
    print('cols =',str(col_ID),str(col_YID), str(col_Jan), str(col_Dec), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        # rowvalues =[]
        rowvalues =[]
        ID = [str(sheet.cell_value(i+k, col_ID)) for i in range(limit)] 
        Year = [str(sheet.cell_value(i+k, col_Year)) for i in range(limit)]              
        ID1 = [str(sheet.cell_value(i+k, col_YID)) for i in range(limit)]        
        Jan1 = [str(sheet.cell_value(i+k, col_Jan)) for i in range(limit)]
        Feb1 = [str(sheet.cell_value(i+k, col_Feb)) for i in range(limit)]
        Mar1 = [str(sheet.cell_value(i+k, col_Mar)) for i in range(limit)]
        Apr1 = [str(sheet.cell_value(i+k, col_Apr)) for i in range(limit)]
        May1 = [str(sheet.cell_value(i+k, col_May)) for i in range(limit)]
        Jun1 = [str(sheet.cell_value(i+k, col_Jun)) for i in range(limit)]
        Jul1 = [str(sheet.cell_value(i+k, col_Jul)) for i in range(limit)]
        Aug1 = [str(sheet.cell_value(i+k, col_Aug)) for i in range(limit)]
        Sep1 = [str(sheet.cell_value(i+k, col_Sep)) for i in range(limit)]
        Oct1 = [str(sheet.cell_value(i+k, col_Oct)) for i in range(limit)]
        Nov1 = [str(sheet.cell_value(i+k, col_Nov)) for i in range(limit)]
        Dec1 = [str(sheet.cell_value(i+k, col_Dec)) for i in range(limit)]        
    
        for i in range(limit):
            
            rowvalues.append((ID1[i],Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip(),Year[i][:4]))
            
        
        InUpSql = """INSERT INTO cdb_rd_effort (ID,Jans, Febs, Mars, Aprs, Mays,Juns, Juls,
        Augs,Seps,Octs,Novs,Decs,Year) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
        print('SQL =',InUpSql,file=fa,flush=True)
        SQLConn.executemany(InUpSql,rowvalues)
        SQLConn.commit()
        print('update nrows =',nrows, str(h), str(t), str(k), str(limit),sLastupdate,file=fc,flush=True)
        k += limit
    rowx = nrows-1    
    SQLConn.close()
    return rowx


def file_cleaning_db(filename):    
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    SQLConn = analyzer_db() 
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)] 
    
    col_release = Title.index('Release/Categories')
    col_rcr = Title.index('RCR/Activities')
    col_cat = Title.index('Description')    
    # print('Title =',Title,file=fa,flush=True)    
    col_ID = Title.index('ID')
    # col_YID = Title.index('YID')
    print('cols =',str(col_ID), str(col_release), str(col_rcr), str(col_cat), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]        
        ID = [str(sheet.cell_value(i+k, col_ID)) for i in range(limit)] 
        release = [str(sheet.cell_value(i+k, col_release)) for i in range(limit)]
        rcr = [str(sheet.cell_value(i+k, col_rcr)) for i in range(limit)]
        cat = [str(sheet.cell_value(i+k, col_cat)) for i in range(limit)]              
        # ID1 = [str(sheet.cell_value(i+k, col_YID)) for i in range(limit)] 
    
        for i in range(limit):
            rowvalues.append((release[i].strip(),rcr[i].strip(),cat[i].strip(),ID[i]))
            
        Sql="""UPDATE cdb_rd_resource SET Releases = %s,RCR = %s, Description  = %s
        WHERE ID = %s""" 
         
        # print(Sql, file=fa,flush=True )     
        SQLConn.executemany(Sql,rowvalues)        
        SQLConn.commit()
        print('update nrows =',nrows, str(h), str(t), str(k), str(limit),sLastupdate,file=fc,flush=True)
        k += limit
    rowx = nrows-1    
    SQLConn.close()
    return rowx


def file_insert_db(filename):    
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    # ncols = sheet.ncols 
    # print('nrows new =',nrows,file=fa,flush=True)
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
    
    cmd = """
          SELECT 
            `Key`,Summary,Status, BusinessLine, FixVersions,SubTasks,
             TotalFeatureTeamEffort,HWEffort,Site,BusinessRationale,ProductLine,PortfolioTypes 
          FROM 
            jira_issues_rcr 
          WHERE 
               FixVersions LIKE CONCAT('%','BBD','%') 
               AND FixVersions NOT LIKE CONCAT('%','BBDR2','%')
         """
    SQLBConn = pymysql.connect(host  = settings.BBD_DB['host'],
                            port     = settings.BBD_DB['port'],
                            user     = settings.BBD_DB['username'],
                            password = settings.BBD_DB['password'],
                            database = settings.BBD_DB['name'],
                            charset  = settings.BBD_DB['charset']
                        )
    SQLBCur = SQLBConn.cursor()
    SQLBCur.execute(cmd)                
    SQLBResult = SQLBCur.fetchall()
    SQLBConn.close()
                
    SQLConn = analyzer_db() 
    
    limitall = 50000
    k = 1
    tblname = 'cdb_rd_resource'    
    stN = tbl_index(tblname,SQLConn)
    # print('nrows =',nrows,stN,file=fa,flush=True)
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]
        rowvalues1 =[]
        rowvalues2 =[]
        rowvalues3 =[]
        ID = [strnum(stN+i) for i in range(limit)]            
        Releases= [str(sheet.cell_value(i+k, 0)) for i in range(limit)]
        RCR = [str(sheet.cell_value(i+k, 1)) for i in range(limit)]
        ProductDomain = [str(sheet.cell_value(i+k, 9)) for i in range(limit)]        
        pType = [str(sheet.cell_value(i+k, 10)) for i in range(limit)]
        Site = [str(sheet.cell_value(i+k, 11)) for i in range(limit)]        
        Competence = [str(sheet.cell_value(i+k, 13)) for i in range(limit)]
        RCRCategories = [str(sheet.cell_value(i+k, 14)) for i in range(limit)]        
       
        Year1 = [str(sheet.cell_value(0, 18))[1:5] for i in range(limit)]
        Year2 = [str(sheet.cell_value(0, 19))[1:5] for i in range(limit)]
        Year3 = [str(sheet.cell_value(0, 20))[1:5] for i in range(limit)]
        
        # ID1 = [str(sheet.cell_value(0, 18))[1:5] + strnum(stN+i) for i in range(limit)]        
        # Jan1 = [str(sheet.cell_value(i+k, 21)) for i in range(limit)]
        # Feb1 = [str(sheet.cell_value(i+k, 22)) for i in range(limit)]
        # Mar1 = [str(sheet.cell_value(i+k, 23)) for i in range(limit)]
        # Apr1 = [str(sheet.cell_value(i+k, 24)) for i in range(limit)]
        # May1 = [str(sheet.cell_value(i+k, 25)) for i in range(limit)]
        # Jun1 = [str(sheet.cell_value(i+k, 26)) for i in range(limit)]
        # Jul1 = [str(sheet.cell_value(i+k, 27)) for i in range(limit)]
        # Aug1 = [str(sheet.cell_value(i+k, 28)) for i in range(limit)]
        # Sep1 = [str(sheet.cell_value(i+k, 29)) for i in range(limit)]
        # Oct1 = [str(sheet.cell_value(i+k, 30)) for i in range(limit)]
        # Nov1 = [str(sheet.cell_value(i+k, 31)) for i in range(limit)]
        # Dec1 = [str(sheet.cell_value(i+k, 32)) for i in range(limit)]
        ID2 = [str(sheet.cell_value(0, 19))[1:5] + strnum(stN+i) for i in range(limit)] 
        Jan2 = [str(sheet.cell_value(i+k, 33)) for i in range(limit)]
        Feb2 = [str(sheet.cell_value(i+k, 34)) for i in range(limit)]
        Mar2 = [str(sheet.cell_value(i+k, 35)) for i in range(limit)]
        Apr2 = [str(sheet.cell_value(i+k, 36)) for i in range(limit)]
        May2 = [str(sheet.cell_value(i+k, 37)) for i in range(limit)]
        Jun2 = [str(sheet.cell_value(i+k, 38)) for i in range(limit)]
        Jul2 = [str(sheet.cell_value(i+k, 39)) for i in range(limit)]
        Aug2 = [str(sheet.cell_value(i+k, 40)) for i in range(limit)]
        Sep2 = [str(sheet.cell_value(i+k, 41)) for i in range(limit)]
        Oct2 = [str(sheet.cell_value(i+k, 42)) for i in range(limit)]
        Nov2 = [str(sheet.cell_value(i+k, 43)) for i in range(limit)]
        Dec2 = [str(sheet.cell_value(i+k, 44)) for i in range(limit)]
        ID3 = [str(sheet.cell_value(0, 20))[1:5] + strnum(stN+i) for i in range(limit)] 
        Jan3 = [str(sheet.cell_value(i+k, 45)) for i in range(limit)]
        Feb3 = [str(sheet.cell_value(i+k, 46)) for i in range(limit)]
        Mar3 = [str(sheet.cell_value(i+k, 47)) for i in range(limit)]
        Apr3 = [str(sheet.cell_value(i+k, 48)) for i in range(limit)]
        May3 = [str(sheet.cell_value(i+k, 49)) for i in range(limit)]
        Jun3 = [str(sheet.cell_value(i+k, 50)) for i in range(limit)]
        Jul3 = [str(sheet.cell_value(i+k, 51)) for i in range(limit)]
        Aug3 = [str(sheet.cell_value(i+k, 52)) for i in range(limit)]
        Sep3 = [str(sheet.cell_value(i+k, 53)) for i in range(limit)]
        Oct3 = [str(sheet.cell_value(i+k, 54)) for i in range(limit)]
        Nov3 = [str(sheet.cell_value(i+k, 55)) for i in range(limit)]
        Dec3 = [str(sheet.cell_value(i+k, 56)) for i in range(limit)]
        
        k+=1
        # print('nrows =',nrows, 'ftype =', Ftype1, Ftype2, Ftype3, str(k), str(limit),file=fa,flush=True)
        # print(City, file=fa,flush=True) 
        for i in range(limit): 
            if Site[i] == 'SHANGHAI':
                Site[i] = 'Shanghai'
            if ProductDomain[i] == 'RGW(ONT)':
                ProductDomain[i] = 'ONT_PS'
            elif ProductDomain[i] == 'Beacon Cloud':
                ProductDomain[i] = 'Cloud'
            elif ProductDomain[i] == 'Beacon MAPP':
                ProductDomain[i] = 'Mobile App'
            elif ProductDomain[i] == 'FWA-4G':
                ProductDomain[i] = 'FWA_PS'
            elif ProductDomain[i] == 'FWA-5G' :
                ProductDomain[i] = 'FWA_PS'
            elif ProductDomain[i] == 'Beacon Mesh':
                ProductDomain[i] = 'Mesh'
            create_data(Releases[i],RCR[i],ProductDomain[i],pType[i],Competence[i],RCRCategories[i],Site[i],Year2[i],Jan2[i],Feb2[i],Mar2[i],Apr2[i],May2[i],Jun2[i],Jul2[i],Aug2[i],Sep2[i],Oct2[i],Nov2[i],Dec2[i],Year3[i],Jan3[i],Feb3[i],Mar3[i],Apr3[i],May3[i],Jun3[i],Jul3[i],Aug3[i],Sep3[i],Oct3[i],Nov3[i],Dec3[i],SQLConn,SQLBResult)
        
        k += limit
    rowx = nrows-1     
    SQLConn.close()    
    return rowx

def create_data(sRelease,sRCR,sProductDomain,pType,sCompetence,sRCRCategories,sSite,sYear2,sJans2,sFebs2,sMars2,sAprs2,sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,sYear3,sJans3,sFebs3,sMars3,sAprs3,sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3,SQLConn,SQLBResult):
    
    flag = 0
    if pType == 'FD':
        sql = """
                SELECT                    
                  a.ID,b.ID,Year1,Year2,Year3,Site,Competence,Description,State,
                  ProjectNumber,ProjectDescription,Businessline,Effortjira,Releases
                FROM
                cdb_rd_resource a
                JOIN
                  cdb_rd_effort b
                WHERE  
                  a.ID = Right(b.ID,8) 
                  AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' 
                  AND Type = '%s' AND Year1 = '%s'
                """  % (sProductDomain,sRCR,sCompetence,pType,sYear2)
           
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
        print('infor =',sProductDomain,sRCR,sSite,sCompetence,pType,sYear2,file=fa, flush=True )
        for row in SQLResult:
            ID = row[0]
            YID= row[1]
            # Y1 = row[2]
            # Y2 = row[3]
            # Y3 = row[4]
            Site = row[5]
            Competence =row[6]
            sDescription = row[7]        
            sState = row[8]        
            sProjectNumber = row[9]
            sProjectDescription = row[10]
            sBusinessLine = row[11]
            sEffort = row[12]
            sRelease = row[13]
            if (Site == '' or (Competence == 'HW' and sSite == Site)) and pType == 'FD':
                flag = 1
                sql2="""
                       UPDATE cdb_rd_effort set Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
                        Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
                        Novs= '%s',Decs = '%s'
                        WHERE ID = '%s'           
                    """ % (sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2,YID)                                   
                # print('A1 update =',YID, sRCR,Site,Competence,file=fa, flush=True ) 
                SQLConn.cur.execute(sql2) 
                if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 != '':
                    InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                    YID2 = sYear3 + ID
                    values = (YID2, sYear3, sJans3, sFebs3,sMars3, sAprs3, sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3)
                    # print('A2 insert =',YID2, sRCR, Site,Competence,file=fa, flush=True )
                    try:
                        SQLConn.cur.execute(InUpSql2, values)
                    except Exception as e:
                        print('A2 insert error =',e, sRCR,file=fa, flush=True )     
                else:
                    sYear3 =''
                sql="""
                    UPDATE cdb_rd_resource set Site = '%s', Year2 = '%s'
                    WHERE ID = '%s'           
                    """ % (sSite,sYear3,ID)
                SQLConn.cur.execute(sql)
                SQLConn.conn.commit()
                
            elif Site != '' and Competence == 'HW' and sSite != Site and pType == 'FD':
                 flag = 1
                 tblname = 'cdb_rd_resource'    
                 stN = tbl_index(tblname,SQLConn)
                 sTD = strnum(stN)
                 if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 == '':
                    sYear3 ='' 
                 InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Effortjira,Year1,Year2) VALUES 
                 (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                 values = (sTD,sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,sEffort,sYear2,sYear3)
                 # print('HW INSERT=',sTD,sRCR,sSite,sCompetence,file=fa, flush=True )
                 SQLConn.cur.execute(InUpSql, values)
                 
                 InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                 (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                 YID2 = sYear2 + sTD
                 values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2)
                 # print('A5 INSERT=',YID2,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                 SQLConn.cur.execute(InUpSql2, values2)
                 if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 != '':
                     InUpSql3 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                     (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                     YID3 = sYear3 + sTD
                     values3 = (YID3, sYear3, sJans3, sFebs3,sMars3, sAprs3, sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3)
                     # print('A6 INSERT=',YID3,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                     SQLConn.cur.execute(InUpSql3, values3)
                 SQLConn.conn.commit()
                 break
            elif Site != sSite and Competence != 'HW' and pType == 'FD':
                # print('infor1 =',sProductDomain,sRCR,Site,sSite,file=fa, flush=True )
                flag = 1
                tblname = 'cdb_rd_resource'    
                stN = tbl_index(tblname,SQLConn)
                sTD = strnum(stN)
                if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 == '':
                   sYear3 ='' 
                InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Effortjira,Year1,Year2) VALUES 
                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                values = (sTD,sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,sEffort,sYear2,sYear3)
                # print('SW INSERT=',sTD,sRCR,sSite,sCompetence,file=fa, flush=True )
                SQLConn.cur.execute(InUpSql, values)
                
                InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                YID2 = sYear2 + sTD
                values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2)
                # print('A3 INSERT=',YID2,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                SQLConn.cur.execute(InUpSql2, values2)
                if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 != '':
                    InUpSql3 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                    YID3 = sYear3 + sTD
                    values3 = (YID3, sYear3, sJans3, sFebs3,sMars3, sAprs3, sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3)
                    # print('A4 INSERT=',YID3,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                    SQLConn.cur.execute(InUpSql3, values3)
                SQLConn.conn.commit()
                break
        if flag == 0:
            sql = """
                    SELECT                    
                      count(ID) as num   
                    FROM
                    cdb_rd_resource                    
                    WHERE                     
                      ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' 
                      AND Type = '%s' AND Site = '%s' AND Year1 = '%s'
                    """  % (sProductDomain,sRCR,sCompetence,pType,sSite,sYear2)               
            SQLConn.cur.execute(sql)      
            SQLResult = SQLConn.cur.fetchall()
            count = SQLResult[0][0]
            if count == 0:
               sDescription = ''        
               sState = ''       
               sProjectNumber = ''
               sProjectDescription = ''
               sBusinessLine = ''
               sEffort = ''
               flag = 1 
               tblname = 'cdb_rd_resource'    
               stN = tbl_index(tblname,SQLConn)
               sTD = strnum(stN)
               if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 == '':
                  sYear3 ='' 
               InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Effortjira,Year1,Year2) VALUES 
               (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
               values = (sTD,sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,sEffort,sYear2,sYear3)
               print('SW INSERT 1=',sTD,sRCR,sSite,sCompetence,file=fa, flush=True )
               SQLConn.cur.execute(InUpSql, values)
               
               InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
               (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
               YID2 = sYear2 + sTD
               values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2)
               print('A7 INSERT=',YID2,sRCR,sSite,sCompetence, file=fa, flush=True )                    
               SQLConn.cur.execute(InUpSql2, values2)
               if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 != '':
                   InUpSql3 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                   (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                   YID3 = sYear3 + sTD
                   values3 = (YID3, sYear3, sJans3, sFebs3,sMars3, sAprs3, sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3)
                   print('A8 INSERT=',YID3,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                   SQLConn.cur.execute(InUpSql3, values3)
               SQLConn.conn.commit() 
    else:
        sql = """
                SELECT                    
                  a.ID,b.ID,Year1,Year2,Year3,Site,Competence,Description,State,
                  ProjectNumber,ProjectDescription,Businessline,Effortjira,RCR
                FROM
                cdb_rd_resource a
                JOIN
                  cdb_rd_effort b
                WHERE  
                  a.ID = Right(b.ID,8) 
                  AND ProductDomain = '%s' AND RCR = '%s'
                """  % (sProductDomain,sRCR)
        SQLConn.cur.execute(sql)      
        SQLResult = SQLConn.cur.fetchall()
                
        for row in SQLResult:
            RCR = row[13] 
            Site =row[5]
            Competence =row[6]
            sDescription = row[7]        
            sState = row[8]        
            sProjectNumber = row[9]
            sProjectDescription = row[10]
            sBusinessLine = row[11]            
            sEffort = row[12]
            if RCR == sRCR:
                
                sql = """
                        SELECT                    
                          count(RCR) as num
                        FROM
                        cdb_rd_resource a
                        JOIN
                          cdb_rd_effort b
                        WHERE  
                          a.ID = Right(b.ID,8) 
                          AND ProductDomain = '%s' AND RCR = '%s' AND Competence = '%s' 
                          AND Type = '%s' AND Site = '%s'
                        """  % (sProductDomain,sRCR,sCompetence,pType,sSite)
                   
                SQLConn.cur.execute(sql)      
                SQLResult = SQLConn.cur.fetchall()
                count = SQLResult[0][0]                
                if count == 0 :
                    flag = 1
                    tblname = 'cdb_rd_resource'    
                    stN = tbl_index(tblname,SQLConn)
                    sTD = strnum(stN)
                    if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 == '':
                       sYear3 ='' 
                    InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Effortjira,Year1,Year2) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
                    values = (sTD,sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,sEffort,sYear2,sYear3)
                    # print('STD 2=',sTD,sRCR, sSite,sCompetence,file=fa, flush=True )
                    SQLConn.cur.execute(InUpSql, values)
                    
                    InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                    YID2 = sYear2 + sTD
                    values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2)
                    # print('A5 INSERT 2=',YID2,sRCR, sSite,sCompetence,file=fa, flush=True )                    
                    SQLConn.cur.execute(InUpSql2, values2)
                    if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 != '':
                        InUpSql3 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                        YID3 = sYear3 + sTD
                        values3 = (YID3, sYear3, sJans3, sFebs3,sMars3, sAprs3, sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3)
                        # print('A6 INSERT=',YID3,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                        SQLConn.cur.execute(InUpSql3, values3)
                    SQLConn.conn.commit()
                    
    if flag == 0:
       flag_r = 0      
       for row in SQLBResult:
           if sRCR == row[0]:
               flag_r = 1
               sRelease = row[4]
               sDescription = row[1]
               sState = row[2]
               sProjectNumber = row[9]
               sProjectDescription = ''
               sBusinessLine = row[3]               
               SubTasks = row[5][:100]
               if row[6] !='' and row[6] !='0.0' and row[6] != '0':
                   sEffort  = row[6]
               else:
                   sEffort  = row[7]
               tblname = 'cdb_rd_resource'    
               stN = tbl_index(tblname,SQLConn)
               sTD = strnum(stN)
               if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 == '':
                  sYear3 ='' 
               InUpSql = """INSERT INTO cdb_rd_resource (ID,Releases,RCR,Description,ProjectNumber,ProjectDescription,Businessline,ProductDomain,Type,Site,Competence,RCRCategories,Effortjira,Year1,Year2) VALUES 
               (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
               values = (sTD,sRelease,sRCR,sDescription,sProjectNumber,sProjectDescription,sBusinessLine,sProductDomain,pType,sSite,sCompetence,sRCRCategories,sEffort,sYear2,sYear3)
               print('PRE INSERT 1=',sTD,sRCR, sSite,sCompetence,file=fa, flush=True )
               SQLConn.cur.execute(InUpSql, values)
               
               InUpSql2 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
               (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
               YID2 = sYear2 + sTD
               values2 = (YID2, sYear2, sJans2, sFebs2,sMars2, sAprs2, sMays2,sJuns2,sJuls2,sAugs2,sSeps2,sOcts2,sNovs2,sDecs2)
               print('PRE INSERT 2=',YID2,sRCR, sSite,sCompetence,file=fa, flush=True )                    
               SQLConn.cur.execute(InUpSql2, values2)
               if sJans3+sFebs3+sMars3+sAprs3+sMays3+sJuns3+sJuls3+sAugs3+sSeps3+sOcts3+sNovs3+sDecs3 != '':
                   InUpSql3 = """INSERT INTO cdb_rd_effort (ID,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
                   (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """                    
                   YID3 = sYear3 + sTD
                   values3 = (YID3, sYear3, sJans3, sFebs3,sMars3, sAprs3, sMays3,sJuns3,sJuls3,sAugs3,sSeps3,sOcts3,sNovs3,sDecs3)
                   print('A6 INSERT=',YID3,sRCR,sSite,sCompetence, file=fa, flush=True )                    
                   SQLConn.cur.execute(InUpSql3, values3)
               SQLConn.conn.commit()
       if flag_r == 0:
           print('RCR not found =',sRCR,sSite,sRelease,sProductDomain,pType,sCompetence,file=fa, flush=True ) 
       
           
def hc_budget_upload(request):    
    file = request.FILES.get('file')
    # print('uplaod:%s'% file,file=fa,flush=True)
    # 创建upload文件夹
    if not os.path.exists(settings.UPLOAD_ROOT):
        os.makedirs(settings.UPLOAD_ROOT)
    try:
        if file is None:
            return HttpResponse('Please select the file you want to upload')
        # 循环二进制写入
        with open(settings.UPLOAD_ROOT + "/" + file.name, 'wb') as f:
            for i in file.readlines():
                f.write(i)

        # 写入 mysql
        filename = settings.UPLOAD_ROOT + "/" + file.name
        budget_insert_db(filename)
    except Exception as e:
        return HttpResponse(e) 
    return HttpResponse('Update successful')


def budget_insert_db(filename):    
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    # ncols = sheet.ncols 
    
    try: 
        while sheet.cell_value(nrows-1, 3) == "":    
            nrows -= 1
    except: 
        print('nrows error =',nrows,file=fa,flush=True)
    # print('nrows new =',nrows,file=fa,flush=True)
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    SQLConn = analyzer_db()    
    SQLConn.cur.execute('truncate table cdb_hc_budget')          
        
    limitall = 50000
    k = 1
    tblname = 'cdb_hc_budget'    
    stN = tbl_index(tblname,SQLConn)
    # print('nrows =',nrows,stN,file=fa,flush=True)
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]
        rowvalues1 =[]
        rowvalues2 =[]
        
        ID = [strnum(stN+i) for i in range(limit)]            
        Country = [str(sheet.cell_value(i+k, 0)) for i in range(limit)]
        City = [str(sheet.cell_value(i+k, 1)) for i in range(limit)]
        BU  = [str(sheet.cell_value(i+k, 2)) for i in range(limit)]
        DU = [str(sheet.cell_value(i+k, 3)) for i in range(limit)]
        DomainPL = [str(sheet.cell_value(i+k, 4)) for i in range(limit)]
        AllocatedDU = [str(sheet.cell_value(i+k, 5)) for i in range(limit)]
        Team = [str(sheet.cell_value(i+k, 6)) for i in range(limit)]
        HCType1 = [str(sheet.cell_value(i+k, 7)) for i in range(limit)]
        HCType2 = [str(sheet.cell_value(i+k, 8)) for i in range(limit)]
        BudgetRollup = [str(sheet.cell_value(i+k, 9)) for i in range(limit)]        
        Year1 = ['2022' for i in range(limit)]                
        Jan1 = [str(sheet.cell_value(i+k, 18)) for i in range(limit)]
        Feb1 = [str(sheet.cell_value(i+k, 19)) for i in range(limit)]
        Mar1 = [str(sheet.cell_value(i+k, 20)) for i in range(limit)]
        Apr1 = [str(sheet.cell_value(i+k, 21)) for i in range(limit)]
        May1 = [str(sheet.cell_value(i+k, 22)) for i in range(limit)]
        Jun1 = [str(sheet.cell_value(i+k, 23)) for i in range(limit)]
        Jul1 = [str(sheet.cell_value(i+k, 24)) for i in range(limit)]
        Aug1 = [str(sheet.cell_value(i+k, 25)) for i in range(limit)]
        Sep1 = [str(sheet.cell_value(i+k, 26)) for i in range(limit)]
        Oct1 = [str(sheet.cell_value(i+k, 27)) for i in range(limit)]
        Nov1 = [str(sheet.cell_value(i+k, 28)) for i in range(limit)]
        Dec1 = [str(sheet.cell_value(i+k, 29)) for i in range(limit)]
        ID1 = [strnum(limit +stN+i) for i in range(limit)]  
        Year2 = ['2023' for i in range(limit)] 
        Jan2 = [str(sheet.cell_value(i+k, 30)) for i in range(limit)]
        Feb2 = [str(sheet.cell_value(i+k, 31)) for i in range(limit)]
        Mar2 = [str(sheet.cell_value(i+k, 32)) for i in range(limit)]
        Apr2 = [str(sheet.cell_value(i+k, 33)) for i in range(limit)]
        May2 = [str(sheet.cell_value(i+k, 34)) for i in range(limit)]
        Jun2 = [str(sheet.cell_value(i+k, 35)) for i in range(limit)]
        Jul2 = [str(sheet.cell_value(i+k, 36)) for i in range(limit)]
        Aug2 = [str(sheet.cell_value(i+k, 37)) for i in range(limit)]
        Sep2 = [str(sheet.cell_value(i+k, 38)) for i in range(limit)]
        Oct2 = [str(sheet.cell_value(i+k, 39)) for i in range(limit)]
        Nov2 = [str(sheet.cell_value(i+k, 40)) for i in range(limit)]
        Dec2 = [str(sheet.cell_value(i+k, 41)) for i in range(limit)]
        ID2 = [ strnum(2*limit + stN+i) for i in range(limit)]  
        Year3 = ['2024' for i in range(limit)]  
        Jan3 = [str(sheet.cell_value(i+k, 42)) for i in range(limit)]
        Feb3 = [str(sheet.cell_value(i+k, 43)) for i in range(limit)]
        Mar3 = [str(sheet.cell_value(i+k, 44)) for i in range(limit)]
        Apr3 = [str(sheet.cell_value(i+k, 45)) for i in range(limit)]
        May3 = [str(sheet.cell_value(i+k, 46)) for i in range(limit)]
        Jun3 = [str(sheet.cell_value(i+k, 47)) for i in range(limit)]
        Jul3 = [str(sheet.cell_value(i+k, 48)) for i in range(limit)]
        Aug3 = [str(sheet.cell_value(i+k, 49)) for i in range(limit)]
        Sep3 = [str(sheet.cell_value(i+k, 50)) for i in range(limit)]
        Oct3 = [str(sheet.cell_value(i+k, 51)) for i in range(limit)]
        Nov3 = [str(sheet.cell_value(i+k, 52)) for i in range(limit)]
        Dec3 = [str(sheet.cell_value(i+k, 53)) for i in range(limit)]
        k+=1        
        for i in range(limit):
            if DomainPL[i] == 'RGW(ONT)':
                DomainPL[i] = 'ONT_PS'
            elif DomainPL[i] == 'SFU':
                DomainPL[i] = 'ONT_PS'
            elif DomainPL[i] == 'ONT':
                DomainPL[i] = 'ONT_PS'
            elif DomainPL[i] == 'Voice(ONT)' :
                DomainPL[i] = 'Voice'
            elif DomainPL[i] == 'Voice(FWA)' :
                DomainPL[i] = 'Voice'
            elif DomainPL[i] == 'FWA' :
                DomainPL[i] = 'FWA_PS'
            elif DomainPL[i] == 'Beacon':
                DomainPL[i] = 'Beacon'
            elif DomainPL[i] == 'Beacon Cloud':
                DomainPL[i] = 'Cloud'
            elif DomainPL[i] == 'Beacon MAPP':
                DomainPL[i] = 'Mobile App'
            elif DomainPL[i] == 'FWA-4G':
                DomainPL[i] = 'FWA_PS'
            elif DomainPL[i] == 'FWA-5G' :
                DomainPL[i] = 'FWA_PS'
            elif DomainPL[i] == 'Beacon Mesh':
                DomainPL[i] = 'Mesh'
            elif DomainPL[i] == 'Container Apps':
                DomainPL[i] = 'Container App' 
            rowvalues.append((ID[i],Country[i].strip(),City[i].strip(),BU[i].strip(),DU[i].strip(),DomainPL[i].strip(),AllocatedDU[i].strip(),Team[i].strip(),HCType1[i].strip(),HCType2[i].strip(),BudgetRollup[i].strip(),Year1[i],Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip()))
            rowvalues1.append((ID1[i],Country[i].strip(),City[i].strip(),BU[i].strip(),DU[i].strip(),DomainPL[i].strip(),AllocatedDU[i].strip(),Team[i].strip(),HCType1[i].strip(),HCType2[i].strip(),BudgetRollup[i].strip(),Year2[i],Jan2[i].strip(),Feb2[i].strip(),Mar2[i].strip(),Apr2[i].strip(),May2[i].strip(),Jun2[i].strip(),Jul2[i].strip(),Aug2[i].strip(),Sep2[i].strip(),Oct2[i].strip(),Nov2[i].strip(),Dec2[i].strip()))
            rowvalues2.append((ID2[i],Country[i].strip(),City[i].strip(),BU[i].strip(),DU[i].strip(),DomainPL[i].strip(),AllocatedDU[i].strip(),Team[i].strip(),HCType1[i].strip(),HCType2[i].strip(),BudgetRollup[i].strip(),Year3[i],Jan3[i].strip(),Feb3[i].strip(),Mar3[i].strip(),Apr3[i].strip(),May3[i].strip(),Jun3[i].strip(),Jul3[i].strip(),Aug3[i].strip(),Sep3[i].strip(),Oct3[i].strip(),Nov3[i].strip(),Dec3[i].strip()))
            
        InUpSql = """INSERT INTO cdb_hc_budget (ID,Country,City,BU,DU,DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """               
        
        SQLConn.cur.executemany(InUpSql,rowvalues)
        SQLConn.cur.executemany(InUpSql,rowvalues1)
        SQLConn.cur.executemany(InUpSql,rowvalues2)                   
        SQLConn.commit()
        # print('new nrows =',nrows, str(h), str(t), str(k), str(limit),file=fa,flush=True)
        k += limit
    rowx = nrows-1     
    SQLConn.close()    
    return rowx

def hc_budget_update(request):    
    file = request.FILES.get('file')
    print('uplaod:%s'% file,file=fa,flush=True)
    # 创建upload文件夹
    if not os.path.exists(settings.UPLOAD_ROOT):
        os.makedirs(settings.UPLOAD_ROOT)
    try:
        if file is None:
            return HttpResponse('Please select the file you want to upload')
        # 循环二进制写入
        with open(settings.UPLOAD_ROOT + "/" + file.name, 'wb') as f:
            for i in file.readlines():
                f.write(i)

        # 写入 mysql
        filename = settings.UPLOAD_ROOT + "/" + file.name
        hc_update_db(filename)
    except Exception as e:
        return HttpResponse(e) 
    return HttpResponse('Upload successful')

def hc_update_db2(filename):
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    # print('nrows =',nrows,file=fa,flush=True)
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    
    cmd = """
            SELECT
               Country, City, BU,DU, DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,
               Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,ID
            FROM
              cdb_hc_budget
                        
            """
    
    SQLConn = analyzer_db() 
    SQLConn.cur.execute(cmd)
    SQLResult = SQLConn.cur.fetchall()
    tblname = 'cdb_hc_budget'    
    stN = tbl_index(tblname,SQLConn)
    
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)] 
    
    col_site = Title.index('Site')
    print('Title =',Title,file=fa,flush=True)
    col_Country = Title.index('Country')
    col_Site = Title.index('Site')
    col_BU = Title.index('BU')
    col_DU = Title.index('DU')
    col_DomainPL = Title.index('TeachArea')
    col_Type = Title.index('Type')
    col_Category = Title.index('Category') 
    col_Rollup = Title.index('Budget/Rollup')
    col_Jan = Title.index('Jan_23')
    col_Feb = Title.index('Feb_23')
    col_Mar = Title.index('Mar_23')
    col_Apr = Title.index('Apr_23')
    col_May = Title.index('May_23')
    col_Jun = Title.index('Jun_23')
    col_Jul = Title.index('Jul_23')
    col_Aug = Title.index('Aug_23')
    col_Sep = Title.index('Sep_23')
    col_Oct = Title.index('Oct_23')
    col_Nov = Title.index('Nov_23')
    col_Dec = Title.index('Dec_23')
    
    print('cols =',str(col_Country),str(col_site),str(col_DomainPL), str(col_Jan), str(col_Dec), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    d = 0
    e = 0
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]        
        Country = [str(sheet.cell_value(i+k, col_Country)) for i in range(limit)] 
        Site = [str(sheet.cell_value(i+k, col_Site)) for i in range(limit)]              
        BU = [str(sheet.cell_value(i+k, col_BU)) for i in range(limit)] 
        DU = [str(sheet.cell_value(i+k, col_DU)) for i in range(limit)] 
        DomainPL = [str(sheet.cell_value(i+k, col_DomainPL)) for i in range(limit)] 
        Type = [str(sheet.cell_value(i+k, col_Type).replace(' HC','')) for i in range(limit)] 
        Category = [str(sheet.cell_value(i+k, col_Category)) for i in range(limit)] 
        BudgetRollup = [str(sheet.cell_value(i+k, col_Rollup)) for i in range(limit)] 
        
        AllocatedDU = [str(sheet.cell_value(i+k, col_DU)) for i in range(limit)]
        HCType2 = ['Own' for i in range(limit)]        
        Year1 = ['2023' for i in range(limit)]
        Flag = ['No' for i in range(limit)]
        
        Jan1 = [str(sheet.cell_value(i+k, col_Jan)) for i in range(limit)]
        Feb1 = [str(sheet.cell_value(i+k, col_Feb)) for i in range(limit)]
        Mar1 = [str(sheet.cell_value(i+k, col_Mar)) for i in range(limit)]
        Apr1 = [str(sheet.cell_value(i+k, col_Apr)) for i in range(limit)]
        May1 = [str(sheet.cell_value(i+k, col_May)) for i in range(limit)]
        Jun1 = [str(sheet.cell_value(i+k, col_Jun)) for i in range(limit)]
        Jul1 = [str(sheet.cell_value(i+k, col_Jul)) for i in range(limit)]
        Aug1 = [str(sheet.cell_value(i+k, col_Aug)) for i in range(limit)]
        Sep1 = [str(sheet.cell_value(i+k, col_Sep)) for i in range(limit)]
        Oct1 = [str(sheet.cell_value(i+k, col_Oct)) for i in range(limit)]
        Nov1 = [str(sheet.cell_value(i+k, col_Nov)) for i in range(limit)]
        Dec1 = [str(sheet.cell_value(i+k, col_Dec)) for i in range(limit)]        
        
        for i in range(limit):        
            if Flag[i] == 'No' :
                ID = strnum(stN+e) 
                e += 1 
                print('New hc = ',Country[i],Site[i],DomainPL[i],Category[i],Type[i],file=fa,flush=True )
                rowvalues.append((ID,Country[i].strip(),Site[i].strip(),BU[i].strip(),DU[i].strip(),DomainPL[i].strip(),AllocatedDU[i].strip(),Category[i].strip(),Type[i].strip(),HCType2[i].strip(),BudgetRollup[i].strip(),Year1[i],Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip()))
            
            
        InUpSql = """INSERT INTO cdb_hc_budget (ID,Country,City,BU,DU,DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """               
        
        SQLConn.cur.executemany(InUpSql,rowvalues)                          
        SQLConn.commit()
        print('update nrows =',nrows, str(d), str(e), str(k),sLastupdate,file=fa,flush=True)
        k += limit
    rowx = nrows-1    
    SQLConn.close()
    return rowx

def hc_update_db1(filename):
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    # print('nrows =',nrows,file=fa,flush=True)
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    
    cmd = """
            SELECT
               Country, City, BU,DU, DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,
               Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,ID
            FROM
              cdb_hc_budget
                        
            """
    
    SQLConn = analyzer_db() 
    SQLConn.cur.execute(cmd)
    SQLResult = SQLConn.cur.fetchall()
    tblname = 'cdb_hc_budget'    
    stN = tbl_index(tblname,SQLConn)
    
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)] 
    
    col_site = Title.index('Site')
    print('Title =',Title,file=fa,flush=True)
    col_Country = Title.index('Country')
    col_Site = Title.index('Site')
    col_BU = Title.index('BU')
    col_DU = Title.index('DU')
    col_DomainPL = Title.index('TeachArea')
    col_Type = Title.index('Type')
    col_Category = Title.index('Category')  
    
    col_Dec = Title.index('Dec_23')
    
    print('cols =',str(col_Country),str(col_site),str(col_DomainPL), str(col_Dec), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    d = 0
    e = 0
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]        
        Country = [str(sheet.cell_value(i+k, col_Country)) for i in range(limit)] 
        Site = [str(sheet.cell_value(i+k, col_Site)) for i in range(limit)]              
        BU = [str(sheet.cell_value(i+k, col_BU)) for i in range(limit)] 
        DU = [str(sheet.cell_value(i+k, col_DU)) for i in range(limit)] 
        DomainPL = [str(sheet.cell_value(i+k, col_DomainPL)) for i in range(limit)] 
        Type = [str(sheet.cell_value(i+k, col_Type).replace(' HC','')) for i in range(limit)] 
        Category = [str(sheet.cell_value(i+k, col_Category)) for i in range(limit)]    
        
        AllocatedDU = [str(sheet.cell_value(i+k, col_DU)) for i in range(limit)]
        HCType2 = ['Own' for i in range(limit)]
        BudgetRollup = ['Rollup' for i in range(limit)]
        Year1 = ['2023' for i in range(limit)]
        Flag = ['No' for i in range(limit)]       
       
        Dec1 = [str(sheet.cell_value(i+k, col_Dec)) for i in range(limit)]        
                
        for i in range(limit):
            if DomainPL[i] == 'ONT_PS&L':
                DomainPL[i] = 'ONT_PS'            
            elif DomainPL[i] == 'Voice(FWA)' :
                DomainPL[i] = 'Voice'                         
            elif DomainPL[i] == 'FWA_PS&L' :
                DomainPL[i] = 'FWA_PS'
            elif DomainPL[i] == 'Beacon Mesh':
                DomainPL[i] = 'Mesh'
            elif DomainPL[i] == 'Container Apps':
                DomainPL[i] = 'Container App' 
                
            elif DomainPL[i] == 'RGW(ONT)':
                DomainPL[i] = 'ONT_PS'
            elif DomainPL[i] == 'Beacon Cloud':
                DomainPL[i] = 'Cloud'
            elif DomainPL[i] == 'Beacon MAPP':
                DomainPL[i] = 'Mobile App'            
            elif DomainPL[i] == 'FWA-4G' :
                DomainPL[i] = 'FWA_PS'
            elif DomainPL[i] == 'FWA-5G' :
                DomainPL[i] = 'FWA_PS'    
            elif DomainPL[i] == 'Mesh':
                DomainPL[i] = 'Mesh'
            elif DomainPL[i] == 'Voice(ONT)':
                DomainPL[i] = 'Voice'    
            
            for row in SQLResult:
                if Country[i] == row[0] and Site[i] == row[1] and DomainPL[i] == row[4] and Category[i] == row[6] and Type[i] == row[7] and row[10] == Year1[i]:                             
                    Sql="""UPDATE cdb_hc_budget SET Decs = '%s'
                    WHERE 
                          ID= '%s' 
                    """ % (Dec1[i],row[23])  
                    print('Update 2023 = ',row[23],Dec1[i],file=fa,flush=True )
                    print('Update SQL= ',Sql,file=fa,flush=True )
                    SQLConn.cur.execute(Sql)                   
                    SQLConn.commit()
                    Flag[i] = 'Yes'
                    d += 1
                    break 
                
        for i in range(limit):        
            if Flag[i] == 'No' :
                ID = strnum(stN+e) 
                e += 1 
                print('New hc = ',Country[i],Site[i],DomainPL[i],Category[i],Type[i],file=fa,flush=True )
                rowvalues.append((ID,Country[i].strip(),Site[i].strip(),BU[i].strip(),DU[i].strip(),DomainPL[i].strip(),AllocatedDU[i].strip(),Category[i].strip(),Type[i].strip(),HCType2[i].strip(),BudgetRollup[i].strip(),Year1[i],Dec1[i].strip()))
            
            
        InUpSql = """INSERT INTO cdb_hc_budget (ID,Country,City,BU,DU,DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,Year,Decs) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """               
        
        SQLConn.cur.executemany(InUpSql,rowvalues)                          
        SQLConn.commit()
        print('update nrows =',nrows, str(d), str(e), str(k),sLastupdate,file=fa,flush=True)
        k += limit
    rowx = nrows-1    
    SQLConn.close()
    return rowx


def hc_update_db(filename):
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    DueDate = datetime.today() + timedelta(weeks=8)
    sYear = str(DueDate)[:4]
    sYear1 = sYear[-2:]
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
    SQLConn = analyzer_db()            
    sql="""
        UPDATE cdb_hc_budget set Modifier = 'No'
        WHERE Year = '%s'     
        """ % sYear
    SQLConn.cur.execute(sql)
    SQLConn.commit() 
    cmd = """
            SELECT
               Country, City, BU,DU, DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,
               Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,ID
            FROM
              cdb_hc_budget
                        
            """
    SQLConn.cur.execute(cmd)
    SQLResult = SQLConn.cur.fetchall()
    tblname = 'cdb_hc_budget'    
    stN = tbl_index(tblname,SQLConn)
    
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)]
    print('Title =',Title,file=fa,flush=True)
    col_Country = Title.index('Country')
    col_City = Title.index('City')
    col_BU = Title.index('BU')
    col_DU = Title.index('DU')
    col_DomainPL = Title.index('Domain(PL)')
    col_AllocatedDU = Title.index('Allocated DU')
    col_Team = Title.index('Abbreviation(Team)')
    col_HCType1 = Title.index('HC Type1')
    col_HCType2 = Title.index('HC Type2')
    col_BudgetRollup = Title.index('Budget/Rollup')       
    col_Jan = Title.index('Jan_'+sYear1)
    col_Feb = Title.index('Feb_'+sYear1)
    col_Mar = Title.index('Mar_'+sYear1)
    col_Apr = Title.index('Apr_'+sYear1)
    col_May = Title.index('May_'+sYear1)
    col_Jun = Title.index('Jun_'+sYear1)
    col_Jul = Title.index('Jul_'+sYear1)
    col_Aug = Title.index('Aug_'+sYear1)
    col_Sep = Title.index('Sep_'+sYear1)
    col_Oct = Title.index('Oct_'+sYear1)
    col_Nov = Title.index('Nov_'+sYear1)
    col_Dec = Title.index('Dec_'+sYear1)
    Modifier = 'yanhui.zhang@nokia-sbell.com'
    print('cols =',str(col_Country),str(col_City),str(col_DomainPL), str(col_Jan), str(col_Dec), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    d = 0
    e = 0
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]        
        Country = [str(sheet.cell_value(i+k, col_Country)) for i in range(limit)] 
        City = [str(sheet.cell_value(i+k, col_City)) for i in range(limit)]              
        BU = [str(sheet.cell_value(i+k, col_BU)) for i in range(limit)] 
        DU = [str(sheet.cell_value(i+k, col_DU)) for i in range(limit)] 
        DomainPL = [str(sheet.cell_value(i+k, col_DomainPL)) for i in range(limit)]
        AllocatedDU = [str(sheet.cell_value(i+k, col_AllocatedDU)) for i in range(limit)]
        Team = [str(sheet.cell_value(i+k, col_Team)) for i in range(limit)]
        HCType1 = [str(sheet.cell_value(i+k, col_HCType1).replace(' HC','')) for i in range(limit)] 
        HCType2 = [str(sheet.cell_value(i+k, col_HCType2)) for i in range(limit)] 
        BudgetRollup = [str(sheet.cell_value(i+k, col_BudgetRollup)) for i in range(limit)]
        Year1 = [sYear for i in range(limit)]
        Flag = ['No' for i in range(limit)]
        
        Jan1 = [str(sheet.cell_value(i+k, col_Jan)) for i in range(limit)]
        Feb1 = [str(sheet.cell_value(i+k, col_Feb)) for i in range(limit)]
        Mar1 = [str(sheet.cell_value(i+k, col_Mar)) for i in range(limit)]
        Apr1 = [str(sheet.cell_value(i+k, col_Apr)) for i in range(limit)]
        May1 = [str(sheet.cell_value(i+k, col_May)) for i in range(limit)]
        Jun1 = [str(sheet.cell_value(i+k, col_Jun)) for i in range(limit)]
        Jul1 = [str(sheet.cell_value(i+k, col_Jul)) for i in range(limit)]
        Aug1 = [str(sheet.cell_value(i+k, col_Aug)) for i in range(limit)]
        Sep1 = [str(sheet.cell_value(i+k, col_Sep)) for i in range(limit)]
        Oct1 = [str(sheet.cell_value(i+k, col_Oct)) for i in range(limit)]
        Nov1 = [str(sheet.cell_value(i+k, col_Nov)) for i in range(limit)]
        Dec1 = [str(sheet.cell_value(i+k, col_Dec)) for i in range(limit)]        
                
        for i in range(limit):
            if DomainPL[i] == 'ONT_PS&L':
                DomainPL[i] = 'ONT_PS'            
            elif DomainPL[i] == 'China' :
                DomainPL[i] = 'China Product'                        
            elif DomainPL[i] == 'FWA_PS&L' :
                DomainPL[i] = 'FWA_PS'
            elif DomainPL[i] == 'Mesh':
                DomainPL[i] = 'Mesh'
            elif DomainPL[i] == 'Container Apps':
                DomainPL[i] = 'Container App' 
            
            for row in SQLResult:
                if Country[i] == row[0] and City[i] == row[1] and DomainPL[i] == row[4] and Team[i] == row[6] and HCType1[i] == row[7] and HCType2[i] == row[8] and BudgetRollup[i] == row[9] and row[10] == Year1[i]:                             
                    Sql="""UPDATE cdb_hc_budget SET Jans= '%s', Febs = '%s',Mars = '%s',Aprs = '%s',
                     Mays = '%s',Juns = '%s',Juls = '%s',Augs = '%s',Seps = '%s',Octs = '%s',
                     Novs= '%s',Decs = '%s',Modifier = '%s' 
                    WHERE 
                          ID= '%s'
                    """ % (Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip(),Modifier,row[23])  
                    print('Update = ',row[23],file=fa,flush=True )
                    # print('Update SQL= ',Sql,file=fa,flush=True )
                    SQLConn.cur.execute(Sql)                   
                    SQLConn.commit()
                    Flag[i] = 'Yes'
                    d += 1
                    break 
                
        for i in range(limit):        
            if Flag[i] == 'No' :
                ID = strnum(stN+e) 
                e += 1 
                print('New hc = ',Country[i],City[i],DomainPL[i],Team[i],HCType1[i],file=fa,flush=True )
                rowvalues.append((ID,Country[i].strip(),City[i].strip(),BU[i].strip(),DU[i].strip(),DomainPL[i].strip(),AllocatedDU[i].strip(),Team[i].strip(),HCType1[i].strip(),HCType2[i].strip(),BudgetRollup[i].strip(),Year1[i],Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip(),Modifier))
                        
        InUpSql = """INSERT INTO cdb_hc_budget (ID,Country,City,BU,DU,DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,Modifier) VALUES 
        (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """               
        
        SQLConn.cur.executemany(InUpSql,rowvalues)                          
        SQLConn.commit()
        
        print('update nrows =',nrows, str(d), str(e), str(k),sLastupdate,file=fa,flush=True)
        k += limit
    rowx = nrows-1
    sql="""DELETE FROM cdb_hc_budget
    WHERE Modifier = 'No'   
    """                                 
    SQLConn.cur.execute(sql)
    SQLConn.commit()    
    SQLConn.close()
    return rowx

def file_update(request):    
    file = request.FILES.get('file')
    print('uplaod:%s'% file,file=fa,flush=True)
    # 创建upload文件夹
    if not os.path.exists(settings.UPLOAD_ROOT):
        os.makedirs(settings.UPLOAD_ROOT)
    try:
        if file is None:
            return HttpResponse('Please select the file you want to upload')
        # 循环二进制写入
        with open(settings.UPLOAD_ROOT + "/" + file.name, 'wb') as f:
            for i in file.readlines():
                f.write(i)

        # 写入 mysql
        filename = settings.UPLOAD_ROOT + "/" + file.name
        update_db(filename)
    except Exception as e:
        return HttpResponse(e) 
    return HttpResponse('Upload successful')

def update_db(filename):
    ws = xlrd.open_workbook(filename)
    sheet = ws.sheet_by_index(0)    
    nrows = sheet.nrows
    ncols = sheet.ncols 
    # print('nrows =',nrows,file=fa,flush=True)
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['items'] = []
                
    SQLConn = analyzer_db() 
    Title = [str(sheet.cell_value(0, i)) for i in range(ncols)] 
    
    col_site = Title.index('Country')
    print('Title =',Title,file=fa,flush=True)
    col_Jan = Title.index('Jan')
    col_Feb = Title.index('Feb')
    col_Mar = Title.index('Mar')
    col_Apr = Title.index('Apr')
    col_May = Title.index('May')
    col_Jun = Title.index('Jun')
    col_Jul = Title.index('Jul')
    col_Aug = Title.index('Aug')
    col_Sep = Title.index('Sep')
    col_Oct = Title.index('Oct')
    col_Nov = Title.index('Nov')
    col_Dec = Title.index('Dec')
    col_ID = Title.index('ID')
    col_YID = Title.index('YID')
    print('cols =',str(col_ID),str(col_YID), str(col_site), str(col_Jan), str(col_Dec), str(ncols),file=fa,flush=True)
    
    limitall = 60000
    k = 1
    
    if nrows - k > limitall :
        m = divmod(nrows-k,limitall)
        h = int(m[0]) + 1
        tar = int(m[1])
        limit = limitall         
    else:        
        h = 1
        limit = nrows - k
        tar = limit
        
    for t in range(h):
        if t == h-1: 
            limit = tar        
        rowvalues =[]
        rowvalues2 =[]
        ID = [str(sheet.cell_value(i+k, col_ID)) for i in range(limit)] 
        Site = [str(sheet.cell_value(i+k, col_site)) for i in range(limit)]              
        ID1 = [str(sheet.cell_value(i+k, col_YID)) for i in range(limit)]        
        Jan1 = [str(sheet.cell_value(i+k, col_Jan)) for i in range(limit)]
        Feb1 = [str(sheet.cell_value(i+k, col_Feb)) for i in range(limit)]
        Mar1 = [str(sheet.cell_value(i+k, col_Mar)) for i in range(limit)]
        Apr1 = [str(sheet.cell_value(i+k, col_Apr)) for i in range(limit)]
        May1 = [str(sheet.cell_value(i+k, col_May)) for i in range(limit)]
        Jun1 = [str(sheet.cell_value(i+k, col_Jun)) for i in range(limit)]
        Jul1 = [str(sheet.cell_value(i+k, col_Jul)) for i in range(limit)]
        Aug1 = [str(sheet.cell_value(i+k, col_Aug)) for i in range(limit)]
        Sep1 = [str(sheet.cell_value(i+k, col_Sep)) for i in range(limit)]
        Oct1 = [str(sheet.cell_value(i+k, col_Oct)) for i in range(limit)]
        Nov1 = [str(sheet.cell_value(i+k, col_Nov)) for i in range(limit)]
        Dec1 = [str(sheet.cell_value(i+k, col_Dec)) for i in range(limit)]        
    
        for i in range(limit):
            rowvalues.append((Site[i].strip(),ID[i]))
            rowvalues2.append((Jan1[i].strip(),Feb1[i].strip(),Mar1[i].strip(),Apr1[i].strip(),May1[i].strip(),Jun1[i].strip(),Jul1[i].strip(),Aug1[i].strip(),Sep1[i].strip(),Oct1[i].strip(),Nov1[i].strip(),Dec1[i].strip(),ID1[i]))
            
        Sql="""UPDATE cdb_rd_resource SET Site = %s
        WHERE ID = %s""" 
        Sql2="""UPDATE cdb_rd_effort SET Jans=%s,Febs=%s,Mars=%s,Aprs=%s,Mays=%s,Juns=%s,Juls=%s,Augs=%s,Seps=%s,Octs=%s,Novs=%s,Decs=%s
        WHERE ID = %s"""  
        # print(Sql, file=fa,flush=True )     
        SQLConn.executemany(Sql,rowvalues)
        SQLConn.executemany(Sql2,rowvalues2)
        SQLConn.commit()
        print('update nrows =',nrows, str(h), str(t), str(k), str(limit),sLastupdate,file=fc,flush=True)
        k += limit
    rowx = nrows-1    
    SQLConn.close()
    return rowx


def hc_budget(request):
    try:
        sUsername = request.GET['username']
        sGrade = request.GET['grade']
        sLevel = request.GET['level']
    except:
        return HttpResponse('Invalid Parameters', content_type='application/json')
    
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['column'] = []
    dResult['data']['items'] = []   
    SQLConn = pymysql.connect(host    = settings.RES_DB['host'],
                            port     = settings.RES_DB['port'],
                            user     = settings.RES_DB['username'],
                            password = settings.RES_DB['password'],
                            database = settings.RES_DB['name'],
                            charset  = settings.RES_DB['charset'],
                            autocommit = True
                            )
    cmd = """
            SELECT
               Country, City, BU,DU, DomainPL,AllocatedDU,Team,HCType1,HCType2,BudgetRollup,
               Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,ID,
               Modifier,RecordTime
            FROM
              cdb_hc_budget
                        
            """
    
    # SQLConn = analyzer_db() 
    # SQLConn.cur.execute(cmd)
    # SQLResult = SQLConn.cur.fetchall()
    SQLCur = SQLConn.cursor()
    SQLCur.execute(cmd)
    SQLResult = SQLCur.fetchall()
    SQLConn.close()

    for row in SQLResult:
        dItem = {}
        dItem['Country'] = row[0]
        dItem['City'] = row[1]
        dItem['BU'] = row[2]
        dItem['DU'] = row[3]
        dItem['DomainPL'] = row[4]
        dItem['AllocatedDU'] = row[5]
        dItem['Team'] = row[6]
        dItem['HCType1'] = row[7]
        dItem['HCType2'] = row[8]
        dItem['BudgetRollup'] = row[9]
        dItem['Year'] = row[10]        
        dItem['Jans'] = row[11]
        dItem['Febs'] = row[12]
        dItem['Mars'] = row[13]
        dItem['Aprs'] = row[14]
        dItem['Mays'] = row[15]
        dItem['Juns'] = row[16]
        dItem['Juls'] = row[17]
        dItem['Augs'] = row[18]
        dItem['Seps'] = row[19]
        dItem['Octs'] = row[20]
        dItem['Novs'] = row[21]
        dItem['Decs'] = row[22]
        dItem['Modifier'] = row[24]
        dItem['RecordTime'] = str(row[25])        
        dItem['ID'] = row[23]        
        
        dResult['data']['items'].append(dItem)
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')

def hc_budget_edit(request):
    try:
        sType = request.GET['type'] 
        # sGrade = request.GET['grade']
        #Release, Pgroup, CMQ, SWbuild
        if sType == '2':
            sFWAID = request.GET['FWAID']
        elif sType == '3':
            sFWAID = request.GET['FWAID'] 
        if sType == '1' or sType == '2' :
            sIDType = 'FWA'
            sRelease = request.GET['Release']
            sProduct = request.GET['Product']
            sSWbuild = request.GET['SWbuild']
            sProductGroup = request.GET['ProductGroup'] 
            sRecordTime = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
            
    except:
        return HttpResponse('Invalid Parameters', content_type='application/json')
    
    
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['status'] = []
            
    # 1 add
    if sType == '1':         
        SQLConn = analyzer_db()
        tblname ='file_issues_sw'
        stN = tbl_filtered_index(tblname,'FWA',SQLConn) 
        sFWAID = strNum(stN, "FWA")  
        
        sqlt="""insert into file_issues_sw (ID, Type, `Release`, Product, SWbuild, Pgroup, RecordTime) 
            values(%s,%s,%s,%s,%s,%s,%s)""" 
        values = (sFWAID, sIDType, sRelease, sProduct, sSWbuild, sProductGroup, sRecordTime)
        SQLConn.cur.execute(sqlt, values)
        SQLConn.conn.commit() 
        SQLConn.conn.close()  
        dResult['data']['status']="successful" 
        
        
    # 2 edit
    elif sType == '2': 
        SQLConn = analyzer_db()
        sql="""
            UPDATE file_issues_sw set `Release`= '%s', Product= '%s', SWbuild= '%s', Pgroup= '%s', RecordTime= '%s'
            WHERE ID = '%s'
            ###### updated til here
            """ % (sRelease, sProduct, sSWbuild, sProductGroup, sRecordTime, sFWAID)
        SQLConn.cur.execute(sql)
        SQLConn.conn.commit() 
        SQLConn.conn.close()  
        dResult['data']['status']="successful" 
    
    # 3 delete
    elif sType == '3': 
        lNTID = sFWAID.split(',')
        SQLConn = analyzer_db()  
        sql="DELETE FROM file_issues_sw WHERE ID IN %s" % List2String(lNTID)        
        SQLConn.cur.execute(sql)
        SQLConn.conn.commit() 
        SQLConn.conn.close()  
        dResult['data']['status']="successful"
        
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')



def parseStr(s):
    if s is None or len(s) == 0:
        return None
    idx1 = s.find('_')
    if idx1 == -1:
        return s
    idx2 = s.find('_', idx1 + 1)
    if idx2 == -1:
        return s[idx1 + 1:]
    return s[idx1 + 1:idx2]


def send_mail(mailadd_list,filename,file_path):
    # sLastupdate = datetime.today().strftime('%Y-%m-%d')
    sLastupdate = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    host_server = '172.24.146.133'  # MAIl server 地址
    sender = "rd-resource-notice@nokia.com"
    

    mail_title = "RD Resource Allocation Notification" 
            
    mail_content = """
    <head>
    
    <style>
    <!--
    /* Font Definitions */
    @font-face
        {font-family:"Cambria Math";
        panose-1:2 4 5 3 5 4 6 3 2 4;
        mso-font-charset:0;
        mso-generic-font-family:roman;
        mso-font-pitch:variable;
        mso-font-signature:3 0 0 0 1 0;}
    @font-face
        {font-family:DengXian;
        panose-1:2 1 6 0 3 1 1 1 1 1;
        mso-font-alt:DengXian;
        mso-font-charset:134;
        mso-generic-font-family:auto;
        mso-font-pitch:variable;
        mso-font-signature:-1610612033 953122042 22 0 262159 0;}
    @font-face
        {font-family:Calibri;
        panose-1:2 15 5 2 2 2 4 3 2 4;
        mso-font-charset:0;
        mso-generic-font-family:swiss;
        mso-font-pitch:variable;
        mso-font-signature:-469750017 -1073732485 9 0 511 0;}
    @font-face
        {font-family:DengXian;
        panose-1:2 1 6 0 3 1 1 1 1 1;
        mso-font-charset:134;
        mso-generic-font-family:auto;
        mso-font-pitch:variable;
        mso-font-signature:-1610612033 953122042 22 0 262159 0;}
    /* Style Definitions */
    p.MsoNormal, li.MsoNormal, div.MsoNormal
        {mso-style-unhide:no;
        mso-style-qformat:yes;
        mso-style-parent:"";
        margin:0cm;
        mso-pagination:widow-orphan;
        font-size:11.0pt;
        font-family:"Calibri",sans-serif;
        mso-fareast-font-family:DengXian;
        mso-fareast-theme-font:minor-fareast;}
    a:link, span.MsoHyperlink
        {mso-style-priority:99;
        color:#0563C1;
        text-decoration:underline;
        text-underline:single;}
    a:visited, span.MsoHyperlinkFollowed
        {mso-style-noshow:yes;
        mso-style-priority:99;
        color:#954F72;
        text-decoration:underline;
        text-underline:single;}
    p.msonormal0, li.msonormal0, div.msonormal0
        {mso-style-name:msonormal;
        mso-style-unhide:no;
        mso-margin-top-alt:auto;
        margin-right:0cm;
        mso-margin-bottom-alt:auto;
        margin-left:0cm;
        mso-pagination:widow-orphan;
        font-size:11.0pt;
        font-family:"Calibri",sans-serif;
        mso-fareast-font-family:DengXian;
        mso-fareast-theme-font:minor-fareast;}
    span.EmailStyle18
        {mso-style-type:personal;
        mso-style-noshow:yes;
        mso-style-unhide:no;
        font-family:"Calibri",sans-serif;
        mso-ascii-font-family:Calibri;
        mso-hansi-font-family:Calibri;
        mso-bidi-font-family:Calibri;
        color:windowtext;}
    .MsoChpDefault
        {mso-style-type:export-only;
        mso-default-props:yes;
        font-size:10.0pt;
        mso-ansi-font-size:10.0pt;
        mso-bidi-font-size:10.0pt;
        mso-ascii-font-family:"Times New Roman";
        mso-fareast-font-family:"Times New Roman";
        mso-hansi-font-family:"Times New Roman";
        mso-font-kerning:0pt;}
    @page WordSection1
        {size:612.0pt 792.0pt;
        margin:72.0pt 90.0pt 72.0pt 90.0pt;
        mso-header-margin:36.0pt;
        mso-footer-margin:36.0pt;
        mso-paper-source:0;}
    div.WordSection1
        {page:WordSection1;}
    -->
    </style>
    <!--[if gte mso 10]>
    <style>
    /* Style Definitions */
    table.MsoNormalTable
        {mso-style-name:\666E\901A\8868\683C;
        mso-tstyle-rowband-size:0;
        mso-tstyle-colband-size:0;
        mso-style-noshow:yes;
        mso-style-priority:99;
        mso-style-parent:"";
        mso-padding-alt:0cm 5.4pt 0cm 5.4pt;
        mso-para-margin:0cm;
        mso-pagination:widow-orphan;
        font-size:10.0pt;
        font-family:"Times New Roman",serif;}
    </style>
    <![endif]-->
    </head>

    <body lang=ZH-CN link="#0563C1" vlink="#954F72" style='tab-interval:21.0pt;
    word-wrap:break-word'>

    <div class=WordSection1>
    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>
    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>
    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <table class=MsoNormalTable border=1 cellspacing=0 cellpadding=0 width="100%"
    style='width:100.0%;mso-cellspacing:0cm;mso-yfti-tbllook:1184;mso-padding-alt:
    0cm 0cm 0cm 0cm'>
    <tr style='mso-yfti-irow:0;mso-yfti-firstrow:yes'>
    <td colspan=4 style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal align=center style='text-align:center'><b><span
    lang=EN-US>RD Resource Allocation Snapshot</span></b><span
    lang=EN-US> </span></p>
    </td>
    </tr>
    <tr style='mso-yfti-irow:1'>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><b><span lang=EN-US>Action</span></b></p>
    </td>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><span lang=EN-US>Description</span></p>
    </td>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><b><span lang=EN-US>File Name</span></b></p>
    </td>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><b><span lang=EN-US>Date<span style='color:#0070C0'></span></span></b></p>
    </td>
    </tr>
    <tr style='mso-yfti-irow:2'>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><i><span lang=EN-US>Data Backup</span></i></p>
    </td>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><b><span lang=EN-US>RD Resource Allocation Snapshot<o:p></o:p></span></b></p>
    </td>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><span lang=EN-US>"""+filename+"""</span></p>
    </td>
    <td style='padding:1.2pt 1.2pt 1.2pt 1.2pt'>
    <p class=MsoNormal><span lang=EN-US>"""+sLastupdate+"""</span></p>
    </td>
    </tr>    
    </table>
    
    <p class=MsoNormal><span lang=EN-US>&nbsp;</span></p>


    <p class=MsoNormal><span lang=EN-US><span style='color:#0070C0'>This is an automatically generated email, please do not reply.</span></span></p>
    </div>

    </body>

    </html>
    """
    # SMTP
    smtp = smtplib.SMTP(host_server)    
    # 发送
    msg = MIMEMultipart()
    msg['Subject'] = mail_title
    msg['From'] = sender
    
    msg.attach(MIMEText(mail_content, 'html', 'utf-8'))
     
    att1 = MIMEText(open(file_path, 'rb').read(), 'xlsx', 'utf-8')
    att1['Content-type'] = 'application/octet-stream'
    att1['Content-Disposition'] = 'attachment; filename= %s' % filename
    msg.attach(att1)
    print('file path :',file_path,file=fd,flush=True)
    if isinstance(mailadd_list,list):
            j=0
            msg['To'] = mailadd_list[0]
            for user in mailadd_list:                
                try:                    
                    msg.replace_header('To', user)
                    print('user = ',user,file=fd,flush=True)
                    smtp.sendmail(sender, user, msg.as_string())
                    print('sendmail to ',user, msg['To'],sLastupdate,file=fd,flush=True)
                    j+=1
                except smtplib.SMTPException as e:
                    print('sendmail to ',user,' error',sLastupdate,file=fd,flush=True)
                    print("smtplib send mail error：", e)                
            print('sendmail =',str(j), sLastupdate,file=fd,flush=True)

def move_file(flname): 
    FILE_ROOT=os.path.join(Path(__file__).resolve().parent.parent)
    os.chdir(FILE_ROOT) 
    path1 = os.path.join(FILE_ROOT,'snapshot')        
    file_path1= os.path.join(FILE_ROOT,flname)     
    file_path2= os.path.join(path1,flname) 
    print('move path :',file_path1,file_path2,file=fd,flush=True)
    if not os.path.exists(path1):
        os.makedirs(path1)    
    if os.path.exists(file_path2):
        os.remove(file_path2)
    
    shutil.move(file_path1, file_path2)
    print('move file success :',file_path2,file=fd,flush=True)
     
    return file_path2


def rd_snapshot(request):
    sLastupdate = datetime.today().strftime('%Y-%m-%d')
    print("Executing rd snapshot",sLastupdate,file=fd,flush=True)
    
    
    cmd = """
            SELECT
            Releases, RCR, Description, State, TechnicalAreas, ProjectNumber, ProjectDescription, ProjectState,
            Businessline,ProductDomain,Type,Site,Activity,Competence,RCRCategories,Siteallocation,Effortjira,
            SumAllocation,Phase,
            Year,Jans,Febs,Mars,Aprs,Mays,Juns,Juls,Augs,Seps,Octs,Novs,Decs,
            Year1,Year2,Year3,a.ID,b.ID,SubTasks,b.Modifier,b.RecordTime, ProjectNumber2
            FROM
            cdb_rd_resource a
            JOIN
              cdb_rd_effort b
            WHERE  
              a.ID = Right(b.ID,8) 
              %s
            """
      
    sRule = 'ORDER BY Releases, RCR, Year'
    
    SQLConn = analyzer_db()
    # print('SQL=', sRule,sYear,sCompetence,sBL,file=fa,flush=True )
    
    SQLConn.cur.execute(cmd % sRule)
    SQLResult = SQLConn.cur.fetchall()
    SQLConn.close()
    dAve = {}    
    dAve['items'] = [] 
    wb = Workbook()
    ws = wb.active
    ws.title = "data"																

    ws['A1'] = 'Release/Categories'
    ws['B1'] = 'RCR/Activities'
    ws['C1'] = 'Year'
    ws['D1'] = 'Country'
    ws['E1'] = 'Description'
    ws['F1'] = 'State'
    ws['G1'] = 'Project Number'
    ws['H1'] = 'Project Description'
    ws['I1'] = 'Businessline'
    ws['J1'] = 'Resource allocation pool'
    ws['K1'] = 'Technical Areas'
    ws['L1'] = 'Type'
    ws['M1'] = 'Competence'
    ws['N1'] = 'Categories'
    ws['O1'] = 'Phase'
    ws['P1'] = 'SubTask ID'
    ws['Q1'] = 'Effort JIRA Subtask'
    ws['R1'] = 'Effort JIRA RCR'
    ws['S1'] = 'Available'
    ws['T1'] = 'Summary'
    ws['U1'] = 'Year average'
    ws['V1'] = 'Jan'
    ws['W1'] = 'Feb'
    ws['X1'] = 'Mar'
    ws['Y1'] = 'Apr'
    ws['Z1'] = 'May'
    ws['AA1'] = 'Jun'
    ws['AB1'] = 'Jul'
    ws['AC1'] = 'Aug'
    ws['AD1'] = 'Sep'
    ws['AE1'] = 'Oct'
    ws['AF1'] = 'Nov'
    ws['AG1'] = 'Dec'
    ws['AH1'] = 'Year1'
    ws['AI1'] = 'Year2'
    ws['AJ1'] = 'Year3'
    ws['AK1'] = 'Modifier'
    ws['AL1'] = 'Last Modified'
    ws['AM1'] = 'ID'
    ws['AN1'] = 'YID'
    
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['L1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['M1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['N1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['P1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['Q1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['R1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['S1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['T1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['U1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['V1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['W1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['X1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['Y1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['Z1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AA1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AB1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AC1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AD1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AE1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AF1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AG1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AH1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AI1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AJ1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AK1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AL1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AM1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AN1'].alignment = Alignment(horizontal='center', vertical='center')
    
    k=0
    for row in SQLResult:       
        Yave = 0        
        if row[20] !='':
           ave2 = float(row[20])
           Yave = Yave + ave2
        if row[21] !='':            
           ave3 = float(row[21])           
           Yave = Yave + ave3
        if row[22] !='':
           ave4 = float(row[22])
           Yave = Yave + ave4
        if row[23] !='':
           ave5 = float(row[23])
           Yave = Yave + ave5
        if row[24] !='':
           ave6 = float(row[24]) 
           Yave = Yave + ave6
        if row[25] !='':
           ave7 = float(row[25])
           Yave = Yave + ave7
        if row[26] !='':
           ave8 = float(row[26])
           Yave = Yave + ave8
        if row[27] !='':
           ave9 = float(row[27])
           Yave = Yave + ave9
        if row[28] !='':
           ave10 = float(row[28])
           Yave = Yave + ave10
        if row[29] !='':
           ave11 = float(row[29])
           Yave = Yave + ave11
        if row[30] !='':
           ave12 = float(row[30])
           Yave = Yave + ave12
        if row[31] !='':
           ave1 = float(row[31])
           Yave = Yave + ave1
        Yave = Yave/12
        RCR = row[1]
        Competence = row[13]
        TA = row[9]
        ID  = row[35]
        YID = row[36]
        Year = row[19]
        Year1 = row[32]
        Year2 = row[33]
        Year3 = row[34]
        flag = 0        
        for item in dAve['items']:
            if RCR == item['RCR'] and TA == item['TA'] and Competence == item['Competence'] and (Year == item['Year1'] or Year == item['Year2'] or Year == item['Year3']):
                flag = 1
                if YID not in item['YID']:                   
                   ave = item['Yave']
                   item['Yave'] = ave + Yave
                   item['YID'].append(YID)
                   
        if flag == 0:            
            dR = {}
            dR['YID'] = []           
            dR['RCR'] = RCR
            dR['TA'] = TA
            dR['Competence'] = Competence
            dR['Year1'] = Year1
            dR['Year2'] = Year2
            dR['Year3'] = Year3
            dR['Yave'] = Yave            
            dR['YID'].append(YID)
            dAve['items'].append(dR)
            
    for row in SQLResult:
        Yave = 0        
        if row[20] !='':
           ave2 = float(row[20])
           Yave = Yave + ave2
        if row[21] !='':
           ave3 = float(row[21])
           Yave = Yave + ave3
        if row[22] !='':
           ave4 = float(row[22])
           Yave = Yave + ave4
        if row[23] !='':
           ave5 = float(row[23])
           Yave = Yave + ave5
        if row[24] !='':
           ave6 = float(row[24]) 
           Yave = Yave + ave6
        if row[25] !='':
           ave7 = float(row[25])
           Yave = Yave + ave7
        if row[26] !='':
           ave8 = float(row[26])
           Yave = Yave + ave8
        if row[27] !='':
           ave9 = float(row[27])
           Yave = Yave + ave9
        if row[28] !='':
           ave10 = float(row[28])
           Yave = Yave + ave10
        if row[29] !='':
           ave11 = float(row[29])
           Yave = Yave + ave11
        if row[30] !='':
           ave12 = float(row[30])
           Yave = Yave + ave12
        if row[31] !='':
           ave1 = float(row[31])
           Yave = Yave + ave1
        Yave = Yave/12
        RCR = row[1]
        TA = row[9]
        Competence =row[13]
        Year = row[19]
        Year1 = row[32]
        Year2 = row[33]
        Year3 = row[34]
        Sumave = 0
        for item in dAve['items']:
            if RCR == item['RCR'] and TA == item['TA'] and Competence == item['Competence'] and (Year == item['Year1'] or Year == item['Year2'] or Year == item['Year3']):
               Sumave = item['Yave']
        Available = 0
        if row[15] != '' and row[15] !=',' :
            Available = round(float(row[15])-Sumave,2)
        elif row[16] !='' and row[16] !=',' :
            try:
                Available = round(float(row[16])-Sumave,2)
            except:
                # print('Yave err =',RCR, row[16], file=fa, flush=True )
                pass
        elif Yave >0 :
            Available = 0 - Sumave
                
        k+=1
        ws.append([row[0],row[1],row[19],row[11],row[2],row[3],row[5],row[6],row[8],row[9],row[4],row[10],row[13],row[14],row[18],row[37],row[15][:5],row[16][:5],str("%.2f" % (Available)),str("%.2f" % (Sumave)),str("%.2f" % (Yave)),row[20],row[21],row[22],row[23],row[24],row[25],row[26],row[27],row[28],row[29],row[30],row[31],row[32],row[33],row[34],row[38],row[39],row[35],row[36]])
                
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15
    ws.column_dimensions['Q'].width = 15
    ws.column_dimensions['R'].width = 15
    ws.column_dimensions['S'].width = 15
    ws.column_dimensions['T'].width = 15
    ws.column_dimensions['U'].width = 15
    ws.column_dimensions['V'].width = 15
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 15
    ws.column_dimensions['Z'].width = 15
    ws.column_dimensions['AA'].width = 15
    ws.column_dimensions['AB'].width = 15
    ws.column_dimensions['AC'].width = 15
    ws.column_dimensions['AD'].width = 15
    ws.column_dimensions['AE'].width = 15
    ws.column_dimensions['AF'].width = 15
    ws.column_dimensions['AG'].width = 15
    ws.column_dimensions['AH'].width = 15
    ws.column_dimensions['AI'].width = 15
    ws.column_dimensions['AJ'].width = 15
    ws.column_dimensions['AK'].width = 15
    ws.column_dimensions['AL'].width = 15
    ws.column_dimensions['AM'].width = 15
    ws.column_dimensions['AN'].width = 15
    
    filename1 = 'RD_resource-('+ str(k) +')-'+ sLastupdate +'.xlsx'
    wb.save(filename1)
    dResult = {}
    dResult['code'] = 20000
    dResult['data'] = {}
    dResult['data']['status'] = []
    file_path= move_file(filename1)
    NKmail1 ='frank.xiao@nokia-sbell.com'    
    NKmail2 ='dandan.yu@nokia-sbell.com'
    NKmail_list =[]    
    NKmail_list.append(NKmail1)
    NKmail_list.append(NKmail2)
    send_mail(NKmail_list,filename1,file_path)
    dResult['data']['status']="RD snapshot successful" 
    
    return HttpResponse(simplejson.dumps(dResult), content_type='application/json')
